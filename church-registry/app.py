"""
교회 교적 관리 시스템 (Church Registry)
AI 채팅 기반 교적 관리
"""
import os
import json
import base64
import requests
from datetime import datetime, date, timedelta
try:
    from zoneinfo import ZoneInfo
except ImportError:
    from backports.zoneinfo import ZoneInfo
from flask import Flask, render_template, redirect, url_for, request, flash, jsonify, send_file
from flask_sqlalchemy import SQLAlchemy
from dotenv import load_dotenv
from openai import OpenAI
from werkzeug.utils import secure_filename
import cloudinary
import cloudinary.uploader

load_dotenv()

# 서울 시간대 (KST)
SEOUL_TZ = ZoneInfo('Asia/Seoul')

def get_seoul_now():
    """서울 시간대 기준 현재 시각"""
    return datetime.now(SEOUL_TZ)

def get_seoul_today():
    """서울 시간대 기준 오늘 날짜"""
    return datetime.now(SEOUL_TZ).date()

# OpenAI 클라이언트 (API 키가 있을 때만 초기화)
openai_api_key = os.getenv('OPENAI_API_KEY')
openai_client = OpenAI(api_key=openai_api_key) if openai_api_key else None

# Cloudinary 설정 (환경변수가 있을 때만 초기화)
cloudinary_configured = False
if os.getenv('CLOUDINARY_CLOUD_NAME'):
    cloudinary.config(
        cloud_name=os.getenv('CLOUDINARY_CLOUD_NAME'),
        api_key=os.getenv('CLOUDINARY_API_KEY'),
        api_secret=os.getenv('CLOUDINARY_API_SECRET'),
        secure=True
    )
    cloudinary_configured = True

# Flask 앱 초기화
app = Flask(__name__)
app.config['SECRET_KEY'] = os.getenv('SECRET_KEY', 'dev-secret-key-change-in-production')

# DATABASE_URL 처리 (Render PostgreSQL용)
database_url = os.getenv('DATABASE_URL', 'sqlite:///church.db')
if database_url.startswith('postgres://'):
    database_url = database_url.replace('postgres://', 'postgresql://', 1)
app.config['SQLALCHEMY_DATABASE_URI'] = database_url
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# 사진 업로드 설정
UPLOAD_FOLDER = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static', 'uploads')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB

# 데이터베이스 초기화
db = SQLAlchemy(app)

# =============================================================================
# 모델 임포트 (나중에 models/ 폴더로 분리)
# =============================================================================

class Member(db.Model):
    """교인 모델"""
    __tablename__ = 'members'

    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)  # 이름
    phone = db.Column(db.String(20))  # 전화번호
    email = db.Column(db.String(120))  # 이메일
    address = db.Column(db.String(200))  # 주소
    birth_date = db.Column(db.Date)  # 생년월일
    gender = db.Column(db.String(10))  # 성별

    # 교회 관련 정보
    baptism_date = db.Column(db.Date)  # 세례일
    registration_date = db.Column(db.Date)  # 등록일
    registration_number = db.Column(db.String(20))  # 등록번호 (2025-43 형식)
    group_id = db.Column(db.Integer, db.ForeignKey('groups.id'))  # 셀/구역/목장

    # 이전 교회 정보
    previous_church = db.Column(db.String(100))  # 이전 교회명
    previous_church_address = db.Column(db.String(200))  # 이전 교회 주소

    # 교회 조직 정보 (god4u 매핑)
    district = db.Column(db.String(50))  # 교구 (range: "3교구[2025]")
    section = db.Column(db.String(50))  # 구역 (range1: "18구역")
    cell_group = db.Column(db.String(50))  # 속회 (range2: "서울1속")
    mission_group = db.Column(db.String(50))  # 선교회 (range3)
    barnabas = db.Column(db.String(50))  # 바나바 (담당 장로/권사)
    referrer = db.Column(db.String(50))  # 인도자/관계

    # 신급 (학습, 세례, 유아세례, 입교)
    faith_level = db.Column(db.String(20))

    # 가족 관계
    family_id = db.Column(db.Integer, db.ForeignKey('families.id'))
    family_role = db.Column(db.String(20))  # 가장, 배우자, 자녀 등

    # 상태 (active: 활동, inactive: 비활동, newcomer: 새신자, deceased: 별세, transferred: 타교회)
    status = db.Column(db.String(20), default='active')
    member_status = db.Column(db.String(30))  # god4u state 필드 (재적, 별세, 타교회 등)
    deceased_date = db.Column(db.Date)  # 별세일자
    transferred_date = db.Column(db.Date)  # 타교회 이적일자
    notes = db.Column(db.Text)  # 메모

    # 성도 구분 (god4u 매핑)
    member_type = db.Column(db.String(20))  # 직분 (cvname: 권사, 집사 등)
    position_detail = db.Column(db.String(30))  # 상세 직분 (cvname1: 시무권사, 은퇴권사 등)
    department = db.Column(db.String(20))   # 교회학교: 유아부, 유치부, 아동부, 청소년부, 청년부 (장년은 null)
    age_group = db.Column(db.String(20))  # 연령대 (state1: 장년, 청년 등)
    attendance_status = db.Column(db.String(20))  # 출석 상태 (state3: 예배출석, 장기결석 등)

    # 생년월일 추가 정보
    birth_lunar = db.Column(db.Boolean, default=False)  # 음력 생일 여부 (solar: 음/양)

    # 마지막 심방일 (god4u lastvisitday)
    last_visit_date = db.Column(db.Date)  # 마지막 심방일

    # 추가 연락처
    tel = db.Column(db.String(20))  # 집 전화번호
    zipcode = db.Column(db.String(20))  # 우편번호

    # 직업
    occupation = db.Column(db.String(100))  # 직업 (occu)

    # 차량번호 (god4u carnum)
    car_number = db.Column(db.String(20))  # 차량번호

    # 가족 정보 (god4u ran1 - 텍스트 형태의 가족 명단)
    family_members = db.Column(db.String(200))  # 가족 (예: "김진일 이효연 김예원 김주원")

    # 사진
    photo_url = db.Column(db.String(500))  # 프로필 사진 URL

    # 외부 시스템 연동
    external_id = db.Column(db.String(50))  # 외부 시스템 ID (god4u 교적번호 등)

    # 배우자 연결
    partner_id = db.Column(db.Integer)  # 배우자 교인 ID (god4u partnerid)

    created_at = db.Column(db.DateTime, default=get_seoul_now)
    updated_at = db.Column(db.DateTime, default=get_seoul_now, onupdate=get_seoul_now)

    @property
    def age(self):
        """나이 자동 계산 (만 나이)"""
        if not self.birth_date:
            return None
        today = get_seoul_today()
        age = today.year - self.birth_date.year
        # 생일이 아직 안 지났으면 1살 빼기
        if (today.month, today.day) < (self.birth_date.month, self.birth_date.day):
            age -= 1
        return age

    @property
    def korean_age(self):
        """한국 나이 (세는 나이) - 새해에 나이가 바뀜"""
        if not self.birth_date:
            return None
        today = get_seoul_today()
        return today.year - self.birth_date.year + 1

    @property
    def church_school_department(self):
        """교회학교 부서 자동 분류 (한국 나이 기준)

        새해가 되면 자동으로 진급됨:
        - 영아부: 1-3세 (0-2세 만 나이)
        - 유치부: 4-7세 (3-6세 만 나이)
        - 유년부: 8-10세 (초등 1-3학년)
        - 초등부: 11-13세 (초등 4-6학년)
        - 중등부: 14-16세 (중학교)
        - 고등부: 17-19세 (고등학교)
        - 청년부: 20-39세
        - 장년부: 40세 이상
        """
        k_age = self.korean_age
        if k_age is None:
            return None

        if k_age <= 3:
            return '영아부'
        elif k_age <= 7:
            return '유치부'
        elif k_age <= 10:
            return '유년부'
        elif k_age <= 13:
            return '초등부'
        elif k_age <= 16:
            return '중등부'
        elif k_age <= 19:
            return '고등부'
        elif k_age <= 39:
            return '청년부'
        else:
            return '장년부'

    @property
    def is_newcomer(self):
        """새신자 여부 (등록 후 2년 이내, 한국 시간 기준)"""
        if not self.registration_date:
            return False
        two_years_ago = get_seoul_today() - timedelta(days=730)  # 2년
        return self.registration_date > two_years_ago

    @property
    def status_display(self):
        """상태 표시 텍스트"""
        # 별세자와 타교회는 우선 표시
        if self.status == 'deceased' or self.member_status in ['별세', '소천']:
            return '별세'
        if self.status == 'transferred' or self.member_status in ['타교회', '타교인', '이명']:
            return '타교회'
        # 새신자 체크 (등록 2년 이내)
        if self.is_newcomer:
            return '새신자'
        # 일반 상태
        if self.status == 'active':
            return '활동'
        if self.status == 'inactive':
            return '비활동'
        return self.status or '활동'

    @property
    def status_display_with_date(self):
        """상태 표시 텍스트 (날짜 포함)"""
        # 별세자
        if self.status == 'deceased' or self.member_status in ['별세', '소천']:
            if self.deceased_date:
                return f'별세 ({self.deceased_date.strftime("%Y.%m.%d")})'
            return '별세'
        # 타교회
        if self.status == 'transferred' or self.member_status in ['타교회', '타교인', '이명']:
            if self.transferred_date:
                return f'타교회 ({self.transferred_date.strftime("%Y.%m.%d")})'
            return '타교회'
        # 새신자 체크 (등록 2년 이내)
        if self.is_newcomer:
            return '새신자'
        # 일반 상태
        if self.status == 'active':
            return '활동'
        if self.status == 'inactive':
            return '비활동'
        return self.status or '활동'

    @property
    def status_class(self):
        """CSS 클래스용 상태"""
        if self.status == 'deceased' or self.member_status in ['별세', '소천']:
            return 'deceased'
        if self.status == 'transferred' or self.member_status in ['타교회', '타교인', '이명']:
            return 'transferred'
        if self.is_newcomer:
            return 'newcomer'
        return self.status or 'active'

    @property
    def display_name(self):
        """표시용 이름 (이명섭 집사(새가족) 형식)"""
        name = self.name
        if self.member_type:
            name = f"{self.name} {self.member_type}"
        if self.is_newcomer:
            name = f"{name}(새가족)"
        return name

    def __repr__(self):
        return f'<Member {self.name}>'

    # ===== 가족 관계 헬퍼 메서드 =====

    def get_family_relationships(self):
        """모든 가족 관계 반환"""
        return FamilyRelationship.query.filter_by(member_id=self.id).all()

    def get_spouse(self):
        """배우자 반환"""
        rel = FamilyRelationship.query.filter_by(
            member_id=self.id,
            relationship_type='spouse'
        ).first()
        return rel.related_member if rel else None

    def get_parents(self):
        """부모 목록 반환 (나의 부모 = 나를 자녀로 둔 사람)"""
        rels = FamilyRelationship.query.filter_by(
            member_id=self.id,
            relationship_type='child'  # 내가 자녀인 관계
        ).all()
        return [rel.related_member for rel in rels]

    def get_children(self):
        """자녀 목록 반환 (나의 자녀 = 나를 부모로 둔 사람)"""
        rels = FamilyRelationship.query.filter_by(
            member_id=self.id,
            relationship_type='parent'  # 내가 부모인 관계
        ).all()
        return [rel.related_member for rel in rels]

    def get_siblings(self):
        """형제자매 목록 반환"""
        rels = FamilyRelationship.query.filter_by(
            member_id=self.id,
            relationship_type='sibling'
        ).all()
        return [rel.related_member for rel in rels]

    def get_family_tree(self):
        """전체 가족 트리 반환 (딕셔너리 형태)"""
        return {
            'spouse': self.get_spouse(),
            'parents': self.get_parents(),
            'children': self.get_children(),
            'siblings': self.get_siblings(),
        }

    def get_extended_family(self):
        """대가족 전체 목록 반환 (중복 제거)"""
        family_members = set()

        # 직계 가족 추가
        spouse = self.get_spouse()
        if spouse:
            family_members.add(spouse)

        for parent in self.get_parents():
            family_members.add(parent)
            # 부모의 배우자 (다른 부모)
            parent_spouse = parent.get_spouse()
            if parent_spouse:
                family_members.add(parent_spouse)
            # 부모의 자녀 (형제자매)
            for sibling in parent.get_children():
                if sibling.id != self.id:
                    family_members.add(sibling)

        for child in self.get_children():
            family_members.add(child)
            # 자녀의 배우자
            child_spouse = child.get_spouse()
            if child_spouse:
                family_members.add(child_spouse)
            # 손주
            for grandchild in child.get_children():
                family_members.add(grandchild)

        for sibling in self.get_siblings():
            family_members.add(sibling)

        return list(family_members)


class Family(db.Model):
    """가족 모델 - 대가족 단위 그룹"""
    __tablename__ = 'families'

    id = db.Column(db.Integer, primary_key=True)
    family_name = db.Column(db.String(100))  # 가족명 (예: 홍길동 가정)
    members = db.relationship('Member', backref='family', lazy=True)

    created_at = db.Column(db.DateTime, default=get_seoul_now)

    def get_all_members(self):
        """가족 구성원 전체 반환"""
        return Member.query.filter_by(family_id=self.id).all()


class FamilyRelationship(db.Model):
    """가족 관계 모델 - 개별 관계 정의 (한국 가족 관계 전체 지원)"""
    __tablename__ = 'family_relationships'

    id = db.Column(db.Integer, primary_key=True)
    member_id = db.Column(db.Integer, db.ForeignKey('members.id'), nullable=False)  # 기준 교인
    related_member_id = db.Column(db.Integer, db.ForeignKey('members.id'), nullable=False)  # 관계 대상

    # 관계 유형: spouse, parent, child, sibling, in_law, grandparent, grandchild, extended
    relationship_type = db.Column(db.String(20), nullable=False)

    # 상세 관계: 아버지, 어머니, 시어머니, 장인, 동서, 시누이, 조카 등
    relationship_detail = db.Column(db.String(30))

    created_at = db.Column(db.DateTime, default=get_seoul_now)

    # 관계 정의
    member = db.relationship('Member', foreign_keys=[member_id], backref='relationships_as_member')
    related_member = db.relationship('Member', foreign_keys=[related_member_id], backref='relationships_as_related')

    # 유니크 제약 (같은 관계 중복 방지)
    __table_args__ = (
        db.UniqueConstraint('member_id', 'related_member_id', 'relationship_type', name='unique_relationship'),
    )

    # 관계 유형 상수
    SPOUSE = 'spouse'           # 배우자
    PARENT = 'parent'           # 부모
    CHILD = 'child'             # 자녀
    SIBLING = 'sibling'         # 형제자매
    IN_LAW = 'in_law'           # 인척 (시댁/처가/사돈)
    GRANDPARENT = 'grandparent' # 조부모
    GRANDCHILD = 'grandchild'   # 손자녀
    EXTENDED = 'extended'       # 확대가족 (삼촌, 고모, 조카 등)

    # 관계 역방향 매핑
    REVERSE_RELATIONSHIPS = {
        'spouse': 'spouse',
        'parent': 'child',
        'child': 'parent',
        'sibling': 'sibling',
        'in_law': 'in_law',
        'grandparent': 'grandchild',
        'grandchild': 'grandparent',
        'extended': 'extended',
    }

    # 한국 가족 관계 상세 목록 (relationship_detail에 사용)
    KOREAN_RELATIONSHIPS = {
        # 배우자
        'spouse': ['남편', '아내', '배우자'],

        # 부모
        'parent': ['아버지', '어머니', '부모'],

        # 자녀
        'child': ['아들', '딸', '자녀'],

        # 형제자매
        'sibling': ['형', '오빠', '누나', '언니', '남동생', '여동생', '동생', '형제', '자매', '남매'],

        # 인척 - 시댁
        'in_law_husband_side': [
            '시아버지', '시어머니', '시부모',  # 배우자의 부모
            '시형', '시동생', '시누이', '시숙', '형님', '아주버님',  # 배우자의 형제자매
            '시조부', '시조모',  # 배우자의 조부모
        ],

        # 인척 - 처가
        'in_law_wife_side': [
            '장인', '장모', '처부모',  # 배우자의 부모
            '처형', '처남', '처제', '처남댁', '형님',  # 배우자의 형제자매
        ],

        # 인척 - 형제자매의 배우자
        'in_law_sibling_spouse': [
            '형수', '제수', '올케',  # 형제의 아내
            '매형', '매부', '제부', '형부',  # 자매의 남편
            '동서',  # 형제자매의 배우자끼리
        ],

        # 인척 - 자녀의 배우자
        'in_law_child_spouse': [
            '며느리', '사위',  # 자녀의 배우자
            '손자며느리', '손녀사위',  # 손자녀의 배우자
        ],

        # 인척 - 사돈
        'in_law_saddon': [
            '사돈', '바깥사돈', '안사돈', '사돈어른',
        ],

        # 조부모
        'grandparent': ['할아버지', '할머니', '조부', '조모', '외할아버지', '외할머니', '외조부', '외조모'],

        # 손자녀
        'grandchild': ['손자', '손녀', '외손자', '외손녀'],

        # 확대가족 - 삼촌/고모/이모
        'extended_parent_sibling': [
            '삼촌', '큰아버지', '작은아버지', '막내삼촌',
            '고모', '큰고모', '작은고모',
            '외삼촌', '큰외삼촌', '작은외삼촌',
            '이모', '큰이모', '작은이모',
            '숙부', '숙모', '고모부', '이모부',
        ],

        # 확대가족 - 조카
        'extended_niece_nephew': [
            '조카', '조카딸', '조카아들', '생질', '생질녀',
            '외조카', '이종조카', '고종조카',
        ],

        # 확대가족 - 사촌
        'extended_cousin': [
            '사촌', '사촌형', '사촌누나', '사촌오빠', '사촌언니', '사촌동생',
            '고종사촌', '이종사촌', '외사촌',
        ],
    }

    # 상세 관계 역방향 매핑 (남성↔여성 기준)
    REVERSE_DETAILS = {
        # 배우자
        '남편': '아내', '아내': '남편',

        # 부모↔자녀
        '아버지': '아들', '어머니': '아들',  # 아들 입장
        '아버지': '딸', '어머니': '딸',      # 딸 입장

        # 형제자매 (성별에 따라 다름)
        '형': '동생', '오빠': '동생', '누나': '동생', '언니': '동생',
        '남동생': '형', '여동생': '언니',

        # 조부모↔손자녀
        '할아버지': '손자', '할머니': '손자',
        '외할아버지': '외손자', '외할머니': '외손자',

        # 시댁 관계
        '시아버지': '며느리', '시어머니': '며느리',
        '며느리': '시어머니',

        # 처가 관계
        '장인': '사위', '장모': '사위',
        '사위': '장인',

        # 형제자매의 배우자
        '형수': '시동생', '제수': '시형',
        '매형': '처제', '매부': '처형',
        '동서': '동서',

        # 삼촌/조카
        '삼촌': '조카', '고모': '조카', '이모': '조카', '외삼촌': '조카',
        '조카': '삼촌',
    }

    @classmethod
    def create_bidirectional(cls, member_id, related_member_id, relationship_type, detail=None, reverse_detail=None):
        """양방향 관계 생성"""
        # 정방향 관계
        rel1 = cls(
            member_id=member_id,
            related_member_id=related_member_id,
            relationship_type=relationship_type,
            relationship_detail=detail
        )

        # 역방향 관계
        reverse_type = cls.REVERSE_RELATIONSHIPS.get(relationship_type, relationship_type)
        rel2 = cls(
            member_id=related_member_id,
            related_member_id=member_id,
            relationship_type=reverse_type,
            relationship_detail=reverse_detail or cls.REVERSE_DETAILS.get(detail)
        )

        return [rel1, rel2]

    @classmethod
    def get_all_relationship_options(cls):
        """UI용 관계 선택 옵션 반환"""
        options = []

        # 직계가족
        options.append({'group': '배우자', 'items': cls.KOREAN_RELATIONSHIPS['spouse']})
        options.append({'group': '부모', 'items': cls.KOREAN_RELATIONSHIPS['parent']})
        options.append({'group': '자녀', 'items': cls.KOREAN_RELATIONSHIPS['child']})
        options.append({'group': '형제자매', 'items': cls.KOREAN_RELATIONSHIPS['sibling']})

        # 조부모/손자녀
        options.append({'group': '조부모', 'items': cls.KOREAN_RELATIONSHIPS['grandparent']})
        options.append({'group': '손자녀', 'items': cls.KOREAN_RELATIONSHIPS['grandchild']})

        # 인척
        options.append({'group': '시댁', 'items': cls.KOREAN_RELATIONSHIPS['in_law_husband_side']})
        options.append({'group': '처가', 'items': cls.KOREAN_RELATIONSHIPS['in_law_wife_side']})
        options.append({'group': '형제자매의 배우자', 'items': cls.KOREAN_RELATIONSHIPS['in_law_sibling_spouse']})
        options.append({'group': '자녀의 배우자', 'items': cls.KOREAN_RELATIONSHIPS['in_law_child_spouse']})
        options.append({'group': '사돈', 'items': cls.KOREAN_RELATIONSHIPS['in_law_saddon']})

        # 확대가족
        options.append({'group': '삼촌/고모/이모', 'items': cls.KOREAN_RELATIONSHIPS['extended_parent_sibling']})
        options.append({'group': '조카', 'items': cls.KOREAN_RELATIONSHIPS['extended_niece_nephew']})
        options.append({'group': '사촌', 'items': cls.KOREAN_RELATIONSHIPS['extended_cousin']})

        return options

    @classmethod
    def detail_to_type(cls, detail):
        """상세 관계에서 관계 유형 추론"""
        if detail in cls.KOREAN_RELATIONSHIPS['spouse']:
            return 'spouse'
        elif detail in cls.KOREAN_RELATIONSHIPS['parent']:
            return 'parent'
        elif detail in cls.KOREAN_RELATIONSHIPS['child']:
            return 'child'
        elif detail in cls.KOREAN_RELATIONSHIPS['sibling']:
            return 'sibling'
        elif detail in cls.KOREAN_RELATIONSHIPS['grandparent']:
            return 'grandparent'
        elif detail in cls.KOREAN_RELATIONSHIPS['grandchild']:
            return 'grandchild'
        elif any(detail in cls.KOREAN_RELATIONSHIPS[k] for k in ['in_law_husband_side', 'in_law_wife_side',
                 'in_law_sibling_spouse', 'in_law_child_spouse', 'in_law_saddon']):
            return 'in_law'
        elif any(detail in cls.KOREAN_RELATIONSHIPS[k] for k in ['extended_parent_sibling',
                 'extended_niece_nephew', 'extended_cousin']):
            return 'extended'
        return 'extended'  # 기본값

    def get_display_text(self):
        """표시용 텍스트 (예: 배우자, 아버지, 딸 등)"""
        if self.relationship_detail:
            return self.relationship_detail

        type_names = {
            'spouse': '배우자',
            'parent': '부모',
            'child': '자녀',
            'sibling': '형제자매',
            'in_law': '인척',
            'grandparent': '조부모',
            'grandchild': '손자녀',
            'extended': '친척',
        }
        return type_names.get(self.relationship_type, self.relationship_type)


class Group(db.Model):
    """그룹 모델 - 계층 구조 지원 (교구/선교회/직분/교회학교/새가족)"""
    __tablename__ = 'groups'

    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)  # 그룹명
    group_type = db.Column(db.String(50))  # district, mission, position, school, newcomer
    leader_id = db.Column(db.Integer)  # 리더 교인 ID
    parent_id = db.Column(db.Integer, db.ForeignKey('groups.id'), nullable=True)  # 상위 그룹
    level = db.Column(db.Integer, default=0)  # 계층 레벨 (0: 최상위)
    description = db.Column(db.String(200))  # 설명

    # 자기 참조 관계
    parent = db.relationship('Group', remote_side=[id], backref='children')
    members = db.relationship('Member', backref='group', lazy=True)

    created_at = db.Column(db.DateTime, default=get_seoul_now)

    def get_full_path(self):
        """전체 경로 반환 (예: 교구 > 1교구 > 1구역)"""
        if self.parent:
            return f"{self.parent.get_full_path()} > {self.name}"
        return self.name

    def get_all_children(self):
        """모든 하위 그룹 반환 (재귀)"""
        result = list(self.children)
        for child in self.children:
            result.extend(child.get_all_children())
        return result

    def get_member_count(self, include_children=True):
        """소속 인원 수 (하위 그룹 포함 옵션)"""
        count = len(self.members)
        if include_children:
            for child in self.get_all_children():
                count += len(child.members)
        return count


class Attendance(db.Model):
    """출석 기록 모델"""
    __tablename__ = 'attendance'

    id = db.Column(db.Integer, primary_key=True)
    member_id = db.Column(db.Integer, db.ForeignKey('members.id'), nullable=False)
    date = db.Column(db.Date, nullable=False)  # 출석 날짜
    service_type = db.Column(db.String(50))  # 주일예배, 수요예배, 금요기도회 등
    attended = db.Column(db.Boolean, default=True)

    member = db.relationship('Member', backref='attendance_records')

    created_at = db.Column(db.DateTime, default=get_seoul_now)


class Visit(db.Model):
    """심방 기록 모델"""
    __tablename__ = 'visits'

    id = db.Column(db.Integer, primary_key=True)
    member_id = db.Column(db.Integer, db.ForeignKey('members.id'), nullable=False)
    visit_date = db.Column(db.Date, nullable=False)
    visitor_name = db.Column(db.String(100))  # 심방자
    purpose = db.Column(db.String(100))  # 심방 목적
    notes = db.Column(db.Text)  # 심방 내용

    member = db.relationship('Member', backref='visit_records')

    created_at = db.Column(db.DateTime, default=get_seoul_now)


class Offering(db.Model):
    """헌금 기록 모델"""
    __tablename__ = 'offerings'

    id = db.Column(db.Integer, primary_key=True)
    member_id = db.Column(db.Integer, db.ForeignKey('members.id'))
    date = db.Column(db.Date, nullable=False)
    amount = db.Column(db.Integer, nullable=False)  # 금액 (원)
    offering_type = db.Column(db.String(50))  # 십일조, 감사헌금, 건축헌금 등
    notes = db.Column(db.Text)

    member = db.relationship('Member', backref='offering_records')

    created_at = db.Column(db.DateTime, default=get_seoul_now)


# =============================================================================
# 라우트
# =============================================================================

@app.route('/')
def index():
    """메인 페이지"""
    return render_template('index.html')


@app.route('/members')
def member_list():
    """교인 목록"""
    import re

    # 검색 파라미터
    query = request.args.get('q', '')
    cat1 = request.args.get('cat1', '')  # 1단계: district, position, reg_status, attendance, mission
    cat2 = request.args.get('cat2', '')  # 2단계: 교구(1교구 등), 직분(장로 등)
    cat3 = request.args.get('cat3', '')  # 3단계: 구역 (교구별에서만)
    cat4 = request.args.get('cat4', '')  # 4단계: 속회 (교구별에서만)

    # 교구 이름에서 기본 교구 추출 (예: "1교구[2012]" → "1교구")
    def extract_base_district(district_str):
        if not district_str:
            return None
        # "1교구", "2교구", "3교구" 패턴 추출
        match = re.match(r'^(\d+교구)', district_str)
        if match:
            return match.group(1)
        # "교회학교" 포함 여부
        if '교회학교' in district_str or '청년교구' in district_str:
            return '교회학교'
        return None

    # 기본 쿼리 (기본적으로 별세/타교인 제외)
    members_query = Member.query

    # 이름 검색
    if query:
        members_query = members_query.filter(Member.name.contains(query))

    # 카테고리 필터링
    if cat1 == 'district':
        # 교구별 필터 (3단계 계층)
        if cat2:
            if cat2 == '미등록':
                members_query = members_query.filter(
                    db.or_(Member.district.is_(None), Member.district == '')
                )
            elif cat2 == '교회학교':
                members_query = members_query.filter(
                    db.or_(
                        Member.district.contains('교회학교'),
                        Member.district.contains('청년교구')
                    )
                )
            else:
                # 1교구, 2교구, 3교구 등
                members_query = members_query.filter(Member.district.contains(cat2))

            # 3단계: 구역 필터
            if cat3:
                members_query = members_query.filter(Member.section.contains(cat3))

                # 4단계: 속회 필터
                if cat4:
                    members_query = members_query.filter(Member.cell_group.contains(cat4))

        # 별세/타교인 제외
        members_query = members_query.filter(Member.status.notin_(['deceased', 'transferred']))

    elif cat1 == 'position':
        # 직분 필터 (장로, 권사, 집사, 성도, 교회학교만)
        if cat2:
            if cat2 == '성도':
                # 성도: 직분이 없거나 '성도'인 경우
                members_query = members_query.filter(
                    db.or_(
                        Member.member_type.is_(None),
                        Member.member_type == '',
                        Member.member_type == '성도'
                    )
                )
            elif cat2 == '교회학교':
                # 교회학교: department가 있거나 교회학교 관련 교구
                members_query = members_query.filter(
                    db.or_(
                        Member.department.isnot(None),
                        Member.district.contains('교회학교'),
                        Member.district.contains('청년교구')
                    )
                )
            else:
                members_query = members_query.filter(Member.member_type == cat2)
        # 별세/타교인 제외
        members_query = members_query.filter(Member.status.notin_(['deceased', 'transferred']))

    elif cat1 == 'reg_status':
        # 등록상태 필터 (교인/별세/타교인)
        if cat2 == 'active_member':
            members_query = members_query.filter(Member.status.notin_(['deceased', 'transferred']))
        elif cat2 == 'deceased':
            members_query = members_query.filter(
                db.or_(Member.status == 'deceased', Member.member_status.in_(['별세', '소천']))
            )
        elif cat2 == 'transferred':
            members_query = members_query.filter(
                db.or_(Member.status == 'transferred', Member.member_status.in_(['타교회', '타교인', '이명']))
            )

    elif cat1 == 'attendance':
        # 출석상태 필터
        if cat2 == 'active':
            members_query = members_query.filter(Member.status == 'active')
        elif cat2 == 'inactive':
            members_query = members_query.filter(Member.status == 'inactive')
        elif cat2 == 'newcomer':
            two_years_ago = get_seoul_today() - timedelta(days=730)
            members_query = members_query.filter(Member.registration_date > two_years_ago)
        # 별세/타교인 제외
        members_query = members_query.filter(Member.status.notin_(['deceased', 'transferred']))

    elif cat1 == 'mission' and cat2:
        # 선교회 필터
        members_query = members_query.filter(Member.mission_group.contains(cat2))
        # 별세/타교인 제외
        members_query = members_query.filter(Member.status.notin_(['deceased', 'transferred']))

    else:
        # 기본: 별세/타교인 제외
        if not query:
            members_query = members_query.filter(Member.status.notin_(['deceased', 'transferred']))

    # 검색어나 필터가 있을 때만 결과 조회 (버퍼링 방지)
    if query or cat1:
        members = members_query.order_by(Member.name).all()
    else:
        members = []  # 검색 전에는 빈 목록

    # 2단계 옵션 데이터 수집
    cat2_options = {}
    cat3_options = {}
    cat4_options = {}

    if cat1 == 'district':
        # 교구 옵션: 1교구, 2교구, 3교구, 교회학교, 미등록만
        cat2_options = {
            '1교구': '1교구',
            '2교구': '2교구',
            '3교구': '3교구',
            '교회학교': '교회학교',
            '미등록': '미등록'
        }

        # 선택된 교구의 구역 목록
        if cat2 and cat2 not in ['미등록', '교회학교']:
            sections = db.session.query(Member.section).filter(
                Member.district.contains(cat2),
                Member.section.isnot(None)
            ).distinct().all()
            cat3_options = {s[0]: s[0] for s in sections if s[0]}

            # 선택된 구역의 속회 목록
            if cat3:
                cells = db.session.query(Member.cell_group).filter(
                    Member.district.contains(cat2),
                    Member.section.contains(cat3),
                    Member.cell_group.isnot(None)
                ).distinct().all()
                cat4_options = {c[0]: c[0] for c in cells if c[0]}

    elif cat1 == 'position':
        # 직분 옵션: 장로, 권사, 집사, 성도, 교회학교만
        cat2_options = {
            '장로': '장로',
            '권사': '권사',
            '집사': '집사',
            '성도': '성도',
            '교회학교': '교회학교'
        }

    elif cat1 == 'reg_status':
        cat2_options = {
            'active_member': '현재 교인',
            'deceased': '별세',
            'transferred': '타교회'
        }

    elif cat1 == 'attendance':
        cat2_options = {
            'active': '활동',
            'inactive': '비활동',
            'newcomer': '새신자'
        }

    elif cat1 == 'mission':
        missions = db.session.query(Member.mission_group).filter(Member.mission_group.isnot(None)).distinct().all()
        cat2_options = {m[0]: m[0] for m in missions if m[0]}

    # 선교회 옵션 (JavaScript에서 동적 로드용)
    all_missions = db.session.query(Member.mission_group).filter(Member.mission_group.isnot(None)).distinct().all()
    mission_options = {m[0]: m[0] for m in all_missions if m[0]}

    return render_template('members/list.html',
                         members=members,
                         query=query,
                         cat1=cat1,
                         cat2=cat2,
                         cat3=cat3,
                         cat4=cat4,
                         cat2_options=cat2_options,
                         cat3_options=cat3_options,
                         cat4_options=cat4_options,
                         mission_options=mission_options)


@app.route('/members/new', methods=['GET', 'POST'])
def member_new():
    """교인 등록"""
    if request.method == 'POST':
        # 폼 데이터 수집 - 기본 정보
        name = request.form.get('name', '').strip()
        phone = request.form.get('phone', '').strip()
        email = request.form.get('email', '').strip()
        address = request.form.get('address', '').strip()
        gender = request.form.get('gender', '')
        status = request.form.get('status', 'active')
        group_id = request.form.get('group_id')
        notes = request.form.get('notes', '').strip()
        member_type = request.form.get('member_type', '')
        department = request.form.get('department', '')
        photo_url = request.form.get('photo_url', '')

        # 새 필드들
        registration_number = request.form.get('registration_number', '').strip()
        previous_church = request.form.get('previous_church', '').strip()
        previous_church_address = request.form.get('previous_church_address', '').strip()
        district = request.form.get('district', '').strip()
        cell_group = request.form.get('cell_group', '').strip()
        mission_group = request.form.get('mission_group', '').strip()
        barnabas = request.form.get('barnabas', '').strip()
        referrer = request.form.get('referrer', '').strip()
        faith_level = request.form.get('faith_level', '').strip()

        # 날짜 처리
        birth_date = None
        if request.form.get('birth_date'):
            birth_date = datetime.strptime(request.form.get('birth_date'), '%Y-%m-%d').date()

        baptism_date = None
        if request.form.get('baptism_date'):
            baptism_date = datetime.strptime(request.form.get('baptism_date'), '%Y-%m-%d').date()

        registration_date = None
        if request.form.get('registration_date'):
            registration_date = datetime.strptime(request.form.get('registration_date'), '%Y-%m-%d').date()
        else:
            registration_date = get_seoul_today()

        # 별세일자 / 타교일자
        deceased_date = None
        if request.form.get('deceased_date'):
            deceased_date = datetime.strptime(request.form.get('deceased_date'), '%Y-%m-%d').date()

        transferred_date = None
        if request.form.get('transferred_date'):
            transferred_date = datetime.strptime(request.form.get('transferred_date'), '%Y-%m-%d').date()

        # 유효성 검사
        if not name:
            flash('이름은 필수 입력 항목입니다.', 'danger')
            return render_template('members/form.html', groups=Group.query.all())

        # 교인 생성
        member = Member(
            name=name,
            phone=phone,
            email=email,
            address=address,
            birth_date=birth_date,
            gender=gender,
            baptism_date=baptism_date,
            registration_date=registration_date,
            registration_number=registration_number if registration_number else None,
            group_id=int(group_id) if group_id else None,
            status=status,
            notes=notes,
            member_type=member_type if member_type else None,
            department=department if department else None,
            photo_url=photo_url if photo_url else None,
            previous_church=previous_church if previous_church else None,
            previous_church_address=previous_church_address if previous_church_address else None,
            district=district if district else None,
            cell_group=cell_group if cell_group else None,
            mission_group=mission_group if mission_group else None,
            barnabas=barnabas if barnabas else None,
            referrer=referrer if referrer else None,
            faith_level=faith_level if faith_level else None,
            deceased_date=deceased_date,
            transferred_date=transferred_date
        )

        db.session.add(member)
        db.session.commit()

        flash(f'{member.display_name}이(가) 등록되었습니다.', 'success')
        return redirect(url_for('member_detail', member_id=member.id))

    groups = Group.query.all()
    return render_template('members/form.html', groups=groups, member=None)


@app.route('/members/<int:member_id>')
def member_detail(member_id):
    """교인 상세 보기"""
    member = Member.query.get_or_404(member_id)
    return render_template('members/detail.html', member=member)


@app.route('/members/<int:member_id>/edit', methods=['GET', 'POST'])
def member_edit(member_id):
    """교인 수정"""
    member = Member.query.get_or_404(member_id)

    if request.method == 'POST':
        # 폼 데이터 수집 - 기본 정보
        member.name = request.form.get('name', '').strip()
        member.phone = request.form.get('phone', '').strip()
        member.email = request.form.get('email', '').strip()
        member.address = request.form.get('address', '').strip()
        member.gender = request.form.get('gender', '')
        member.status = request.form.get('status', 'active')
        member.notes = request.form.get('notes', '').strip()

        group_id = request.form.get('group_id')
        member.group_id = int(group_id) if group_id else None

        # 성도 구분
        member_type = request.form.get('member_type', '')
        member.member_type = member_type if member_type else None
        department = request.form.get('department', '')
        member.department = department if department else None

        # 사진 URL
        photo_url = request.form.get('photo_url', '')
        member.photo_url = photo_url if photo_url else None

        # 새 필드들
        registration_number = request.form.get('registration_number', '').strip()
        member.registration_number = registration_number if registration_number else None

        previous_church = request.form.get('previous_church', '').strip()
        member.previous_church = previous_church if previous_church else None

        previous_church_address = request.form.get('previous_church_address', '').strip()
        member.previous_church_address = previous_church_address if previous_church_address else None

        district = request.form.get('district', '').strip()
        member.district = district if district else None

        cell_group = request.form.get('cell_group', '').strip()
        member.cell_group = cell_group if cell_group else None

        mission_group = request.form.get('mission_group', '').strip()
        member.mission_group = mission_group if mission_group else None

        barnabas = request.form.get('barnabas', '').strip()
        member.barnabas = barnabas if barnabas else None

        referrer = request.form.get('referrer', '').strip()
        member.referrer = referrer if referrer else None

        faith_level = request.form.get('faith_level', '').strip()
        member.faith_level = faith_level if faith_level else None

        # 날짜 처리
        if request.form.get('birth_date'):
            member.birth_date = datetime.strptime(request.form.get('birth_date'), '%Y-%m-%d').date()
        else:
            member.birth_date = None

        if request.form.get('baptism_date'):
            member.baptism_date = datetime.strptime(request.form.get('baptism_date'), '%Y-%m-%d').date()
        else:
            member.baptism_date = None

        if request.form.get('registration_date'):
            member.registration_date = datetime.strptime(request.form.get('registration_date'), '%Y-%m-%d').date()

        # 별세일자 / 타교일자
        if request.form.get('deceased_date'):
            member.deceased_date = datetime.strptime(request.form.get('deceased_date'), '%Y-%m-%d').date()
        else:
            member.deceased_date = None

        if request.form.get('transferred_date'):
            member.transferred_date = datetime.strptime(request.form.get('transferred_date'), '%Y-%m-%d').date()
        else:
            member.transferred_date = None

        # 유효성 검사
        if not member.name:
            flash('이름은 필수 입력 항목입니다.', 'danger')
            return render_template('members/form.html', groups=Group.query.all(), member=member)

        db.session.commit()

        flash(f'{member.display_name} 정보가 수정되었습니다.', 'success')
        return redirect(url_for('member_detail', member_id=member.id))

    groups = Group.query.all()
    return render_template('members/form.html', groups=groups, member=member)


@app.route('/members/<int:member_id>/delete', methods=['POST'])
def member_delete(member_id):
    """교인 삭제"""
    member = Member.query.get_or_404(member_id)
    name = member.name

    db.session.delete(member)
    db.session.commit()

    flash(f'{name} 교인이 삭제되었습니다.', 'success')
    return redirect(url_for('member_list'))


@app.route('/health')
def health():
    """헬스 체크 (Render용)"""
    return {'status': 'ok'}


# =============================================================================
# 그룹 관리 라우트
# =============================================================================

@app.route('/groups')
def group_list():
    """그룹 목록 - 계층 구조로 표시"""
    # 최상위 그룹만 조회 (parent_id가 None인 그룹)
    top_groups = Group.query.filter_by(parent_id=None).order_by(Group.name).all()
    return render_template('groups/list.html', top_groups=top_groups)


@app.route('/groups/new', methods=['GET', 'POST'])
def group_new():
    """그룹 등록"""
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        group_type = request.form.get('group_type', '')
        leader_id = request.form.get('leader_id')
        parent_id = request.form.get('parent_id')
        description = request.form.get('description', '').strip()

        if not name:
            flash('그룹명은 필수 입력 항목입니다.', 'danger')
            members = Member.query.order_by(Member.name).all()
            all_groups = Group.query.order_by(Group.name).all()
            return render_template('groups/form.html', members=members, all_groups=all_groups)

        # 상위 그룹이 있으면 레벨과 타입 결정
        level = 0
        if parent_id:
            parent = Group.query.get(int(parent_id))
            if parent:
                level = parent.level + 1
                if not group_type:
                    group_type = parent.group_type

        group = Group(
            name=name,
            group_type=group_type,
            leader_id=int(leader_id) if leader_id else None,
            parent_id=int(parent_id) if parent_id else None,
            level=level,
            description=description
        )

        db.session.add(group)
        db.session.commit()

        flash(f'{name} 그룹이 생성되었습니다.', 'success')
        return redirect(url_for('group_list'))

    members = Member.query.order_by(Member.name).all()
    all_groups = Group.query.order_by(Group.level, Group.name).all()
    parent_id = request.args.get('parent_id', type=int)
    parent_group = Group.query.get(parent_id) if parent_id else None
    return render_template('groups/form.html', members=members, group=None,
                         all_groups=all_groups, parent_group=parent_group)


@app.route('/groups/<int:group_id>')
def group_detail(group_id):
    """그룹 상세 보기"""
    group = Group.query.get_or_404(group_id)
    return render_template('groups/detail.html', group=group)


@app.route('/groups/<int:group_id>/edit', methods=['GET', 'POST'])
def group_edit(group_id):
    """그룹 수정"""
    group = Group.query.get_or_404(group_id)

    if request.method == 'POST':
        group.name = request.form.get('name', '').strip()
        group.group_type = request.form.get('group_type', '')
        leader_id = request.form.get('leader_id')
        parent_id = request.form.get('parent_id')
        group.description = request.form.get('description', '').strip()
        group.leader_id = int(leader_id) if leader_id else None

        # 상위 그룹 변경
        new_parent_id = int(parent_id) if parent_id else None
        if new_parent_id != group.parent_id:
            group.parent_id = new_parent_id
            if new_parent_id:
                parent = Group.query.get(new_parent_id)
                group.level = parent.level + 1 if parent else 0
            else:
                group.level = 0

        if not group.name:
            flash('그룹명은 필수 입력 항목입니다.', 'danger')
            members = Member.query.order_by(Member.name).all()
            all_groups = Group.query.filter(Group.id != group.id).order_by(Group.name).all()
            return render_template('groups/form.html', members=members, group=group, all_groups=all_groups)

        db.session.commit()

        flash(f'{group.name} 그룹이 수정되었습니다.', 'success')
        return redirect(url_for('group_detail', group_id=group.id))

    members = Member.query.order_by(Member.name).all()
    # 자기 자신과 자식 그룹은 상위 그룹으로 선택 불가
    exclude_ids = [group.id] + [c.id for c in group.get_all_children()]
    all_groups = Group.query.filter(~Group.id.in_(exclude_ids)).order_by(Group.level, Group.name).all()
    return render_template('groups/form.html', members=members, group=group, all_groups=all_groups)


@app.route('/groups/<int:group_id>/delete', methods=['POST'])
def group_delete(group_id):
    """그룹 삭제 (하위 그룹 포함)"""
    group = Group.query.get_or_404(group_id)
    name = group.name

    # 재귀적으로 모든 하위 그룹과 소속 교인 해제
    def delete_group_recursive(g):
        # 하위 그룹 먼저 삭제
        for child in g.children:
            delete_group_recursive(child)
        # 소속 교인들의 그룹 해제
        for member in g.members:
            member.group_id = None
        db.session.delete(g)

    delete_group_recursive(group)
    db.session.commit()

    flash(f'{name} 그룹이 삭제되었습니다.', 'success')
    return redirect(url_for('group_list'))


# =============================================================================
# 출석 관리 라우트
# =============================================================================

@app.route('/attendance')
def attendance_list():
    """출석 현황"""
    # 날짜 파라미터 (기본: 오늘)
    date_str = request.args.get('date')
    if date_str:
        selected_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    else:
        selected_date = get_seoul_today()

    # 예배 종류 필터
    service_type = request.args.get('service', '주일예배')

    # 해당 날짜의 출석 기록
    attendance_records = Attendance.query.filter_by(
        date=selected_date,
        service_type=service_type
    ).all()

    # 출석한 교인 ID 목록
    attended_ids = {r.member_id for r in attendance_records if r.attended}

    # 전체 활동 교인
    members = Member.query.filter(Member.status.in_(['active', 'newcomer'])).order_by(Member.name).all()

    return render_template('attendance/list.html',
                         members=members,
                         attended_ids=attended_ids,
                         selected_date=selected_date,
                         service_type=service_type)


@app.route('/attendance/check', methods=['POST'])
def attendance_check():
    """출석 체크 처리"""
    selected_date = datetime.strptime(request.form.get('date'), '%Y-%m-%d').date()
    service_type = request.form.get('service_type', '주일예배')
    member_ids = request.form.getlist('member_ids')

    # 기존 출석 기록 삭제 (해당 날짜 + 예배 종류)
    Attendance.query.filter_by(date=selected_date, service_type=service_type).delete()

    # 새 출석 기록 생성
    for member_id in member_ids:
        attendance = Attendance(
            member_id=int(member_id),
            date=selected_date,
            service_type=service_type,
            attended=True
        )
        db.session.add(attendance)

    db.session.commit()

    flash(f'{selected_date.strftime("%Y-%m-%d")} {service_type} 출석이 저장되었습니다. ({len(member_ids)}명)', 'success')
    return redirect(url_for('attendance_list', date=selected_date.strftime('%Y-%m-%d'), service=service_type))


@app.route('/attendance/stats')
def attendance_stats():
    """출석 통계"""
    # 최근 4주 출석 통계
    from sqlalchemy import func

    stats = db.session.query(
        Attendance.date,
        Attendance.service_type,
        func.count(Attendance.id).label('count')
    ).filter(
        Attendance.attended == True
    ).group_by(
        Attendance.date,
        Attendance.service_type
    ).order_by(
        Attendance.date.desc()
    ).limit(20).all()

    return render_template('attendance/stats.html', stats=stats)


# =============================================================================
# 심방 기록 라우트
# =============================================================================

@app.route('/visits')
def visit_list():
    """심방 기록 목록"""
    visits = Visit.query.order_by(Visit.visit_date.desc()).limit(50).all()
    return render_template('visits/list.html', visits=visits)


@app.route('/visits/new', methods=['GET', 'POST'])
def visit_new():
    """심방 기록 등록"""
    if request.method == 'POST':
        member_id = request.form.get('member_id')
        visit_date = datetime.strptime(request.form.get('visit_date'), '%Y-%m-%d').date()
        visitor_name = request.form.get('visitor_name', '').strip()
        purpose = request.form.get('purpose', '').strip()
        notes = request.form.get('notes', '').strip()

        visit = Visit(
            member_id=int(member_id),
            visit_date=visit_date,
            visitor_name=visitor_name,
            purpose=purpose,
            notes=notes
        )

        db.session.add(visit)
        db.session.commit()

        flash('심방 기록이 저장되었습니다.', 'success')
        return redirect(url_for('visit_list'))

    members = Member.query.order_by(Member.name).all()
    return render_template('visits/form.html', members=members, visit=None)


@app.route('/visits/<int:visit_id>/edit', methods=['GET', 'POST'])
def visit_edit(visit_id):
    """심방 기록 수정"""
    visit = Visit.query.get_or_404(visit_id)

    if request.method == 'POST':
        visit.member_id = int(request.form.get('member_id'))
        visit.visit_date = datetime.strptime(request.form.get('visit_date'), '%Y-%m-%d').date()
        visit.visitor_name = request.form.get('visitor_name', '').strip()
        visit.purpose = request.form.get('purpose', '').strip()
        visit.notes = request.form.get('notes', '').strip()

        db.session.commit()

        flash('심방 기록이 수정되었습니다.', 'success')
        return redirect(url_for('visit_list'))

    members = Member.query.order_by(Member.name).all()
    return render_template('visits/form.html', members=members, visit=visit)


@app.route('/visits/<int:visit_id>/delete', methods=['POST'])
def visit_delete(visit_id):
    """심방 기록 삭제"""
    visit = Visit.query.get_or_404(visit_id)
    db.session.delete(visit)
    db.session.commit()
    flash('심방 기록이 삭제되었습니다.', 'success')
    return redirect(url_for('visit_list'))


# =============================================================================
# 헌금 기록 라우트
# =============================================================================

@app.route('/offerings')
def offering_list():
    """헌금 기록 목록"""
    # 날짜 필터
    month = request.args.get('month')
    if month:
        year, m = map(int, month.split('-'))
        start_date = date(year, m, 1)
        if m == 12:
            end_date = date(year + 1, 1, 1)
        else:
            end_date = date(year, m + 1, 1)
        offerings = Offering.query.filter(
            Offering.date >= start_date,
            Offering.date < end_date
        ).order_by(Offering.date.desc()).all()
    else:
        offerings = Offering.query.order_by(Offering.date.desc()).limit(100).all()

    # 월별 합계
    from sqlalchemy import func
    total = sum(o.amount for o in offerings)

    return render_template('offerings/list.html', offerings=offerings, total=total, month=month)


@app.route('/offerings/new', methods=['GET', 'POST'])
def offering_new():
    """헌금 기록 등록"""
    if request.method == 'POST':
        member_id = request.form.get('member_id')
        offering_date = datetime.strptime(request.form.get('date'), '%Y-%m-%d').date()
        amount = int(request.form.get('amount', 0))
        offering_type = request.form.get('offering_type', '').strip()
        notes = request.form.get('notes', '').strip()

        offering = Offering(
            member_id=int(member_id) if member_id else None,
            date=offering_date,
            amount=amount,
            offering_type=offering_type,
            notes=notes
        )

        db.session.add(offering)
        db.session.commit()

        flash('헌금 기록이 저장되었습니다.', 'success')
        return redirect(url_for('offering_list'))

    members = Member.query.order_by(Member.name).all()
    return render_template('offerings/form.html', members=members, offering=None)


@app.route('/offerings/<int:offering_id>/delete', methods=['POST'])
def offering_delete(offering_id):
    """헌금 기록 삭제"""
    offering = Offering.query.get_or_404(offering_id)
    db.session.delete(offering)
    db.session.commit()
    flash('헌금 기록이 삭제되었습니다.', 'success')
    return redirect(url_for('offering_list'))


# =============================================================================
# 새신자 관리 라우트
# =============================================================================

@app.route('/newcomers')
def newcomer_list():
    """새신자 목록 (등록 후 2년 이내)"""
    from sqlalchemy import or_
    two_years_ago = get_seoul_today() - timedelta(days=730)
    # 등록일이 2년 이내인 교인 (별세자, 타교회 제외)
    newcomers = Member.query.filter(
        Member.registration_date > two_years_ago,
        Member.status.notin_(['deceased', 'transferred']),
        or_(
            Member.member_status.is_(None),
            ~Member.member_status.in_(['별세', '소천', '타교회', '타교인', '이명'])
        )
    ).order_by(Member.registration_date.desc()).all()
    return render_template('newcomers/list.html', newcomers=newcomers)


# =============================================================================
# 생일/기념일 알림 라우트
# =============================================================================

@app.route('/birthdays')
def birthday_list():
    """이번 달 생일자 목록"""
    from sqlalchemy import extract

    current_month = get_seoul_today().month
    birthdays = Member.query.filter(
        extract('month', Member.birth_date) == current_month,
        Member.status.in_(['active', 'newcomer'])
    ).order_by(extract('day', Member.birth_date)).all()

    return render_template('birthdays/list.html', birthdays=birthdays, month=current_month)


# =============================================================================
# 엑셀 내보내기/가져오기 라우트
# =============================================================================

@app.route('/export/members')
def export_members():
    """교인 목록 엑셀 내보내기"""
    from openpyxl import Workbook
    from io import BytesIO
    from flask import send_file

    wb = Workbook()
    ws = wb.active
    ws.title = "교인 목록"

    # 헤더
    headers = ['이름', '전화번호', '이메일', '주소', '생년월일', '성별', '등록일', '세례일', '소속그룹', '상태', '메모']
    ws.append(headers)

    # 데이터
    members = Member.query.order_by(Member.name).all()
    for m in members:
        ws.append([
            m.name,
            m.phone or '',
            m.email or '',
            m.address or '',
            m.birth_date.strftime('%Y-%m-%d') if m.birth_date else '',
            m.gender or '',
            m.registration_date.strftime('%Y-%m-%d') if m.registration_date else '',
            m.baptism_date.strftime('%Y-%m-%d') if m.baptism_date else '',
            m.group.name if m.group else '',
            m.status or '',
            m.notes or ''
        ])

    # 파일로 저장
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'교인목록_{get_seoul_today().strftime("%Y%m%d")}.xlsx'
    )


@app.route('/import/members', methods=['GET', 'POST'])
def import_members():
    """교인 목록 엑셀 가져오기 (헤더 기반 동적 매핑)"""
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('파일을 선택해주세요.', 'danger')
            return redirect(request.url)

        file = request.files['file']
        if file.filename == '':
            flash('파일을 선택해주세요.', 'danger')
            return redirect(request.url)

        if file and file.filename.endswith('.xlsx'):
            from openpyxl import load_workbook

            wb = load_workbook(file)
            ws = wb.active

            # 헤더 매핑 (한글 헤더 → 필드명)
            HEADER_MAP = {
                '이름': 'name',
                '성명': 'name',
                'ID': 'registration_number',
                '등록번호': 'registration_number',
                '직분': 'member_type',
                '교구': 'district',
                '구역': 'cell_group',  # 구역도 cell_group으로 저장
                '속회': 'cell_group',
                '핸드폰': 'phone',
                '전화번호': 'phone',
                '연락처': 'phone',
                '생년월일': 'birth_date',
                '생일': 'birth_date',
                '주소': 'address',
                '이메일': 'email',
                '성별': 'gender',
                '세례일': 'baptism_date',
                '등록일': 'registration_date',
                '상태': 'status',
                '메모': 'notes',
                '비고': 'notes',
                '배우자': 'spouse_name',  # 임시 필드로 저장 후 notes에 추가
                '가족전체': 'family_info',  # 임시 필드
                '선교회': 'mission_group',
                '바나바': 'barnabas',
                '인도자': 'referrer',
            }

            # 1행에서 헤더 읽기
            headers = []
            for cell in ws[1]:
                header_name = str(cell.value).strip() if cell.value else ''
                headers.append(header_name)

            # 헤더 → 열 인덱스 매핑
            col_map = {}
            for idx, header in enumerate(headers):
                if header in HEADER_MAP:
                    field = HEADER_MAP[header]
                    if field not in col_map:  # 첫 번째 매칭만 사용
                        col_map[field] = idx

            count = 0
            skipped = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                # 이름 필드 확인
                name_idx = col_map.get('name')
                if name_idx is None or not row[name_idx]:
                    continue

                name = str(row[name_idx]).strip()

                # 전화번호 가져오기
                phone = None
                phone_idx = col_map.get('phone')
                if phone_idx is not None and row[phone_idx]:
                    phone = str(row[phone_idx]).strip()

                # 기존 교인 확인 (이름 + 전화번호로 중복 체크)
                existing = Member.query.filter_by(name=name, phone=phone).first()
                if existing:
                    skipped += 1
                    continue

                # Member 객체 생성
                member = Member(name=name, phone=phone, status='active')

                # 등록번호
                if 'registration_number' in col_map and row[col_map['registration_number']]:
                    member.registration_number = str(row[col_map['registration_number']]).strip()

                # 직분 (member_type)
                if 'member_type' in col_map and row[col_map['member_type']]:
                    member.member_type = str(row[col_map['member_type']]).strip()

                # 교구
                if 'district' in col_map and row[col_map['district']]:
                    district_val = str(row[col_map['district']]).strip()
                    # "2교구" → "2"
                    member.district = district_val.replace('교구', '').strip()

                # 속회/구역
                if 'cell_group' in col_map and row[col_map['cell_group']]:
                    member.cell_group = str(row[col_map['cell_group']]).strip()

                # 선교회
                if 'mission_group' in col_map and row[col_map['mission_group']]:
                    member.mission_group = str(row[col_map['mission_group']]).strip()

                # 바나바
                if 'barnabas' in col_map and row[col_map['barnabas']]:
                    member.barnabas = str(row[col_map['barnabas']]).strip()

                # 인도자
                if 'referrer' in col_map and row[col_map['referrer']]:
                    member.referrer = str(row[col_map['referrer']]).strip()

                # 주소
                if 'address' in col_map and row[col_map['address']]:
                    member.address = str(row[col_map['address']]).strip()

                # 이메일
                if 'email' in col_map and row[col_map['email']]:
                    member.email = str(row[col_map['email']]).strip()

                # 성별
                if 'gender' in col_map and row[col_map['gender']]:
                    member.gender = str(row[col_map['gender']]).strip()

                # 상태
                if 'status' in col_map and row[col_map['status']]:
                    status_val = str(row[col_map['status']]).strip()
                    if status_val in ['active', 'inactive', 'newcomer']:
                        member.status = status_val
                    elif status_val in ['활동', '재적']:
                        member.status = 'active'
                    elif status_val in ['비활동', '휴적']:
                        member.status = 'inactive'
                    elif status_val in ['새신자', '새가족']:
                        member.status = 'newcomer'

                # 메모 구성
                notes_parts = []
                if 'notes' in col_map and row[col_map['notes']]:
                    notes_parts.append(str(row[col_map['notes']]).strip())
                if 'spouse_name' in col_map and row[col_map['spouse_name']]:
                    notes_parts.append(f"배우자: {row[col_map['spouse_name']]}")
                if 'family_info' in col_map and row[col_map['family_info']]:
                    notes_parts.append(f"가족: {row[col_map['family_info']]}")
                if notes_parts:
                    member.notes = '\n'.join(notes_parts)

                # 날짜 필드 파싱 함수
                def parse_date(value):
                    if not value:
                        return None
                    if isinstance(value, date):
                        return value
                    if isinstance(value, datetime):
                        return value.date()
                    try:
                        value_str = str(value).strip()
                        # 다양한 형식 시도
                        for fmt in ['%Y-%m-%d', '%Y/%m/%d', '%y/%m/%d', '%m/%d/%y', '%Y.%m.%d']:
                            try:
                                return datetime.strptime(value_str, fmt).date()
                            except ValueError:
                                continue
                    except:
                        pass
                    return None

                # 생년월일
                if 'birth_date' in col_map:
                    member.birth_date = parse_date(row[col_map['birth_date']])

                # 세례일
                if 'baptism_date' in col_map:
                    member.baptism_date = parse_date(row[col_map['baptism_date']])

                # 등록일
                if 'registration_date' in col_map:
                    member.registration_date = parse_date(row[col_map['registration_date']])

                db.session.add(member)
                count += 1

            db.session.commit()
            msg = f'{count}명의 교인이 등록되었습니다.'
            if skipped > 0:
                msg += f' ({skipped}명 중복으로 건너뜀)'
            flash(msg, 'success')
            return redirect(url_for('member_list'))

        flash('xlsx 파일만 업로드 가능합니다.', 'danger')
        return redirect(request.url)

    return render_template('import/members.html')


# =============================================================================
# AI 채팅 API
# =============================================================================

# AI 도구 정의 (Function Calling)
AI_TOOLS = [
    {
        "type": "function",
        "function": {
            "name": "search_members",
            "description": "교인을 검색합니다. 이름, 상태, 그룹 등으로 검색할 수 있습니다.",
            "parameters": {
                "type": "object",
                "properties": {
                    "name": {"type": "string", "description": "교인 이름 (부분 일치)"},
                    "status": {"type": "string", "enum": ["active", "inactive", "newcomer"], "description": "상태"},
                    "group_name": {"type": "string", "description": "소속 그룹명"},
                    "gender": {"type": "string", "enum": ["남", "여"], "description": "성별"}
                }
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "register_member",
            "description": "새 교인을 등록하거나 기존 교인 정보를 업데이트합니다. 동명이인이 있으면 확인 후 처리합니다.",
            "parameters": {
                "type": "object",
                "properties": {
                    "name": {"type": "string", "description": "이름 (필수, 직분 포함 가능: 홍길동 집사)"},
                    "phone": {"type": "string", "description": "전화번호"},
                    "email": {"type": "string", "description": "이메일"},
                    "address": {"type": "string", "description": "주소"},
                    "birth_date": {"type": "string", "description": "생년월일 (YYYY-MM-DD)"},
                    "gender": {"type": "string", "enum": ["남", "여"], "description": "성별"},
                    "status": {"type": "string", "enum": ["active", "newcomer"], "description": "상태 (기본: active)"},
                    "notes": {"type": "string", "description": "메모/특이사항"},
                    "registration_number": {"type": "string", "description": "등록번호"},
                    "previous_church": {"type": "string", "description": "이전교회/출석교회"},
                    "previous_church_address": {"type": "string", "description": "이전교회 주소"},
                    "district": {"type": "string", "description": "교구 (숫자)"},
                    "cell_group": {"type": "string", "description": "속회"},
                    "mission_group": {"type": "string", "description": "선교회"},
                    "barnabas": {"type": "string", "description": "바나바 (담당 장로/권사)"},
                    "referrer": {"type": "string", "description": "인도자/관계"},
                    "faith_level": {"type": "string", "description": "신급 (학습/세례/유아세례/입교)"},
                    "update_existing_id": {"type": "integer", "description": "기존 교인 ID - 동명이인 중 특정 교인 정보를 업데이트할 때 사용"},
                    "force_new": {"type": "boolean", "description": "true이면 동명이인이 있어도 새 교인으로 등록 (이름에 번호 추가)"}
                },
                "required": ["name"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "update_member",
            "description": "기존 교인 정보를 수정합니다.",
            "parameters": {
                "type": "object",
                "properties": {
                    "member_id": {"type": "integer", "description": "교인 ID"},
                    "name": {"type": "string", "description": "이름"},
                    "phone": {"type": "string", "description": "전화번호"},
                    "email": {"type": "string", "description": "이메일"},
                    "address": {"type": "string", "description": "주소"},
                    "birth_date": {"type": "string", "description": "생년월일 (YYYY-MM-DD)"},
                    "gender": {"type": "string", "enum": ["남", "여"], "description": "성별"},
                    "status": {"type": "string", "enum": ["active", "inactive", "newcomer"], "description": "상태"},
                    "notes": {"type": "string", "description": "메모"}
                },
                "required": ["member_id"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "get_member_detail",
            "description": "특정 교인의 상세 정보를 조회합니다.",
            "parameters": {
                "type": "object",
                "properties": {
                    "member_id": {"type": "integer", "description": "교인 ID"},
                    "name": {"type": "string", "description": "교인 이름 (정확히 일치)"}
                }
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "get_newcomers",
            "description": "새신자 목록을 조회합니다.",
            "parameters": {
                "type": "object",
                "properties": {
                    "days": {"type": "integer", "description": "최근 N일 이내 등록자 (기본: 전체)"}
                }
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "get_birthdays",
            "description": "생일자 목록을 조회합니다.",
            "parameters": {
                "type": "object",
                "properties": {
                    "month": {"type": "integer", "description": "월 (1-12, 기본: 이번 달)"}
                }
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "get_absent_members",
            "description": "장기 결석자 목록을 조회합니다.",
            "parameters": {
                "type": "object",
                "properties": {
                    "weeks": {"type": "integer", "description": "N주 이상 결석 (기본: 3)"}
                }
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "recommend_visits",
            "description": "심방 우선순위를 추천합니다. 새신자, 장기결석자, 심방 안 간 분 등을 분석합니다.",
            "parameters": {
                "type": "object",
                "properties": {
                    "limit": {"type": "integer", "description": "추천 인원 수 (기본: 5)"}
                }
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "record_visit",
            "description": "심방 기록을 등록합니다.",
            "parameters": {
                "type": "object",
                "properties": {
                    "member_name": {"type": "string", "description": "심방 대상 교인 이름"},
                    "member_id": {"type": "integer", "description": "심방 대상 교인 ID"},
                    "visit_date": {"type": "string", "description": "심방 날짜 (YYYY-MM-DD, 기본: 오늘)"},
                    "visitor_name": {"type": "string", "description": "심방자 이름"},
                    "purpose": {"type": "string", "description": "심방 목적"},
                    "notes": {"type": "string", "description": "심방 내용"}
                }
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "get_statistics",
            "description": "교적 통계를 조회합니다.",
            "parameters": {
                "type": "object",
                "properties": {
                    "type": {"type": "string", "enum": ["overview", "attendance", "offering", "group"], "description": "통계 유형"}
                }
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "manage_group",
            "description": "그룹(셀/구역/목장)을 관리합니다.",
            "parameters": {
                "type": "object",
                "properties": {
                    "action": {"type": "string", "enum": ["list", "create", "add_member", "remove_member"], "description": "작업 유형"},
                    "group_name": {"type": "string", "description": "그룹명"},
                    "group_type": {"type": "string", "enum": ["cell", "district", "mokjang"], "description": "그룹 유형"},
                    "member_name": {"type": "string", "description": "교인 이름 (add_member/remove_member 시)"}
                }
            }
        }
    }
]


def execute_ai_function(function_name: str, arguments: dict) -> str:
    """AI 함수 호출 실행"""

    if function_name == "search_members":
        query = Member.query
        if arguments.get("name"):
            query = query.filter(Member.name.contains(arguments["name"]))
        if arguments.get("status"):
            query = query.filter(Member.status == arguments["status"])
        if arguments.get("gender"):
            query = query.filter(Member.gender == arguments["gender"])
        if arguments.get("group_name"):
            group = Group.query.filter(Group.name.contains(arguments["group_name"])).first()
            if group:
                query = query.filter(Member.group_id == group.id)

        members = query.order_by(Member.name).limit(20).all()
        if not members:
            return json.dumps({"result": "검색 결과가 없습니다.", "count": 0}, ensure_ascii=False)

        result = []
        for m in members:
            result.append({
                "id": m.id,
                "name": m.name,
                "phone": m.phone or "",
                "status": m.status,
                "group": m.group.name if m.group else "",
                "gender": m.gender or ""
            })
        return json.dumps({"members": result, "count": len(result)}, ensure_ascii=False)

    elif function_name == "register_member":
        # 직분 목록 (이름에서 분리)
        titles = ["목사", "장로", "권사", "집사", "전도사", "선교사", "성도"]

        # 이름에서 직분 분리
        input_name = arguments["name"].strip()
        base_name = input_name
        input_title = None
        for title in titles:
            if input_name.endswith(" " + title):
                base_name = input_name[:-len(title)-1].strip()
                input_title = title
                break
            elif input_name.endswith(title):
                base_name = input_name[:-len(title)].strip()
                input_title = title
                break

        # 동명이인 체크 - 기본 이름으로 검색 (직분 무관)
        all_members = Member.query.all()
        existing_members = []
        for m in all_members:
            member_base_name = m.name
            for title in titles:
                if m.name.endswith(" " + title):
                    member_base_name = m.name[:-len(title)-1].strip()
                    break
                elif m.name.endswith(title):
                    member_base_name = m.name[:-len(title)].strip()
                    break
            # 번호 접미사 제거 (홍길동(1) -> 홍길동)
            if "(" in member_base_name and member_base_name.endswith(")"):
                member_base_name = member_base_name[:member_base_name.rfind("(")].strip()

            if member_base_name == base_name:
                existing_members.append(m)

        # force_new가 True이면 동명이인으로 새로 등록 (이름에 번호 추가)
        if arguments.get("force_new") and existing_members:
            # 동명이인 번호 계산 - 기존 멤버들 중 최대 번호 찾기
            max_suffix = 0
            for m in existing_members:
                # 이미 번호가 붙은 경우 (홍길동(1), 홍길동(2) ...)
                if "(" in m.name and m.name.endswith(")"):
                    try:
                        suffix = int(m.name[m.name.rfind("(")+1:-1])
                        max_suffix = max(max_suffix, suffix + 1)
                    except:
                        max_suffix = max(max_suffix, 1)
                else:
                    max_suffix = max(max_suffix, 1)

            # 직분 유지하면서 번호 추가 (이명섭 집사 -> 이명섭(1) 집사)
            if input_title:
                new_name = f"{base_name}({max_suffix}) {input_title}"
            else:
                new_name = f"{base_name}({max_suffix})"
            arguments["name"] = new_name
            existing_members = []  # 새 이름으로 등록하므로 중복 없음

        # update_existing_id가 있으면 기존 교인 정보 업데이트
        if arguments.get("update_existing_id"):
            member = Member.query.get(arguments["update_existing_id"])
            if not member:
                return json.dumps({"error": "해당 교인을 찾을 수 없습니다."}, ensure_ascii=False)

            # 제공된 정보만 업데이트 - 기본 정보
            if arguments.get("phone"):
                member.phone = arguments["phone"]
            if arguments.get("email"):
                member.email = arguments["email"]
            if arguments.get("address"):
                member.address = arguments["address"]
            if arguments.get("gender"):
                member.gender = arguments["gender"]
            if arguments.get("status"):
                member.status = arguments["status"]
            if arguments.get("notes"):
                member.notes = arguments["notes"]
            if arguments.get("birth_date"):
                try:
                    member.birth_date = datetime.strptime(arguments["birth_date"], "%Y-%m-%d").date()
                except:
                    pass

            # 교회 관련 정보
            if arguments.get("registration_number"):
                member.registration_number = arguments["registration_number"]
            if arguments.get("previous_church"):
                member.previous_church = arguments["previous_church"]
            if arguments.get("previous_church_address"):
                member.previous_church_address = arguments["previous_church_address"]
            if arguments.get("district"):
                member.district = arguments["district"]
            if arguments.get("cell_group"):
                member.cell_group = arguments["cell_group"]
            if arguments.get("mission_group"):
                member.mission_group = arguments["mission_group"]
            if arguments.get("barnabas"):
                member.barnabas = arguments["barnabas"]
            if arguments.get("referrer"):
                member.referrer = arguments["referrer"]
            if arguments.get("faith_level"):
                member.faith_level = arguments["faith_level"]

            db.session.commit()
            return json.dumps({
                "success": True,
                "action": "updated",
                "message": f"'{member.display_name}' 교인 정보가 업데이트되었습니다.",
                "member_id": member.id
            }, ensure_ascii=False)

        # 동명이인이 있는 경우 사용자에게 확인 요청
        if existing_members:
            existing_info = []
            for m in existing_members:
                info = {
                    "id": m.id,
                    "name": m.name,
                    "display_name": m.display_name,
                    "phone": m.phone or "연락처 없음",
                    "birth_date": m.birth_date.strftime("%Y-%m-%d") if m.birth_date else "생년월일 없음",
                    "status": m.status,
                    "registration_date": m.registration_date.strftime("%Y-%m-%d") if m.registration_date else ""
                }
                existing_info.append(info)

            return json.dumps({
                "duplicate_found": True,
                "message": f"'{arguments['name']}' 이름의 교인이 이미 {len(existing_members)}명 등록되어 있습니다.",
                "existing_members": existing_info,
                "suggestion": "기존 교인 정보를 업데이트하려면 update_existing_id에 해당 교인 ID를 지정하세요. 동명이인으로 새로 등록하려면 force_new=true를 사용하세요.",
                "pending_data": {
                    "name": arguments.get("name"),
                    "phone": arguments.get("phone"),
                    "email": arguments.get("email"),
                    "address": arguments.get("address"),
                    "birth_date": arguments.get("birth_date"),
                    "gender": arguments.get("gender"),
                    "status": arguments.get("status"),
                    "notes": arguments.get("notes"),
                    "registration_number": arguments.get("registration_number"),
                    "previous_church": arguments.get("previous_church"),
                    "previous_church_address": arguments.get("previous_church_address"),
                    "district": arguments.get("district"),
                    "cell_group": arguments.get("cell_group"),
                    "mission_group": arguments.get("mission_group"),
                    "barnabas": arguments.get("barnabas"),
                    "referrer": arguments.get("referrer"),
                    "faith_level": arguments.get("faith_level")
                }
            }, ensure_ascii=False)

        # 새 교인 등록
        member = Member(
            name=arguments["name"],
            phone=arguments.get("phone"),
            email=arguments.get("email"),
            address=arguments.get("address"),
            gender=arguments.get("gender"),
            status=arguments.get("status", "active"),
            notes=arguments.get("notes"),
            registration_date=get_seoul_today(),
            # 교회 관련 정보
            registration_number=arguments.get("registration_number"),
            previous_church=arguments.get("previous_church"),
            previous_church_address=arguments.get("previous_church_address"),
            district=arguments.get("district"),
            cell_group=arguments.get("cell_group"),
            mission_group=arguments.get("mission_group"),
            barnabas=arguments.get("barnabas"),
            referrer=arguments.get("referrer"),
            faith_level=arguments.get("faith_level")
        )

        if arguments.get("birth_date"):
            try:
                member.birth_date = datetime.strptime(arguments["birth_date"], "%Y-%m-%d").date()
            except:
                pass

        db.session.add(member)
        db.session.commit()

        return json.dumps({
            "success": True,
            "action": "created",
            "message": f"'{member.name}' 교인이 등록되었습니다.",
            "member_id": member.id,
            "member": {
                "id": member.id,
                "name": member.name,
                "phone": member.phone,
                "status": member.status
            }
        }, ensure_ascii=False)

    elif function_name == "update_member":
        member = Member.query.get(arguments["member_id"])
        if not member:
            return json.dumps({"error": "해당 교인을 찾을 수 없습니다."}, ensure_ascii=False)

        if arguments.get("name"):
            member.name = arguments["name"]
        if arguments.get("phone"):
            member.phone = arguments["phone"]
        if arguments.get("email"):
            member.email = arguments["email"]
        if arguments.get("address"):
            member.address = arguments["address"]
        if arguments.get("gender"):
            member.gender = arguments["gender"]
        if arguments.get("status"):
            member.status = arguments["status"]
        if arguments.get("notes"):
            member.notes = arguments["notes"]
        if arguments.get("birth_date"):
            try:
                member.birth_date = datetime.strptime(arguments["birth_date"], "%Y-%m-%d").date()
            except:
                pass

        db.session.commit()
        return json.dumps({"success": True, "message": f"'{member.name}' 교인 정보가 수정되었습니다."}, ensure_ascii=False)

    elif function_name == "get_member_detail":
        member = None
        if arguments.get("member_id"):
            member = Member.query.get(arguments["member_id"])
        elif arguments.get("name"):
            member = Member.query.filter_by(name=arguments["name"]).first()

        if not member:
            return json.dumps({"error": "해당 교인을 찾을 수 없습니다."}, ensure_ascii=False)

        # 최근 출석
        recent_attendance = Attendance.query.filter_by(member_id=member.id, attended=True).order_by(Attendance.date.desc()).first()
        # 최근 심방
        recent_visit = Visit.query.filter_by(member_id=member.id).order_by(Visit.visit_date.desc()).first()

        return json.dumps({
            "id": member.id,
            "name": member.name,
            "phone": member.phone or "",
            "email": member.email or "",
            "address": member.address or "",
            "birth_date": member.birth_date.strftime("%Y-%m-%d") if member.birth_date else "",
            "gender": member.gender or "",
            "status": member.status,
            "group": member.group.name if member.group else "",
            "registration_date": member.registration_date.strftime("%Y-%m-%d") if member.registration_date else "",
            "baptism_date": member.baptism_date.strftime("%Y-%m-%d") if member.baptism_date else "",
            "notes": member.notes or "",
            "photo_url": member.photo_url or "",
            "last_attendance": recent_attendance.date.strftime("%Y-%m-%d") if recent_attendance else "기록 없음",
            "last_visit": recent_visit.visit_date.strftime("%Y-%m-%d") if recent_visit else "기록 없음"
        }, ensure_ascii=False)

    elif function_name == "get_newcomers":
        query = Member.query.filter_by(status="newcomer")
        if arguments.get("days"):
            cutoff = get_seoul_today() - timedelta(days=arguments["days"])
            query = query.filter(Member.registration_date >= cutoff)

        newcomers = query.order_by(Member.registration_date.desc()).all()
        result = []
        for m in newcomers:
            # 심방 여부 확인
            visited = Visit.query.filter_by(member_id=m.id).first() is not None
            result.append({
                "id": m.id,
                "name": m.name,
                "phone": m.phone or "",
                "registration_date": m.registration_date.strftime("%Y-%m-%d") if m.registration_date else "",
                "visited": visited
            })

        return json.dumps({"newcomers": result, "count": len(result)}, ensure_ascii=False)

    elif function_name == "get_birthdays":
        from sqlalchemy import extract
        month = arguments.get("month", get_seoul_today().month)

        birthdays = Member.query.filter(
            extract('month', Member.birth_date) == month,
            Member.status.in_(['active', 'newcomer'])
        ).order_by(extract('day', Member.birth_date)).all()

        result = []
        for m in birthdays:
            result.append({
                "id": m.id,
                "name": m.name,
                "birth_date": m.birth_date.strftime("%m월 %d일") if m.birth_date else "",
                "phone": m.phone or ""
            })

        return json.dumps({"birthdays": result, "count": len(result), "month": month}, ensure_ascii=False)

    elif function_name == "get_absent_members":
        weeks = arguments.get("weeks", 3)
        cutoff = get_seoul_today() - timedelta(weeks=weeks)

        # 최근 출석 기록이 있는 교인 ID
        recent_attended = db.session.query(Attendance.member_id).filter(
            Attendance.date >= cutoff,
            Attendance.attended == True
        ).distinct().subquery()

        # 활동 교인 중 최근 출석 안 한 사람
        absent = Member.query.filter(
            Member.status == 'active',
            ~Member.id.in_(db.session.query(recent_attended))
        ).all()

        result = []
        for m in absent:
            last_attendance = Attendance.query.filter_by(member_id=m.id, attended=True).order_by(Attendance.date.desc()).first()
            result.append({
                "id": m.id,
                "name": m.name,
                "phone": m.phone or "",
                "last_attendance": last_attendance.date.strftime("%Y-%m-%d") if last_attendance else "기록 없음"
            })

        return json.dumps({"absent_members": result, "count": len(result), "weeks": weeks}, ensure_ascii=False)

    elif function_name == "recommend_visits":
        limit = arguments.get("limit", 5)
        recommendations = []

        # 1. 새신자 중 심방 안 간 분
        newcomers = Member.query.filter_by(status="newcomer").all()
        for m in newcomers:
            if not Visit.query.filter_by(member_id=m.id).first():
                recommendations.append({
                    "id": m.id,
                    "name": m.name,
                    "phone": m.phone or "",
                    "reason": "새신자 - 아직 심방 전",
                    "priority": 1
                })

        # 2. 장기 결석자 (3주 이상)
        cutoff = get_seoul_today() - timedelta(weeks=3)
        recent_attended = db.session.query(Attendance.member_id).filter(
            Attendance.date >= cutoff,
            Attendance.attended == True
        ).distinct().subquery()

        absent = Member.query.filter(
            Member.status == 'active',
            ~Member.id.in_(db.session.query(recent_attended))
        ).all()

        for m in absent:
            if not any(r["id"] == m.id for r in recommendations):
                recommendations.append({
                    "id": m.id,
                    "name": m.name,
                    "phone": m.phone or "",
                    "reason": "3주 이상 출석 안 함",
                    "priority": 2
                })

        # 우선순위 정렬 후 제한
        recommendations.sort(key=lambda x: x["priority"])
        recommendations = recommendations[:limit]

        return json.dumps({"recommendations": recommendations, "count": len(recommendations)}, ensure_ascii=False)

    elif function_name == "record_visit":
        # 교인 찾기
        member = None
        if arguments.get("member_id"):
            member = Member.query.get(arguments["member_id"])
        elif arguments.get("member_name"):
            member = Member.query.filter_by(name=arguments["member_name"]).first()

        if not member:
            return json.dumps({"error": "해당 교인을 찾을 수 없습니다."}, ensure_ascii=False)

        visit_date = get_seoul_today()
        if arguments.get("visit_date"):
            try:
                visit_date = datetime.strptime(arguments["visit_date"], "%Y-%m-%d").date()
            except:
                pass

        visit = Visit(
            member_id=member.id,
            visit_date=visit_date,
            visitor_name=arguments.get("visitor_name", ""),
            purpose=arguments.get("purpose", ""),
            notes=arguments.get("notes", "")
        )

        db.session.add(visit)
        db.session.commit()

        return json.dumps({
            "success": True,
            "message": f"'{member.name}' 교인 심방 기록이 등록되었습니다.",
            "visit_id": visit.id
        }, ensure_ascii=False)

    elif function_name == "get_statistics":
        stat_type = arguments.get("type", "overview")

        if stat_type == "overview":
            total = Member.query.count()
            active = Member.query.filter_by(status="active").count()
            newcomer = Member.query.filter_by(status="newcomer").count()
            inactive = Member.query.filter_by(status="inactive").count()
            groups = Group.query.count()

            return json.dumps({
                "type": "overview",
                "total_members": total,
                "active": active,
                "newcomer": newcomer,
                "inactive": inactive,
                "groups": groups
            }, ensure_ascii=False)

        elif stat_type == "attendance":
            # 최근 4주 출석 통계
            from sqlalchemy import func
            four_weeks_ago = get_seoul_today() - timedelta(weeks=4)

            stats = db.session.query(
                Attendance.date,
                func.count(Attendance.id).label('count')
            ).filter(
                Attendance.date >= four_weeks_ago,
                Attendance.attended == True
            ).group_by(Attendance.date).order_by(Attendance.date.desc()).all()

            result = [{"date": s.date.strftime("%Y-%m-%d"), "count": s.count} for s in stats]
            return json.dumps({"type": "attendance", "stats": result}, ensure_ascii=False)

        elif stat_type == "group":
            groups = Group.query.all()
            result = []
            for g in groups:
                result.append({
                    "name": g.name,
                    "type": g.group_type or "",
                    "member_count": len(g.members)
                })
            return json.dumps({"type": "group", "groups": result}, ensure_ascii=False)

        return json.dumps({"error": "지원하지 않는 통계 유형입니다."}, ensure_ascii=False)

    elif function_name == "manage_group":
        action = arguments.get("action", "list")

        if action == "list":
            groups = Group.query.all()
            result = []
            for g in groups:
                leader = Member.query.get(g.leader_id) if g.leader_id else None
                result.append({
                    "id": g.id,
                    "name": g.name,
                    "type": g.group_type or "",
                    "leader": leader.name if leader else "",
                    "member_count": len(g.members)
                })
            return json.dumps({"groups": result, "count": len(result)}, ensure_ascii=False)

        elif action == "create":
            if not arguments.get("group_name"):
                return json.dumps({"error": "그룹명이 필요합니다."}, ensure_ascii=False)

            group = Group(
                name=arguments["group_name"],
                group_type=arguments.get("group_type", "cell")
            )
            db.session.add(group)
            db.session.commit()

            return json.dumps({
                "success": True,
                "message": f"'{group.name}' 그룹이 생성되었습니다.",
                "group_id": group.id
            }, ensure_ascii=False)

        elif action == "add_member":
            group = Group.query.filter_by(name=arguments.get("group_name")).first()
            member = Member.query.filter_by(name=arguments.get("member_name")).first()

            if not group:
                return json.dumps({"error": "해당 그룹을 찾을 수 없습니다."}, ensure_ascii=False)
            if not member:
                return json.dumps({"error": "해당 교인을 찾을 수 없습니다."}, ensure_ascii=False)

            member.group_id = group.id
            db.session.commit()

            return json.dumps({
                "success": True,
                "message": f"'{member.name}' 교인이 '{group.name}' 그룹에 추가되었습니다."
            }, ensure_ascii=False)

        return json.dumps({"error": "지원하지 않는 작업입니다."}, ensure_ascii=False)

    return json.dumps({"error": "알 수 없는 함수입니다."}, ensure_ascii=False)


def process_ai_chat(user_message: str, image_data: str = None) -> str:
    """AI 채팅 처리"""

    if not openai_client:
        return "OpenAI API 키가 설정되지 않았습니다. 환경변수 OPENAI_API_KEY를 설정해주세요."

    messages = [
        {
            "role": "system",
            "content": """당신은 교회 교적 관리 AI 어시스턴트입니다.

교인 등록, 검색, 수정, 심방 기록, 출석 관리 등을 도와드립니다.

사용자가 자연어로 요청하면 적절한 도구를 사용하여 작업을 수행하세요.

예시:
- "홍길동 집사님 등록해줘, 전화번호 010-1234-5678" → register_member 호출
- "김영희 권사님 정보 알려줘" → get_member_detail 호출
- "새신자 목록 보여줘" → get_newcomers 호출
- "이번 주 심방 갈 분 추천해줘" → recommend_visits 호출
- "지난 3주간 안 나온 분들" → get_absent_members 호출
- "이번 달 생일자" → get_birthdays 호출
- "전체 교인 수" → get_statistics 호출

동명이인 처리:
register_member 호출 시 동명이인이 있으면 duplicate_found: true 응답이 옵니다.
이 경우 사용자에게 기존 교인 목록을 보여주고 다음을 물어보세요:
1. 기존 교인 정보 업데이트: "홍길동 생년월일 추가해줘" → update_existing_id 사용
2. 동명이인으로 새로 등록: "새 홍길동으로 등록해줘" → force_new=true 사용

예시 응답:
"'홍길동' 이름의 교인이 이미 등록되어 있습니다:
1. 홍길동 집사 (010-1234-5678, 1985-03-15생)
2. 홍길동 성도 (연락처 없음, 생년월일 없음)

기존 교인 정보를 업데이트할까요, 아니면 동명이인으로 새로 등록할까요?"

사진이 첨부되면:
- 명함/등록카드 사진: 정보를 추출하여 등록 제안
- 인물 사진: 어떤 교인의 프로필 사진으로 등록할지 확인

응답은 친절하고 자연스럽게 해주세요. 한국어로 응답합니다."""
        }
    ]

    # 이미지가 있으면 Vision API 사용
    if image_data:
        messages.append({
            "role": "user",
            "content": [
                {"type": "text", "text": user_message if user_message else "이 사진을 분석해주세요. 명함이나 등록카드면 정보를 추출하고, 인물 사진이면 알려주세요."},
                {"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{image_data}"}}
            ]
        })
    else:
        messages.append({"role": "user", "content": user_message})

    try:
        response = openai_client.chat.completions.create(
            model="gpt-4o",
            messages=messages,
            tools=AI_TOOLS,
            tool_choice="auto",
            max_tokens=2000
        )

        assistant_message = response.choices[0].message

        # 도구 호출이 있는 경우
        if assistant_message.tool_calls:
            # 도구 실행 결과 수집
            tool_results = []
            for tool_call in assistant_message.tool_calls:
                function_name = tool_call.function.name
                arguments = json.loads(tool_call.function.arguments)
                result = execute_ai_function(function_name, arguments)
                tool_results.append({
                    "tool_call_id": tool_call.id,
                    "role": "tool",
                    "content": result
                })

            # 도구 결과를 포함하여 다시 요청
            messages.append(assistant_message)
            messages.extend(tool_results)

            final_response = openai_client.chat.completions.create(
                model="gpt-4o",
                messages=messages,
                max_tokens=2000
            )

            return final_response.choices[0].message.content

        return assistant_message.content

    except Exception as e:
        return f"죄송합니다, 오류가 발생했습니다: {str(e)}"


@app.route('/api/chat', methods=['POST'])
def api_chat():
    """AI 채팅 API"""
    data = request.get_json()
    message = data.get('message', '')
    image_data = data.get('image')  # base64 encoded image

    if not message and not image_data:
        return jsonify({"error": "메시지 또는 이미지가 필요합니다."}), 400

    response = process_ai_chat(message, image_data)
    return jsonify({"response": response})


@app.route('/api/upload-photo', methods=['POST'])
def api_upload_photo():
    """사진 업로드 API - Cloudinary 또는 로컬 저장소 사용"""
    if 'photo' not in request.files:
        return jsonify({"error": "파일이 없습니다."}), 400

    file = request.files['photo']
    member_id = request.form.get('member_id')

    if file.filename == '':
        return jsonify({"error": "파일이 선택되지 않았습니다."}), 400

    if file:
        try:
            # Cloudinary가 설정되어 있으면 클라우드에 업로드
            if cloudinary_configured:
                # 고유 public_id 생성
                public_id = f"church-registry/member_{member_id}_{datetime.now().strftime('%Y%m%d%H%M%S')}"

                # Cloudinary에 업로드
                result = cloudinary.uploader.upload(
                    file,
                    public_id=public_id,
                    folder="church-registry",
                    transformation=[
                        {"width": 400, "height": 400, "crop": "fill", "gravity": "face"}
                    ]
                )

                photo_url = result['secure_url']

                # 교인 사진 URL 업데이트
                if member_id:
                    member = Member.query.get(int(member_id))
                    if member:
                        member.photo_url = photo_url
                        db.session.commit()

                return jsonify({
                    "success": True,
                    "url": photo_url,
                    "storage": "cloudinary"
                })

            else:
                # Cloudinary가 없으면 로컬에 저장 (개발용)
                ext = file.filename.rsplit('.', 1)[1].lower() if '.' in file.filename else 'jpg'
                filename = f"member_{member_id}_{datetime.now().strftime('%Y%m%d%H%M%S')}.{ext}"
                filename = secure_filename(filename)

                filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                file.save(filepath)

                photo_url = f"/static/uploads/{filename}"

                # 교인 사진 URL 업데이트
                if member_id:
                    member = Member.query.get(int(member_id))
                    if member:
                        member.photo_url = photo_url
                        db.session.commit()

                return jsonify({
                    "success": True,
                    "url": photo_url,
                    "storage": "local"
                })

        except Exception as e:
            return jsonify({"error": f"업로드 실패: {str(e)}"}), 500

    return jsonify({"error": "업로드 실패"}), 500


@app.route('/api/analyze-excel', methods=['POST'])
def api_analyze_excel():
    """엑셀 파일 분석 API - AI가 열 매핑을 자동 분석"""
    if 'file' not in request.files:
        return jsonify({"error": "파일이 없습니다."}), 400

    file = request.files['file']
    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({"error": "엑셀 파일(.xlsx, .xls)만 지원합니다."}), 400

    try:
        from openpyxl import load_workbook

        wb = load_workbook(file)
        ws = wb.active

        # 첫 10행 정도만 샘플로 추출
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= 15:  # 헤더 + 14개 샘플
                break
            rows.append([str(cell) if cell is not None else "" for cell in row])

        if not rows:
            return jsonify({"error": "빈 엑셀 파일입니다."}), 400

        # AI에게 분석 요청
        if not openai_client:
            return jsonify({"error": "OpenAI API 키가 설정되지 않았습니다."}), 500

        # 엑셀 데이터를 텍스트로 변환
        excel_text = "엑셀 데이터 샘플:\n"
        for i, row in enumerate(rows):
            excel_text += f"행 {i+1}: {' | '.join(row)}\n"

        response = openai_client.chat.completions.create(
            model="gpt-4o",
            messages=[
                {
                    "role": "system",
                    "content": """당신은 교회 교적 데이터 분석 전문가입니다.
엑셀 파일의 열을 분석하여 교인 정보와 매핑해주세요.

분석 결과를 다음 JSON 형식으로 반환하세요:
{
    "header_row": 1,  // 헤더가 있는 행 번호 (1부터 시작)
    "mapping": {
        "name": 0,  // 이름 열 인덱스 (0부터 시작, 없으면 null)
        "phone": 1,
        "email": null,
        "address": 2,
        "birth_date": 3,
        "gender": 4,
        "registration_date": null,
        "baptism_date": null,
        "status": null,
        "notes": null
    },
    "sample_data": [
        {"name": "홍길동", "phone": "010-1234-5678", ...},
        {"name": "김영희", "phone": "010-9876-5432", ...}
    ],
    "total_rows": 100,  // 추정 데이터 행 수
    "analysis": "이 엑셀은 교인 명부로 보입니다. 이름, 연락처, 주소, 생년월일, 성별 정보가 있습니다."
}

주의:
- 헤더 행을 정확히 파악하세요
- 이름 열은 필수입니다
- 날짜 형식이 다양할 수 있습니다 (예: 1990-01-01, 1990.01.01, 90/1/1)
- 성별은 남/여, M/F, 남자/여자 등 다양할 수 있습니다"""
                },
                {"role": "user", "content": excel_text}
            ],
            response_format={"type": "json_object"},
            max_tokens=2000
        )

        analysis = json.loads(response.choices[0].message.content)

        # 전체 데이터 행 수 계산
        total_rows = sum(1 for row in ws.iter_rows(values_only=True)) - analysis.get("header_row", 1)
        analysis["total_rows"] = total_rows

        return jsonify({
            "success": True,
            "analysis": analysis,
            "raw_headers": rows[analysis.get("header_row", 1) - 1] if rows else []
        })

    except Exception as e:
        return jsonify({"error": f"엑셀 분석 중 오류: {str(e)}"}), 500


@app.route('/api/import-analyzed', methods=['POST'])
def api_import_analyzed():
    """분석된 매핑으로 엑셀 데이터 일괄 등록"""
    if 'file' not in request.files:
        return jsonify({"error": "파일이 없습니다."}), 400

    file = request.files['file']
    mapping_json = request.form.get('mapping')

    if not mapping_json:
        return jsonify({"error": "매핑 정보가 없습니다."}), 400

    try:
        mapping = json.loads(mapping_json)
        header_row = mapping.get('header_row', 1)
        col_mapping = mapping.get('mapping', {})

        from openpyxl import load_workbook
        wb = load_workbook(file)
        ws = wb.active

        imported = 0
        skipped = 0
        errors = []

        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i < header_row:  # 헤더 행까지 스킵
                continue

            row_data = list(row)

            # 이름 추출 (필수)
            name_idx = col_mapping.get('name')
            if name_idx is None or name_idx >= len(row_data) or not row_data[name_idx]:
                continue

            name = str(row_data[name_idx]).strip()
            if not name:
                continue

            # 중복 체크
            phone = None
            phone_idx = col_mapping.get('phone')
            if phone_idx is not None and phone_idx < len(row_data):
                phone = str(row_data[phone_idx]).strip() if row_data[phone_idx] else None

            existing = Member.query.filter_by(name=name, phone=phone).first()
            if existing:
                skipped += 1
                continue

            # 교인 생성
            member = Member(name=name, phone=phone, registration_date=get_seoul_today())

            # 나머지 필드 매핑
            for field in ['email', 'address', 'gender', 'notes']:
                idx = col_mapping.get(field)
                if idx is not None and idx < len(row_data) and row_data[idx]:
                    setattr(member, field, str(row_data[idx]).strip())

            # 날짜 필드 처리
            for date_field in ['birth_date', 'registration_date', 'baptism_date']:
                idx = col_mapping.get(date_field)
                if idx is not None and idx < len(row_data) and row_data[idx]:
                    try:
                        val = row_data[idx]
                        if isinstance(val, datetime):
                            setattr(member, date_field, val.date())
                        elif isinstance(val, date):
                            setattr(member, date_field, val)
                        elif isinstance(val, str):
                            # 다양한 날짜 형식 시도
                            for fmt in ['%Y-%m-%d', '%Y.%m.%d', '%Y/%m/%d', '%y-%m-%d', '%y.%m.%d']:
                                try:
                                    setattr(member, date_field, datetime.strptime(val.strip(), fmt).date())
                                    break
                                except:
                                    pass
                    except Exception as e:
                        pass

            # 상태 필드
            status_idx = col_mapping.get('status')
            if status_idx is not None and status_idx < len(row_data) and row_data[status_idx]:
                status_val = str(row_data[status_idx]).strip().lower()
                if '새신자' in status_val or 'new' in status_val:
                    member.status = 'newcomer'
                elif '비활동' in status_val or 'inactive' in status_val:
                    member.status = 'inactive'
                else:
                    member.status = 'active'

            try:
                db.session.add(member)
                db.session.commit()
                imported += 1
            except Exception as e:
                db.session.rollback()
                errors.append(f"행 {i+1}: {str(e)}")

        return jsonify({
            "success": True,
            "imported": imported,
            "skipped": skipped,
            "errors": errors[:10]  # 최대 10개 에러만
        })

    except Exception as e:
        return jsonify({"error": f"가져오기 중 오류: {str(e)}"}), 500


@app.route('/api/analyze-image', methods=['POST'])
def api_analyze_image():
    """이미지 분석 API - 명함/등록카드에서 정보 추출"""
    data = request.get_json()
    image_data = data.get('image')

    if not image_data:
        return jsonify({"error": "이미지가 없습니다."}), 400

    if not openai_client:
        return jsonify({"error": "OpenAI API 키가 설정되지 않았습니다."}), 500

    try:
        response = openai_client.chat.completions.create(
            model="gpt-4o",
            messages=[
                {
                    "role": "system",
                    "content": """이미지에서 사람 정보를 추출하세요.
명함, 교회 등록카드, 신분증 등에서 다음 정보를 찾아 JSON으로 반환하세요:

{
    "type": "namecard" | "registration_form" | "id_card" | "portrait" | "other",
    "extracted": {
        "name": "이름 (직분 포함, 예: 홍길동 집사)",
        "phone": "전화번호",
        "email": "이메일",
        "address": "주소",
        "birth_date": "생년월일 (YYYY-MM-DD 형식)",
        "gender": "남" | "여",
        "company": "소속/직장",
        "position": "직책",
        "registration_number": "등록번호",
        "previous_church": "이전교회/출석교회",
        "previous_church_address": "이전교회 주소",
        "district": "교구 (숫자만, 예: 1, 2, 3)",
        "cell_group": "속회",
        "mission_group": "선교회 (예: 14남선교회)",
        "barnabas": "바나바 (담당 장로/권사 이름)",
        "referrer": "인도자/관계 (예: 스스로, 지인소개)",
        "faith_level": "신급 (학습/세례/유아세례/입교 중 체크된 것)",
        "baptism_year": "세례연도",
        "baptism_church": "세례받은 교회",
        "registration_reason": "등록동기 (교회등록, 이사, 기타)",
        "notes": "기타 메모/특이사항 (여러 줄인 경우 모두 포함)"
    },
    "confidence": 0.9,
    "description": "이미지 설명"
}

찾을 수 없는 정보는 null로 설정하세요.
인물 사진만 있는 경우 type을 "portrait"로 설정하세요.
교회 등록 양식인 경우 모든 필드를 꼼꼼히 확인하세요."""
                },
                {
                    "role": "user",
                    "content": [
                        {"type": "text", "text": "이 이미지를 분석해주세요."},
                        {"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{image_data}"}}
                    ]
                }
            ],
            response_format={"type": "json_object"},
            max_tokens=1000
        )

        analysis = json.loads(response.choices[0].message.content)
        return jsonify({"success": True, "analysis": analysis})

    except Exception as e:
        return jsonify({"error": f"이미지 분석 중 오류: {str(e)}"}), 500


# =============================================================================
# 데이터베이스 마이그레이션
# =============================================================================

@app.route('/migrate-db')
def migrate_db():
    """데이터베이스 스키마 마이그레이션 (새 컬럼 추가)"""
    from sqlalchemy import text

    migrations = [
        # Member 테이블 기존 컬럼들
        "ALTER TABLE members ADD COLUMN IF NOT EXISTS registration_number VARCHAR(20)",
        "ALTER TABLE members ADD COLUMN IF NOT EXISTS previous_church VARCHAR(100)",
        "ALTER TABLE members ADD COLUMN IF NOT EXISTS previous_church_address VARCHAR(200)",
        "ALTER TABLE members ADD COLUMN IF NOT EXISTS district VARCHAR(50)",
        "ALTER TABLE members ADD COLUMN IF NOT EXISTS cell_group VARCHAR(50)",
        "ALTER TABLE members ADD COLUMN IF NOT EXISTS mission_group VARCHAR(50)",
        "ALTER TABLE members ADD COLUMN IF NOT EXISTS barnabas VARCHAR(50)",
        "ALTER TABLE members ADD COLUMN IF NOT EXISTS referrer VARCHAR(50)",
        "ALTER TABLE members ADD COLUMN IF NOT EXISTS faith_level VARCHAR(20)",

        # Member 테이블 새 컬럼들 (2025-01 추가)
        "ALTER TABLE members ADD COLUMN IF NOT EXISTS section VARCHAR(50)",  # 구역
        "ALTER TABLE members ADD COLUMN IF NOT EXISTS position_detail VARCHAR(30)",  # 상세 직분
        "ALTER TABLE members ADD COLUMN IF NOT EXISTS age_group VARCHAR(20)",  # 연령대
        "ALTER TABLE members ADD COLUMN IF NOT EXISTS attendance_status VARCHAR(20)",  # 출석 상태
        "ALTER TABLE members ADD COLUMN IF NOT EXISTS birth_lunar BOOLEAN DEFAULT FALSE",  # 음력 생일
        "ALTER TABLE members ADD COLUMN IF NOT EXISTS last_visit_date DATE",  # 마지막 심방일
        "ALTER TABLE members ADD COLUMN IF NOT EXISTS tel VARCHAR(20)",  # 집 전화
        "ALTER TABLE members ADD COLUMN IF NOT EXISTS zipcode VARCHAR(20)",  # 우편번호
        "ALTER TABLE members ADD COLUMN IF NOT EXISTS occupation VARCHAR(100)",  # 직업
        "ALTER TABLE members ADD COLUMN IF NOT EXISTS partner_id INTEGER",  # 배우자 ID
        "ALTER TABLE members ADD COLUMN IF NOT EXISTS member_status VARCHAR(30)",  # god4u state (재적, 별세, 타교회 등)
        "ALTER TABLE members ADD COLUMN IF NOT EXISTS deceased_date DATE",  # 별세일자
        "ALTER TABLE members ADD COLUMN IF NOT EXISTS transferred_date DATE",  # 타교회 이적일자
        "ALTER TABLE members ADD COLUMN IF NOT EXISTS car_number VARCHAR(20)",  # 차량번호
        "ALTER TABLE members ADD COLUMN IF NOT EXISTS family_members VARCHAR(200)",  # 가족 정보 (텍스트)

        # Group 테이블 새 컬럼들 (계층 구조)
        "ALTER TABLE groups ADD COLUMN IF NOT EXISTS parent_id INTEGER REFERENCES groups(id)",
        "ALTER TABLE groups ADD COLUMN IF NOT EXISTS level INTEGER DEFAULT 0",
        "ALTER TABLE groups ADD COLUMN IF NOT EXISTS description VARCHAR(200)",

        # FamilyRelationship 테이블 생성
        """CREATE TABLE IF NOT EXISTS family_relationships (
            id SERIAL PRIMARY KEY,
            member_id INTEGER NOT NULL REFERENCES members(id),
            related_member_id INTEGER NOT NULL REFERENCES members(id),
            relationship_type VARCHAR(20) NOT NULL,
            relationship_detail VARCHAR(30),
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(member_id, related_member_id, relationship_type)
        )""",

        # relationship_detail 컬럼 크기 확장 (기존 테이블용)
        "ALTER TABLE family_relationships ALTER COLUMN relationship_detail TYPE VARCHAR(30)",
    ]

    results = []
    for sql in migrations:
        try:
            db.session.execute(text(sql))
            db.session.commit()
            results.append(f"OK: {sql[:60]}...")
        except Exception as e:
            db.session.rollback()
            results.append(f"SKIP: {sql[:60]}... ({str(e)[:50]})")

    return jsonify({
        "message": "마이그레이션 완료",
        "results": results
    })


@app.route('/seed-groups')
def seed_groups():
    """기본 그룹 계층 구조 생성"""
    created = []
    skipped = []

    # 상위 그룹 정의
    top_groups = [
        {"name": "교구", "type": "district", "desc": "교구별 조직"},
        {"name": "선교회", "type": "mission", "desc": "선교회 조직"},
        {"name": "직분", "type": "position", "desc": "직분별 분류"},
        {"name": "교회학교", "type": "school", "desc": "교회학교 조직"},
        {"name": "새가족", "type": "newcomer", "desc": "새가족 관리"},
    ]

    # 하위 그룹 정의
    sub_groups = {
        "교구": ["1교구", "2교구", "3교구", "미등록"],
        "선교회": ["남선교회", "여선교회"],
        "직분": ["목사", "장로", "권사", "집사", "성도"],
        "교회학교": ["청년부", "청소년부", "아동부", "유치부", "유아부"],
        "새가족": [],
    }

    # 상위 그룹 생성
    for tg in top_groups:
        existing = Group.query.filter_by(name=tg["name"], parent_id=None).first()
        if existing:
            skipped.append(tg["name"])
            parent = existing
        else:
            parent = Group(
                name=tg["name"],
                group_type=tg["type"],
                level=0,
                description=tg["desc"],
                parent_id=None
            )
            db.session.add(parent)
            db.session.flush()  # ID 생성
            created.append(tg["name"])

        # 하위 그룹 생성
        for sub_name in sub_groups.get(tg["name"], []):
            existing_sub = Group.query.filter_by(name=sub_name, parent_id=parent.id).first()
            if existing_sub:
                skipped.append(f"{tg['name']} > {sub_name}")
            else:
                sub = Group(
                    name=sub_name,
                    group_type=tg["type"],
                    level=1,
                    parent_id=parent.id
                )
                db.session.add(sub)
                created.append(f"{tg['name']} > {sub_name}")

    db.session.commit()

    return jsonify({
        "message": "그룹 시드 완료",
        "created": created,
        "skipped": skipped
    })


# =============================================================================
# REST API (외부 연동용)
# =============================================================================

@app.route('/api/members', methods=['GET'])
def api_list_members():
    """교인 목록 조회 API (검색 지원)"""
    search = request.args.get('search', '').strip()
    limit = request.args.get('limit', type=int)

    query = Member.query

    # 검색어가 있으면 이름/전화번호로 검색
    if search:
        query = query.filter(
            db.or_(
                Member.name.ilike(f'%{search}%'),
                Member.phone.ilike(f'%{search}%')
            )
        )

    # 정렬
    query = query.order_by(Member.name)

    # 제한
    if limit:
        query = query.limit(limit)

    members = query.all()

    # 검색 모드일 때는 간단한 형식으로 반환
    if search or limit:
        return jsonify([{
            "id": m.id,
            "name": m.name,
            "phone": m.phone,
            "member_type": m.member_type,
            "photo_url": m.photo_url,
        } for m in members])

    return jsonify({
        "members": [_member_to_dict(m) for m in members],
        "total": len(members)
    })


@app.route('/api/members', methods=['POST'])
def api_create_member():
    """교인 등록 API (JSON)"""
    data = request.get_json()
    if not data:
        return jsonify({"error": "JSON 데이터가 필요합니다"}), 400

    if not data.get('name'):
        return jsonify({"error": "이름은 필수입니다"}), 400

    # 외부 ID 중복 체크
    if data.get('external_id'):
        existing = Member.query.filter_by(external_id=data['external_id']).first()
        if existing:
            return jsonify({
                "error": "이미 등록된 외부 ID입니다",
                "existing_id": existing.id
            }), 409

    member = Member(
        name=data.get('name'),
        phone=data.get('phone'),
        email=data.get('email'),
        address=data.get('address'),
        birth_date=_parse_date(data.get('birth_date')),
        gender=data.get('gender'),
        baptism_date=_parse_date(data.get('baptism_date')),
        registration_date=_parse_date(data.get('registration_date')) or get_seoul_today(),
        member_type=data.get('position') or data.get('member_type'),
        status=data.get('status', 'active'),
        notes=data.get('notes'),
        external_id=data.get('external_id'),
        photo_url=data.get('photo_url'),
    )

    db.session.add(member)
    db.session.commit()

    return jsonify({
        "success": True,
        "message": "교인이 등록되었습니다",
        "member": _member_to_dict(member)
    }), 201


@app.route('/api/members/natural-search', methods=['POST'])
def api_natural_search_members():
    """
    자연어 검색 API (GPT-5.1 사용)

    예시 쿼리:
    - "2025년에 등록한 남자 성도"
    - "1985년생 남자"
    - "홍길동 가족 전부"
    - "차량번호 12가 1234"
    - "의정부 거주하는 권사님"
    - "3교구 집사"
    - "010-1234로 시작하는 번호"
    """
    data = request.get_json() or {}
    query_text = data.get('query', '').strip()

    if not query_text:
        return jsonify({"error": "검색어(query)가 필요합니다"}), 400

    # === 빠른 검색 (GPT 없이 직접 처리) ===
    # 1. 단순 이름 검색 (한글 2-4자, 공백 없음)
    import re
    is_simple_name = re.match(r'^[가-힣]{2,4}$', query_text)

    # 2. 숫자만 있는 경우 (차량번호 일부)
    is_car_number_partial = re.match(r'^[0-9]{2,4}$', query_text)

    # 3. 전화번호 일부 (숫자와 하이픈)
    is_phone_partial = re.match(r'^[0-9-]{4,}$', query_text) and not is_car_number_partial

    if is_simple_name or is_car_number_partial or is_phone_partial:
        # GPT 호출 없이 직접 검색 (토큰 절약)
        try:
            if is_simple_name:
                members = Member.query.filter(
                    Member.name.ilike(f'%{query_text}%')
                ).order_by(Member.name).limit(50).all()
                explanation = f"'{query_text}' 이름으로 직접 검색"
            elif is_car_number_partial:
                members = Member.query.filter(
                    Member.car_number.ilike(f'%{query_text}%')
                ).order_by(Member.name).limit(50).all()
                explanation = f"차량번호 '{query_text}' 포함 검색"
            else:  # is_phone_partial
                members = Member.query.filter(
                    Member.phone.ilike(f'%{query_text}%')
                ).order_by(Member.name).limit(50).all()
                explanation = f"전화번호 '{query_text}' 포함 검색"

            return jsonify({
                "success": True,
                "query": query_text,
                "criteria": {"explanation": explanation, "gpt_used": False},
                "count": len(members),
                "members": [_member_to_dict(m) for m in members]
            })
        except Exception as e:
            app.logger.error(f"[빠른검색] 오류: {str(e)}")
            # 에러 발생 시 GPT 검색으로 fallback

    if not openai_client:
        return jsonify({"error": "OpenAI API 키가 설정되지 않았습니다"}), 500

    # GPT-5.1에 전달할 시스템 프롬프트 (교인 DB 스키마 설명)
    system_prompt = """당신은 교회 교인 데이터베이스 검색 도우미입니다.
사용자의 자연어 검색 요청을 분석하여 JSON 형식의 검색 조건으로 변환해주세요.

## 교인 데이터베이스 필드:
- name: 이름 (문자열)
- phone: 전화번호 (010-xxxx-xxxx 형식)
- address: 주소 (문자열)
- birth_date: 생년월일 (YYYY-MM-DD)
- birth_year: 출생연도 (정수, birth_date에서 추출)
- gender: 성별 ("남" 또는 "여")
- registration_date: 등록일 (YYYY-MM-DD)
- registration_year: 등록연도 (정수)
- member_type: 직분 (장로, 권사, 집사, 성도, 전도사 등)
- position_detail: 상세 직분 (시무장로, 은퇴권사 등)
- district: 교구 ("1교구[2025]", "2교구[2025]" 등)
- section: 구역 ("1구역", "2구역" 등)
- cell_group: 속회 (문자열)
- age_group: 연령대 (장년, 청년, 교회학교 등)
- attendance_status: 출석상태 (예배출석, 장기결석 등)
- car_number: 차량번호 ("12가 1234" 형식)
- occupation: 직업
- family_members: 가족 이름들 (공백/쉼표로 구분된 문자열)
- notes: 메모

## 출력 JSON 형식:
{
  "filters": [
    {"field": "필드명", "operator": "연산자", "value": "값"}
  ],
  "family_search": "가족검색할 이름" (선택사항),
  "limit": 결과수 (선택사항, 기본 50),
  "explanation": "검색 조건 설명"
}

## 연산자 종류:
- "eq": 정확히 일치
- "contains": 포함 (부분 일치)
- "starts_with": ~로 시작
- "gt": 보다 큼
- "gte": 이상
- "lt": 보다 작음
- "lte": 이하
- "year_eq": 연도 일치 (날짜 필드용)
- "in": 여러 값 중 하나 (value는 배열)

## 예시:
입력: "2025년에 등록한 남자 성도"
출력: {
  "filters": [
    {"field": "registration_year", "operator": "eq", "value": 2025},
    {"field": "gender", "operator": "eq", "value": "남"},
    {"field": "member_type", "operator": "eq", "value": "성도"}
  ],
  "explanation": "2025년에 등록한 남성 성도를 검색합니다"
}

입력: "1985년생 남자"
출력: {
  "filters": [
    {"field": "birth_year", "operator": "eq", "value": 1985},
    {"field": "gender", "operator": "eq", "value": "남"}
  ],
  "explanation": "1985년에 태어난 남성을 검색합니다"
}

입력: "홍길동 가족 전부"
출력: {
  "family_search": "홍길동",
  "explanation": "홍길동의 가족 구성원 전체를 검색합니다"
}

입력: "의정부 사는 권사님"
출력: {
  "filters": [
    {"field": "address", "operator": "contains", "value": "의정부"},
    {"field": "member_type", "operator": "eq", "value": "권사"}
  ],
  "explanation": "의정부에 거주하는 권사님을 검색합니다"
}

반드시 유효한 JSON만 출력하세요. 다른 텍스트 없이 JSON만 출력합니다."""

    try:
        # GPT-5.1 Responses API 호출
        response = openai_client.responses.create(
            model="gpt-5.1",
            input=[
                {"role": "system", "content": [{"type": "input_text", "text": system_prompt}]},
                {"role": "user", "content": [{"type": "input_text", "text": query_text}]}
            ],
            temperature=0.3
        )

        # 응답 텍스트 추출
        if getattr(response, "output_text", None):
            result_text = response.output_text.strip()
        else:
            text_chunks = []
            for item in getattr(response, "output", []) or []:
                for content in getattr(item, "content", []) or []:
                    if getattr(content, "type", "") == "text":
                        text_chunks.append(getattr(content, "text", ""))
            result_text = "\n".join(text_chunks).strip()

        # JSON 파싱
        import json
        # JSON 블록 추출 (마크다운 코드 블록 처리)
        if "```json" in result_text:
            result_text = result_text.split("```json")[1].split("```")[0].strip()
        elif "```" in result_text:
            result_text = result_text.split("```")[1].split("```")[0].strip()

        search_criteria = json.loads(result_text)

        # DB 세션 리프레시 (GPT API 호출 중 SSL 연결이 끊어질 수 있음)
        try:
            db.session.remove()  # 세션 완전 제거 후 새 연결 획득
        except Exception:
            pass

        # 검색 실행
        members = _execute_natural_search(search_criteria)

        return jsonify({
            "success": True,
            "query": query_text,
            "criteria": search_criteria,
            "count": len(members),
            "members": [_member_to_dict(m) for m in members]
        })

    except json.JSONDecodeError as e:
        app.logger.error(f"[자연어검색] JSON 파싱 실패: {result_text}")
        return jsonify({
            "error": "검색 조건 해석에 실패했습니다",
            "detail": str(e),
            "raw_response": result_text
        }), 500
    except Exception as e:
        app.logger.error(f"[자연어검색] 오류: {str(e)}")
        return jsonify({"error": f"검색 실패: {str(e)}"}), 500


def _execute_natural_search(criteria: dict) -> list:
    """GPT가 파싱한 검색 조건으로 실제 DB 검색 수행"""
    from sqlalchemy import extract, func

    query = Member.query

    # 가족 검색 모드
    if criteria.get("family_search"):
        family_name = criteria["family_search"]
        # 1. 해당 이름의 교인 찾기
        target_member = Member.query.filter(Member.name.ilike(f'%{family_name}%')).first()
        if target_member:
            # 2. 가족 관계로 연결된 모든 교인 찾기
            family_ids = {target_member.id}

            # FamilyRelationship에서 연결된 모든 교인 ID 수집
            rels = FamilyRelationship.query.filter(
                db.or_(
                    FamilyRelationship.member_id == target_member.id,
                    FamilyRelationship.related_member_id == target_member.id
                )
            ).all()

            for rel in rels:
                family_ids.add(rel.member_id)
                family_ids.add(rel.related_member_id)

            # 3. family_members 필드에서도 검색
            members_with_family = Member.query.filter(
                Member.family_members.ilike(f'%{family_name}%')
            ).all()
            for m in members_with_family:
                family_ids.add(m.id)

            return Member.query.filter(Member.id.in_(family_ids)).order_by(Member.name).all()
        else:
            # 이름으로 못 찾으면 family_members 필드에서 검색
            return Member.query.filter(
                Member.family_members.ilike(f'%{family_name}%')
            ).order_by(Member.name).all()

    # 필터 적용
    filters = criteria.get("filters", [])
    for f in filters:
        field = f.get("field")
        operator = f.get("operator")
        value = f.get("value")

        if not field or not operator:
            continue

        # 특수 필드 처리
        if field == "birth_year":
            if operator == "eq":
                query = query.filter(extract('year', Member.birth_date) == value)
            elif operator == "gte":
                query = query.filter(extract('year', Member.birth_date) >= value)
            elif operator == "lte":
                query = query.filter(extract('year', Member.birth_date) <= value)
            continue

        if field == "registration_year":
            if operator == "eq":
                query = query.filter(extract('year', Member.registration_date) == value)
            elif operator == "gte":
                query = query.filter(extract('year', Member.registration_date) >= value)
            elif operator == "lte":
                query = query.filter(extract('year', Member.registration_date) <= value)
            continue

        if field == "age":
            # 나이로 검색 (현재 날짜 기준)
            today = get_seoul_today()
            if operator == "eq":
                birth_year = today.year - value
                query = query.filter(extract('year', Member.birth_date) == birth_year)
            elif operator == "gte":
                birth_year = today.year - value
                query = query.filter(extract('year', Member.birth_date) <= birth_year)
            elif operator == "lte":
                birth_year = today.year - value
                query = query.filter(extract('year', Member.birth_date) >= birth_year)
            continue

        # 일반 필드
        column = getattr(Member, field, None)
        if column is None:
            continue

        if operator == "eq":
            query = query.filter(column == value)
        elif operator == "contains":
            query = query.filter(column.ilike(f'%{value}%'))
        elif operator == "starts_with":
            query = query.filter(column.ilike(f'{value}%'))
        elif operator == "gt":
            query = query.filter(column > value)
        elif operator == "gte":
            query = query.filter(column >= value)
        elif operator == "lt":
            query = query.filter(column < value)
        elif operator == "lte":
            query = query.filter(column <= value)
        elif operator == "in" and isinstance(value, list):
            query = query.filter(column.in_(value))
        elif operator == "year_eq":
            query = query.filter(extract('year', column) == value)

    # 결과 제한
    limit = criteria.get("limit", 50)
    query = query.order_by(Member.name).limit(limit)

    return query.all()


@app.route('/api/members/rebuild-family-relations', methods=['POST'])
def api_rebuild_family_relations():
    """
    가족 관계 재구축 API (동기화 없이 기존 데이터로 추론)

    기존 parent/child 관계를 기반으로:
    1. 같은 자녀를 둔 부모들 → 배우자로 연결
    2. 같은 부모를 둔 자녀들 → 형제로 연결
    """
    results = {
        "spouse_created": 0,
        "sibling_created": 0
    }

    try:
        # DB 세션 리프레시 (SSL 연결 오류 방지)
        try:
            db.session.remove()
        except Exception:
            pass

        # 1. 같은 자녀를 둔 부모들을 배우자로 연결
        app.logger.info("[가족관계 재구축] 부모 배우자 추론 시작...")

        child_relationships = FamilyRelationship.query.filter_by(relationship_type='child').all()
        child_parents = {}
        for rel in child_relationships:
            child_id = rel.member_id
            parent_id = rel.related_member_id
            if child_id not in child_parents:
                child_parents[child_id] = []
            if parent_id not in child_parents[child_id]:
                child_parents[child_id].append(parent_id)

        for child_id, parent_ids in child_parents.items():
            if len(parent_ids) >= 2:
                for i, p1_id in enumerate(parent_ids):
                    for p2_id in parent_ids[i+1:]:
                        existing = FamilyRelationship.query.filter(
                            db.or_(
                                db.and_(FamilyRelationship.member_id == p1_id,
                                       FamilyRelationship.related_member_id == p2_id,
                                       FamilyRelationship.relationship_type == 'spouse'),
                                db.and_(FamilyRelationship.member_id == p2_id,
                                       FamilyRelationship.related_member_id == p1_id,
                                       FamilyRelationship.relationship_type == 'spouse')
                            )
                        ).first()

                        if not existing:
                            p1 = Member.query.get(p1_id)
                            p2 = Member.query.get(p2_id)
                            if p1 and p2:
                                if p1.gender in ['M', '남', '남성', '남자']:
                                    detail1, detail2 = '남편', '아내'
                                else:
                                    detail1, detail2 = '아내', '남편'

                                db.session.add(FamilyRelationship(
                                    member_id=p1_id, related_member_id=p2_id,
                                    relationship_type='spouse', relationship_detail=detail2
                                ))
                                db.session.add(FamilyRelationship(
                                    member_id=p2_id, related_member_id=p1_id,
                                    relationship_type='spouse', relationship_detail=detail1
                                ))
                                results["spouse_created"] += 1
                                app.logger.info(f"[가족관계] 부모 배우자: {p1.name} ↔ {p2.name}")

        db.session.commit()

        # 2. 같은 부모를 둔 자녀들을 형제로 연결
        app.logger.info("[가족관계 재구축] 형제 관계 추론 시작...")

        parent_relationships = FamilyRelationship.query.filter_by(relationship_type='parent').all()
        parent_children = {}
        for rel in parent_relationships:
            parent_id = rel.member_id
            child_id = rel.related_member_id
            if parent_id not in parent_children:
                parent_children[parent_id] = []
            if child_id not in parent_children[parent_id]:
                parent_children[parent_id].append(child_id)

        for parent_id, child_ids in parent_children.items():
            if len(child_ids) >= 2:
                for i, c1_id in enumerate(child_ids):
                    for c2_id in child_ids[i+1:]:
                        existing = FamilyRelationship.query.filter(
                            db.or_(
                                db.and_(FamilyRelationship.member_id == c1_id,
                                       FamilyRelationship.related_member_id == c2_id,
                                       FamilyRelationship.relationship_type == 'sibling'),
                                db.and_(FamilyRelationship.member_id == c2_id,
                                       FamilyRelationship.related_member_id == c1_id,
                                       FamilyRelationship.relationship_type == 'sibling')
                            )
                        ).first()

                        if not existing:
                            c1 = Member.query.get(c1_id)
                            c2 = Member.query.get(c2_id)
                            if c1 and c2:
                                detail1, detail2 = '형제', '형제'
                                c1_male = c1.gender in ['M', '남', '남성', '남자']
                                c2_male = c2.gender in ['M', '남', '남성', '남자']
                                c1_older = True
                                if c1.birth_date and c2.birth_date:
                                    c1_older = c1.birth_date < c2.birth_date

                                if c1_male and c2_male:
                                    detail1 = '형' if not c1_older else '동생'
                                    detail2 = '형' if c1_older else '동생'
                                elif not c1_male and not c2_male:
                                    detail1 = '언니' if not c1_older else '동생'
                                    detail2 = '언니' if c1_older else '동생'
                                elif c1_male:
                                    detail1 = '오빠' if not c1_older else '남동생'
                                    detail2 = '누나' if c1_older else '여동생'
                                else:
                                    detail1 = '누나' if not c1_older else '여동생'
                                    detail2 = '오빠' if c1_older else '남동생'

                                db.session.add(FamilyRelationship(
                                    member_id=c1_id, related_member_id=c2_id,
                                    relationship_type='sibling', relationship_detail=detail2
                                ))
                                db.session.add(FamilyRelationship(
                                    member_id=c2_id, related_member_id=c1_id,
                                    relationship_type='sibling', relationship_detail=detail1
                                ))
                                results["sibling_created"] += 1
                                app.logger.info(f"[가족관계] 형제: {c1.name} ↔ {c2.name}")

        db.session.commit()

        return jsonify({
            "success": True,
            "message": f"가족 관계 재구축 완료: 배우자 {results['spouse_created']}쌍, 형제 {results['sibling_created']}쌍 생성",
            "results": results
        })

    except Exception as e:
        db.session.rollback()
        app.logger.error(f"[가족관계 재구축] 오류: {str(e)}")
        return jsonify({"error": str(e)}), 500


@app.route('/api/members/<int:member_id>', methods=['GET'])
def api_get_member(member_id):
    """교인 상세 조회 API"""
    member = Member.query.get_or_404(member_id)
    return jsonify(_member_to_dict(member))


@app.route('/api/members/<int:member_id>', methods=['PUT'])
def api_update_member(member_id):
    """교인 수정 API (JSON)"""
    member = Member.query.get_or_404(member_id)
    data = request.get_json()
    if not data:
        return jsonify({"error": "JSON 데이터가 필요합니다"}), 400

    # 업데이트 가능한 필드들
    if 'name' in data:
        member.name = data['name']
    if 'phone' in data:
        member.phone = data['phone']
    if 'email' in data:
        member.email = data['email']
    if 'address' in data:
        member.address = data['address']
    if 'birth_date' in data:
        member.birth_date = _parse_date(data['birth_date'])
    if 'gender' in data:
        member.gender = data['gender']
    if 'baptism_date' in data:
        member.baptism_date = _parse_date(data['baptism_date'])
    if 'registration_date' in data:
        member.registration_date = _parse_date(data['registration_date'])
    if 'position' in data or 'member_type' in data:
        member.member_type = data.get('position') or data.get('member_type')
    if 'status' in data:
        member.status = data['status']
    if 'notes' in data:
        member.notes = data['notes']
    if 'external_id' in data:
        member.external_id = data['external_id']
    if 'photo_url' in data:
        member.photo_url = data['photo_url']

    db.session.commit()

    return jsonify({
        "success": True,
        "message": "교인 정보가 수정되었습니다",
        "member": _member_to_dict(member)
    })


@app.route('/api/members/by-external-id/<external_id>', methods=['GET'])
def api_get_member_by_external_id(external_id):
    """외부 ID로 교인 조회 API"""
    member = Member.query.filter_by(external_id=external_id).first()
    if not member:
        return jsonify({"error": "교인을 찾을 수 없습니다"}), 404
    return jsonify(_member_to_dict(member))


@app.route('/api/members/<int:member_id>/photo', methods=['POST'])
def api_upload_member_photo(member_id):
    """교인 사진 업로드 API (base64)"""
    member = Member.query.get_or_404(member_id)
    data = request.get_json()

    if not data or not data.get('photo'):
        return jsonify({"error": "photo (base64) 데이터가 필요합니다"}), 400

    try:
        photo_base64 = data['photo']

        # Cloudinary 사용 가능하면 업로드
        if cloudinary_configured:
            result = cloudinary.uploader.upload(
                f"data:image/jpeg;base64,{photo_base64}",
                folder="church-registry/members",
                public_id=f"member_{member_id}"
            )
            photo_url = result['secure_url']
        else:
            # 로컬 저장
            import base64 as b64
            photo_data = b64.b64decode(photo_base64)
            filename = f"member_{member_id}.jpg"
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            with open(filepath, 'wb') as f:
                f.write(photo_data)
            photo_url = f"/static/uploads/{filename}"

        member.photo_url = photo_url
        db.session.commit()

        return jsonify({
            "success": True,
            "photo_url": photo_url
        })

    except Exception as e:
        return jsonify({"error": f"사진 업로드 실패: {str(e)}"}), 500


@app.route('/api/members/bulk', methods=['POST'])
def api_bulk_create_members():
    """교인 일괄 등록 API"""
    data = request.get_json()
    if not data or not data.get('members'):
        return jsonify({"error": "members 배열이 필요합니다"}), 400

    results = {"created": 0, "updated": 0, "failed": 0, "errors": []}

    for member_data in data['members']:
        try:
            external_id = member_data.get('external_id')

            # 기존 회원 확인
            existing = None
            if external_id:
                existing = Member.query.filter_by(external_id=external_id).first()

            if existing:
                # 업데이트
                for key, value in member_data.items():
                    if key in ['birth_date', 'baptism_date', 'registration_date']:
                        value = _parse_date(value)
                    if hasattr(existing, key) and value:
                        setattr(existing, key, value)
                results["updated"] += 1
            else:
                # 새로 생성
                member = Member(
                    name=member_data.get('name'),
                    phone=member_data.get('phone'),
                    email=member_data.get('email'),
                    address=member_data.get('address'),
                    birth_date=_parse_date(member_data.get('birth_date')),
                    gender=member_data.get('gender'),
                    registration_date=_parse_date(member_data.get('registration_date')) or get_seoul_today(),
                    member_type=member_data.get('position') or member_data.get('member_type'),
                    status=member_data.get('status', 'active'),
                    notes=member_data.get('notes'),
                    external_id=member_data.get('external_id'),
                )
                db.session.add(member)
                results["created"] += 1

        except Exception as e:
            results["failed"] += 1
            results["errors"].append(f"{member_data.get('name', 'Unknown')}: {str(e)}")

    db.session.commit()

    return jsonify({
        "success": True,
        "results": results
    })


# =============================================================================
# 가족 관계 API
# =============================================================================

@app.route('/api/members/<int:member_id>/family', methods=['GET'])
def api_get_member_family(member_id):
    """교인의 가족 관계 조회 (양방향)"""
    member = Member.query.get_or_404(member_id)

    def member_summary(m):
        if not m:
            return None
        return {
            "id": m.id,
            "name": m.name,
            "member_type": m.member_type,
            "photo_url": m.photo_url,
            "gender": m.gender,
            "age": m.age,
        }

    # 관계 유형 역방향 매핑
    REVERSE_TYPE = {
        'spouse': 'spouse',
        'parent': 'child',      # 상대가 나의 부모 → 나는 상대의 자녀
        'child': 'parent',      # 상대가 나의 자녀 → 나는 상대의 부모
        'sibling': 'sibling',
        'in_law': 'in_law',
        'grandparent': 'grandchild',
        'grandchild': 'grandparent',
        'extended': 'extended',
    }

    # 역방향 조회 시 detail 변환 (상대 관점 → 내 관점)
    def get_reverse_detail(rel_type, original_detail, related_member):
        """역방향 관계의 detail을 내 관점으로 변환"""
        if rel_type == 'child':
            # 원본: 상대가 child (나의 자녀) → 역방향: 상대가 parent (나의 부모)
            # detail '부모' 유지
            return original_detail
        elif rel_type == 'parent':
            # 원본: 상대가 parent (나의 부모) → 역방향: 상대가 child (나의 자녀)
            # detail을 성별에 따라 '아들/딸'로 변환
            if related_member and related_member.gender:
                if related_member.gender in ['M', '남', '남성', '남자']:
                    return '아들'
                else:
                    return '딸'
            return '자녀'
        elif rel_type == 'sibling':
            # 형제 관계는 양방향으로 다를 수 있음
            return original_detail
        else:
            return original_detail

    family_data = {
        "member": member_summary(member),
        "spouse": None,
        "parents": [],
        "children": [],
        "siblings": [],
        "in_laws": [],       # 인척 (시댁/처가/형제자매 배우자)
        "grandparents": [],  # 조부모
        "grandchildren": [], # 손자녀
        "extended": [],      # 확대가족 (삼촌/고모/조카/사촌 등)
    }

    # 1. 정방향 관계: 내가 member_id인 경우
    relationships = FamilyRelationship.query.filter_by(member_id=member_id).all()
    for rel in relationships:
        related = rel.related_member
        rel_info = {
            **member_summary(related),
            "relationship_id": rel.id,
            "relationship_detail": rel.relationship_detail,
            "relationship_type": rel.relationship_type,
        }

        if rel.relationship_type == 'spouse':
            family_data["spouse"] = rel_info
        elif rel.relationship_type == 'child':  # 내가 자녀 = 상대가 부모
            family_data["parents"].append(rel_info)
        elif rel.relationship_type == 'parent':  # 내가 부모 = 상대가 자녀
            family_data["children"].append(rel_info)
        elif rel.relationship_type == 'sibling':
            family_data["siblings"].append(rel_info)
        elif rel.relationship_type == 'in_law':
            family_data["in_laws"].append(rel_info)
        elif rel.relationship_type == 'grandparent':
            family_data["grandchildren"].append(rel_info)
        elif rel.relationship_type == 'grandchild':
            family_data["grandparents"].append(rel_info)
        elif rel.relationship_type == 'extended':
            family_data["extended"].append(rel_info)

    # 2. 역방향 관계: 내가 related_member_id인 경우 (중복 제외)
    existing_ids = set()
    if family_data["spouse"]:
        existing_ids.add(family_data["spouse"]["id"])
    for lst in [family_data["parents"], family_data["children"], family_data["siblings"],
                family_data["in_laws"], family_data["grandparents"], family_data["grandchildren"],
                family_data["extended"]]:
        for item in lst:
            existing_ids.add(item["id"])

    reverse_relationships = FamilyRelationship.query.filter_by(related_member_id=member_id).all()
    for rel in reverse_relationships:
        if rel.member_id in existing_ids:
            continue  # 이미 정방향에서 추가됨

        related = rel.member  # 역방향이므로 member가 상대방
        reversed_type = REVERSE_TYPE.get(rel.relationship_type, 'extended')

        # 역방향 관계의 detail을 내 관점으로 변환
        reversed_detail = get_reverse_detail(rel.relationship_type, rel.relationship_detail, related)

        rel_info = {
            **member_summary(related),
            "relationship_id": rel.id,
            "relationship_detail": reversed_detail,
            "relationship_type": reversed_type,
        }

        if reversed_type == 'spouse' and not family_data["spouse"]:
            family_data["spouse"] = rel_info
        elif reversed_type == 'parent':  # 역방향: 상대가 child → 나는 parent → 상대는 내 자녀
            family_data["children"].append(rel_info)
        elif reversed_type == 'child':  # 역방향: 상대가 parent → 나는 child → 상대는 내 부모
            family_data["parents"].append(rel_info)
        elif reversed_type == 'sibling':
            family_data["siblings"].append(rel_info)
        elif reversed_type == 'in_law':
            family_data["in_laws"].append(rel_info)
        elif reversed_type == 'grandparent':
            family_data["grandchildren"].append(rel_info)
        elif reversed_type == 'grandchild':
            family_data["grandparents"].append(rel_info)
        elif reversed_type == 'extended':
            family_data["extended"].append(rel_info)

        existing_ids.add(related.id)

    # 대가족 (Family 모델)
    if member.family_id:
        family = Family.query.get(member.family_id)
        if family:
            family_data["family_group"] = {
                "id": family.id,
                "name": family.family_name,
                "member_count": len(family.members),
            }

    # 3. 형제자매 가족 정보 (형제의 배우자, 자녀) + 배우자의 형제 가족
    sibling_families = []

    # 헬퍼 함수: 특정 형제의 가족 정보 수집
    def build_sibling_family(sibling_info, is_spouse_sibling=False):
        sibling_id = sibling_info["id"]
        sibling_family = {
            "sibling": sibling_info,
            "spouse": None,
            "children": [],
            "is_spouse_sibling": is_spouse_sibling  # 배우자의 형제인지 표시
        }

        # 형제의 배우자 찾기
        sibling_spouse_rel = FamilyRelationship.query.filter_by(
            member_id=sibling_id, relationship_type='spouse'
        ).first()
        if not sibling_spouse_rel:
            sibling_spouse_rel = FamilyRelationship.query.filter_by(
                related_member_id=sibling_id, relationship_type='spouse'
            ).first()

        if sibling_spouse_rel:
            spouse_member = (sibling_spouse_rel.related_member
                           if sibling_spouse_rel.member_id == sibling_id
                           else sibling_spouse_rel.member)
            if spouse_member:
                sibling_family["spouse"] = {
                    **member_summary(spouse_member),
                    "relationship_id": sibling_spouse_rel.id,
                    "relationship_detail": "형제 배우자" if not is_spouse_sibling else "처남댁/형수" if spouse_member.gender not in ['M', '남', '남성', '남자'] else "매형/매제",
                }

        # 형제의 자녀 찾기 (중복 제거용 set)
        seen_children_ids = set()
        sibling_children_rels = FamilyRelationship.query.filter_by(
            member_id=sibling_id, relationship_type='parent'
        ).all()
        sibling_children_rels += FamilyRelationship.query.filter_by(
            related_member_id=sibling_id, relationship_type='child'
        ).all()

        for rel in sibling_children_rels:
            child_member = (rel.related_member
                          if rel.member_id == sibling_id
                          else rel.member)
            if child_member and child_member.id not in seen_children_ids:
                seen_children_ids.add(child_member.id)
                sibling_family["children"].append({
                    **member_summary(child_member),
                    "relationship_id": rel.id,
                    "relationship_detail": "조카",
                })

        return sibling_family

    # 내 형제 가족
    for sibling_info in family_data["siblings"]:
        sibling_family = build_sibling_family(sibling_info, is_spouse_sibling=False)
        if sibling_family["spouse"] or sibling_family["children"]:
            sibling_families.append(sibling_family)

    # 배우자의 형제 가족 (인척 중 형제 관계인 사람들)
    if family_data["spouse"]:
        spouse_id = family_data["spouse"]["id"]
        # 배우자의 형제 찾기
        spouse_sibling_rels = FamilyRelationship.query.filter(
            db.or_(
                db.and_(FamilyRelationship.member_id == spouse_id,
                       FamilyRelationship.relationship_type == 'sibling'),
                db.and_(FamilyRelationship.related_member_id == spouse_id,
                       FamilyRelationship.relationship_type == 'sibling')
            )
        ).all()

        seen_spouse_siblings = set()
        for rel in spouse_sibling_rels:
            spouse_sibling_id = rel.related_member_id if rel.member_id == spouse_id else rel.member_id
            if spouse_sibling_id in seen_spouse_siblings or spouse_sibling_id == member_id:
                continue
            seen_spouse_siblings.add(spouse_sibling_id)

            spouse_sibling = Member.query.get(spouse_sibling_id)
            if spouse_sibling:
                spouse_sibling_info = {
                    **member_summary(spouse_sibling),
                    "relationship_detail": rel.relationship_detail or "처남/처제" if family_data["spouse"].get("gender") not in ['M', '남'] else "시누이/시동생"
                }
                sibling_family = build_sibling_family(spouse_sibling_info, is_spouse_sibling=True)
                if sibling_family["spouse"] or sibling_family["children"]:
                    sibling_families.append(sibling_family)

    family_data["sibling_families"] = sibling_families

    # 4. 인척 가족 정보 (인척의 배우자, 자녀)
    in_law_families = []
    for in_law_info in family_data["in_laws"]:
        in_law_id = in_law_info["id"]

        # 이미 형제 가족에 포함된 경우 스킵
        already_in_sibling = any(
            sf["sibling"]["id"] == in_law_id for sf in sibling_families
        )
        if already_in_sibling:
            continue

        in_law_family = {
            "in_law": in_law_info,
            "spouse": None,
            "children": []
        }

        # 인척의 배우자 찾기
        in_law_spouse_rel = FamilyRelationship.query.filter_by(
            member_id=in_law_id, relationship_type='spouse'
        ).first()
        if not in_law_spouse_rel:
            in_law_spouse_rel = FamilyRelationship.query.filter_by(
                related_member_id=in_law_id, relationship_type='spouse'
            ).first()

        if in_law_spouse_rel:
            spouse_member = (in_law_spouse_rel.related_member
                           if in_law_spouse_rel.member_id == in_law_id
                           else in_law_spouse_rel.member)
            # 배우자가 나 자신이거나 이미 표시된 경우 스킵
            if spouse_member and spouse_member.id != member_id:
                in_law_family["spouse"] = {
                    **member_summary(spouse_member),
                    "relationship_id": in_law_spouse_rel.id,
                    "relationship_detail": "인척 배우자",
                }

        # 인척의 자녀 찾기
        seen_children_ids = set()
        in_law_children_rels = FamilyRelationship.query.filter_by(
            member_id=in_law_id, relationship_type='parent'
        ).all()
        in_law_children_rels += FamilyRelationship.query.filter_by(
            related_member_id=in_law_id, relationship_type='child'
        ).all()

        for rel in in_law_children_rels:
            child_member = (rel.related_member
                          if rel.member_id == in_law_id
                          else rel.member)
            if child_member and child_member.id not in seen_children_ids:
                seen_children_ids.add(child_member.id)
                in_law_family["children"].append({
                    **member_summary(child_member),
                    "relationship_id": rel.id,
                    "relationship_detail": "조카",
                })

        # 배우자나 자녀가 있으면 추가
        if in_law_family["spouse"] or in_law_family["children"]:
            in_law_families.append(in_law_family)

    family_data["in_law_families"] = in_law_families

    return jsonify(family_data)


def propagate_extended_family_relationships(member, related_member, relationship_type, detail):
    """
    확대 가족 관계 자동 전파
    예: C와 D가 형제이고, C의 배우자가 E, D의 배우자가 F이면:
    - E와 D는 형제자매의 배우자 관계 (올케/시누이/형수 등)
    - F와 C는 형제자매의 배우자 관계 (올케/시누이/형수 등)
    - E와 F는 동서 관계
    """
    created_relations = []

    def get_spouse(m):
        """배우자 찾기"""
        rel = FamilyRelationship.query.filter_by(member_id=m.id, relationship_type='spouse').first()
        return Member.query.get(rel.related_member_id) if rel else None

    def add_relation_if_not_exists(m1_id, m2_id, rel_type, det1, det2):
        """중복 없이 관계 추가"""
        existing = FamilyRelationship.query.filter_by(
            member_id=m1_id, related_member_id=m2_id
        ).first()
        if not existing and m1_id != m2_id:
            rels = FamilyRelationship.create_bidirectional(m1_id, m2_id, rel_type, det1, det2)
            for r in rels:
                db.session.add(r)
            created_relations.extend(rels)

    # 1. 형제자매 관계 추가 시: 배우자들 사이의 관계 자동 생성
    if relationship_type == 'sibling':
        member_spouse = get_spouse(member)
        related_spouse = get_spouse(related_member)

        # member의 배우자 ↔ related_member (형제자매의 배우자)
        if member_spouse:
            # 내 배우자가 상대방의 형제자매의 배우자가 됨
            if member.gender == 'M':  # 남자의 배우자(아내) → 상대에게 형수/제수/올케
                detail_to_related = '형수' if detail in ['형', '오빠'] else '올케'
                detail_from_related = '시동생' if detail in ['형', '오빠'] else '시누이'
            else:  # 여자의 배우자(남편) → 상대에게 매형/매부/형부
                detail_to_related = '매형' if detail in ['누나', '언니'] else '매부'
                detail_from_related = '처제' if detail in ['누나', '언니'] else '처형'
            add_relation_if_not_exists(
                member_spouse.id, related_member.id, 'in_law',
                detail_to_related, detail_from_related
            )

        # related_member의 배우자 ↔ member (형제자매의 배우자)
        if related_spouse:
            if related_member.gender == 'M':
                detail_to_member = '형수' if detail in ['동생', '남동생', '여동생'] else '올케'
                detail_from_member = '시누이' if detail in ['동생', '남동생', '여동생'] else '시동생'
            else:
                detail_to_member = '매형' if detail in ['동생', '남동생', '여동생'] else '매부'
                detail_from_member = '처형' if detail in ['동생', '남동생', '여동생'] else '처제'
            add_relation_if_not_exists(
                related_spouse.id, member.id, 'in_law',
                detail_to_member, detail_from_member
            )

        # 두 배우자끼리 동서 관계
        if member_spouse and related_spouse:
            add_relation_if_not_exists(
                member_spouse.id, related_spouse.id, 'in_law', '동서', '동서'
            )

    # 2. 배우자 관계 추가 시: 기존 형제자매의 배우자와의 관계 자동 생성
    elif relationship_type == 'spouse':
        # member의 형제자매 찾기
        siblings = FamilyRelationship.query.filter_by(
            member_id=member.id, relationship_type='sibling'
        ).all()

        for sib_rel in siblings:
            sibling = Member.query.get(sib_rel.related_member_id)
            sibling_spouse = get_spouse(sibling)

            # member가 sibling보다 연상인지 확인
            member_is_older = False
            if member.birth_date and sibling.birth_date:
                member_is_older = member.birth_date < sibling.birth_date

            # 새 배우자(related_member) ↔ 형제자매
            # member가 남자인 경우: related_member(아내)가 sibling에게
            if member.gender == 'M':
                if sibling.gender == 'F':  # 여자 형제
                    detail_new = '올케'  # 오빠/남동생의 아내
                    detail_sib = '시누이'  # 오빠/남동생의 아내에게
                else:  # 남자 형제
                    # 형수: 형의 아내, 제수: 동생의 아내
                    detail_new = '형수' if member_is_older else '제수'
                    detail_sib = '시동생' if member_is_older else '시숙'
            # member가 여자인 경우: related_member(남편)이 sibling에게
            else:
                if sibling.gender == 'M':  # 남자 형제
                    # 매형: 누나/언니의 남편, 매부: 여동생의 남편
                    detail_new = '매형' if member_is_older else '매부'
                    detail_sib = '처남'  # 아내의 남자 형제
                else:  # 여자 형제
                    # 형부: 언니의 남편, 제부: 동생의 남편
                    detail_new = '형부' if member_is_older else '제부'
                    detail_sib = '처제' if member_is_older else '처형'

            add_relation_if_not_exists(
                related_member.id, sibling.id, 'in_law', detail_new, detail_sib
            )

            # 새 배우자(related_member) ↔ 형제자매의 배우자
            if sibling_spouse:
                add_relation_if_not_exists(
                    related_member.id, sibling_spouse.id, 'in_law', '동서', '동서'
                )

    # 3. 부모 관계 추가 시: 자녀들끼리 형제자매 관계
    elif relationship_type == 'parent':
        # related_member(자녀)의 다른 부모의 자녀들을 찾아서 형제자매로
        children_of_member = FamilyRelationship.query.filter_by(
            member_id=member.id, relationship_type='parent'
        ).all()

        for child_rel in children_of_member:
            if child_rel.related_member_id != related_member.id:
                other_child = Member.query.get(child_rel.related_member_id)
                # 형제자매 관계 추가
                add_relation_if_not_exists(
                    related_member.id, other_child.id, 'sibling', '형제', '형제'
                )

    return created_relations


@app.route('/api/members/<int:member_id>/family', methods=['POST'])
def api_add_family_relationship(member_id):
    """가족 관계 추가 (확대 가족 자동 연결 지원)"""
    member = Member.query.get_or_404(member_id)
    data = request.get_json() or {}

    related_member_id = data.get('related_member_id')
    relationship_type = data.get('relationship_type')  # spouse, parent, child, sibling, in_law, extended
    detail = data.get('detail')  # 아버지, 어머니, 아들, 딸, 시누이, 동서 등
    reverse_detail = data.get('reverse_detail')  # 역방향 상세 관계

    if not related_member_id:
        return jsonify({"error": "related_member_id가 필요합니다"}), 400

    # relationship_type이 없으면 detail에서 추론
    if not relationship_type and detail:
        relationship_type = FamilyRelationship.detail_to_type(detail)

    if not relationship_type:
        return jsonify({"error": "relationship_type 또는 detail이 필요합니다"}), 400

    valid_types = ['spouse', 'parent', 'child', 'sibling', 'in_law', 'grandparent', 'grandchild', 'extended']
    if relationship_type not in valid_types:
        return jsonify({"error": f"relationship_type은 {', '.join(valid_types)} 중 하나여야 합니다"}), 400

    related_member = Member.query.get(related_member_id)
    if not related_member:
        return jsonify({"error": "관계 대상 교인을 찾을 수 없습니다"}), 404

    # 이미 존재하는 관계 확인
    existing = FamilyRelationship.query.filter_by(
        member_id=member_id,
        related_member_id=related_member_id,
        relationship_type=relationship_type
    ).first()

    if existing:
        return jsonify({"error": "이미 등록된 관계입니다"}), 409

    try:
        # 양방향 관계 생성
        relationships = FamilyRelationship.create_bidirectional(
            member_id=member_id,
            related_member_id=related_member_id,
            relationship_type=relationship_type,
            detail=detail,
            reverse_detail=reverse_detail
        )

        for rel in relationships:
            db.session.add(rel)

        # 확대 가족 관계 자동 전파
        extended_rels = propagate_extended_family_relationships(
            member, related_member, relationship_type, detail
        )

        # 같은 Family 그룹으로 연결 (없으면 생성)
        if not member.family_id and not related_member.family_id:
            # 둘 다 가족이 없으면 새 가족 생성
            family = Family(family_name=f"{member.name} 가정")
            db.session.add(family)
            db.session.flush()
            member.family_id = family.id
            related_member.family_id = family.id
        elif member.family_id and not related_member.family_id:
            related_member.family_id = member.family_id
        elif not member.family_id and related_member.family_id:
            member.family_id = related_member.family_id

        db.session.commit()

        return jsonify({
            "success": True,
            "message": f"{member.name}님과 {related_member.name}님의 관계가 등록되었습니다",
            "relationship_ids": [rel.id for rel in relationships],
            "extended_relations_created": len(extended_rels) // 2 if extended_rels else 0,
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({"error": f"관계 등록 실패: {str(e)}"}), 500


@app.route('/api/relationship-options', methods=['GET'])
def api_get_relationship_options():
    """한국 가족 관계 옵션 목록 반환 (UI용)"""
    return jsonify(FamilyRelationship.get_all_relationship_options())


@app.route('/api/members/<int:member_id>/notes', methods=['PUT'])
def api_update_member_notes(member_id):
    """교인 메모 수정 API"""
    member = Member.query.get_or_404(member_id)
    data = request.get_json() or {}

    notes = data.get('notes', '').strip()

    try:
        member.notes = notes
        db.session.commit()

        return jsonify({
            "success": True,
            "message": "메모가 저장되었습니다",
            "notes": member.notes
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": f"메모 저장 실패: {str(e)}"}), 500


@app.route('/api/family-relationships/<int:relationship_id>', methods=['DELETE'])
def api_delete_family_relationship(relationship_id):
    """가족 관계 삭제 (양방향)"""
    rel = FamilyRelationship.query.get_or_404(relationship_id)

    member_id = rel.member_id
    related_member_id = rel.related_member_id
    relationship_type = rel.relationship_type

    try:
        # 역방향 관계도 삭제
        reverse_type = FamilyRelationship.REVERSE_RELATIONSHIPS.get(relationship_type, relationship_type)
        reverse_rel = FamilyRelationship.query.filter_by(
            member_id=related_member_id,
            related_member_id=member_id,
            relationship_type=reverse_type
        ).first()

        db.session.delete(rel)
        if reverse_rel:
            db.session.delete(reverse_rel)

        db.session.commit()

        return jsonify({"success": True, "message": "관계가 삭제되었습니다"})

    except Exception as e:
        db.session.rollback()
        return jsonify({"error": f"관계 삭제 실패: {str(e)}"}), 500


@app.route('/api/family-relationships/reset', methods=['GET', 'POST'])
def api_reset_family_relationships():
    """모든 가족 관계 삭제 (재생성 전 초기화용)"""
    try:
        count = FamilyRelationship.query.delete()
        db.session.commit()
        return jsonify({
            "success": True,
            "message": f"가족 관계 {count}건 삭제 완료. god4u 동기화를 다시 실행하세요."
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": f"삭제 실패: {str(e)}"}), 500


@app.route('/api/families', methods=['GET'])
def api_get_families():
    """대가족 목록 조회"""
    families = Family.query.all()

    result = []
    for family in families:
        members = Member.query.filter_by(family_id=family.id).all()
        result.append({
            "id": family.id,
            "family_name": family.family_name,
            "member_count": len(members),
            "members": [{
                "id": m.id,
                "name": m.name,
                "member_type": m.member_type,
                "photo_url": m.photo_url,
            } for m in members]
        })

    return jsonify(result)


@app.route('/api/families/<int:family_id>', methods=['GET'])
def api_get_family_detail(family_id):
    """대가족 상세 조회 (가계도 포함)"""
    family = Family.query.get_or_404(family_id)
    members = Member.query.filter_by(family_id=family_id).all()

    # 각 멤버의 관계 정보 포함
    members_with_relations = []
    for member in members:
        relationships = FamilyRelationship.query.filter_by(member_id=member.id).all()
        members_with_relations.append({
            "id": member.id,
            "name": member.name,
            "member_type": member.member_type,
            "photo_url": member.photo_url,
            "gender": member.gender,
            "age": member.age,
            "birth_date": member.birth_date.isoformat() if member.birth_date else None,
            "relationships": [{
                "related_id": rel.related_member_id,
                "type": rel.relationship_type,
                "detail": rel.relationship_detail,
            } for rel in relationships]
        })

    return jsonify({
        "id": family.id,
        "family_name": family.family_name,
        "members": members_with_relations,
    })


@app.route('/api/organization', methods=['GET'])
def api_get_organization_structure():
    """교회 조직 구조 자동 분류 API (교구/구역/속회/선교회)"""
    from collections import defaultdict

    members = Member.query.filter_by(status='active').all()

    # 조직별로 그룹화
    org_structure = {
        "districts": defaultdict(list),      # 교구별
        "sections": defaultdict(list),       # 구역별
        "cell_groups": defaultdict(list),    # 속회별
        "mission_groups": defaultdict(list), # 선교회별
    }

    for member in members:
        member_info = {
            "id": member.id,
            "name": member.name,
            "member_type": member.member_type,
            "photo_url": member.photo_url,
            "phone": member.phone,
        }

        if member.district:
            org_structure["districts"][member.district].append(member_info)
        if member.section:
            org_structure["sections"][member.section].append(member_info)
        if member.cell_group:
            org_structure["cell_groups"][member.cell_group].append(member_info)
        if member.mission_group:
            org_structure["mission_groups"][member.mission_group].append(member_info)

    # dict로 변환하고 각 그룹 인원수 추가
    result = {}
    for org_type, groups in org_structure.items():
        result[org_type] = [
            {
                "name": name,
                "member_count": len(members_list),
                "members": members_list
            }
            for name, members_list in sorted(groups.items())
        ]

    # 전체 통계
    result["statistics"] = {
        "total_active": len(members),
        "district_count": len(result["districts"]),
        "section_count": len(result["sections"]),
        "cell_group_count": len(result["cell_groups"]),
        "mission_group_count": len(result["mission_groups"]),
    }

    return jsonify(result)


@app.route('/api/organization/<org_type>', methods=['GET'])
def api_get_organization_by_type(org_type):
    """특정 조직 유형별 분류 (district, section, cell_group, mission_group)"""
    from collections import defaultdict

    valid_types = ['district', 'section', 'cell_group', 'mission_group']
    if org_type not in valid_types:
        return jsonify({"error": f"유효한 조직 유형: {', '.join(valid_types)}"}), 400

    field_map = {
        'district': 'district',
        'section': 'section',
        'cell_group': 'cell_group',
        'mission_group': 'mission_group',
    }

    members = Member.query.filter_by(status='active').all()
    groups = defaultdict(list)

    for member in members:
        group_name = getattr(member, field_map[org_type])
        if group_name:
            groups[group_name].append({
                "id": member.id,
                "name": member.name,
                "member_type": member.member_type,
                "position_detail": member.position_detail,
                "photo_url": member.photo_url,
                "phone": member.phone,
                "age": member.age,
            })

    result = [
        {
            "name": name,
            "member_count": len(members_list),
            "members": members_list
        }
        for name, members_list in sorted(groups.items())
    ]

    return jsonify({
        "org_type": org_type,
        "groups": result,
        "total_groups": len(result),
        "total_members": sum(g["member_count"] for g in result),
    })


@app.route('/api/families', methods=['POST'])
def api_create_family():
    """대가족 그룹 생성"""
    data = request.get_json() or {}

    family_name = data.get('family_name')
    member_ids = data.get('member_ids', [])

    if not family_name:
        return jsonify({"error": "family_name이 필요합니다"}), 400

    family = Family(family_name=family_name)
    db.session.add(family)
    db.session.flush()

    # 멤버들을 가족에 추가
    for member_id in member_ids:
        member = Member.query.get(member_id)
        if member:
            member.family_id = family.id

    db.session.commit()

    return jsonify({
        "success": True,
        "family_id": family.id,
        "family_name": family.family_name,
    })


@app.route('/api/families/<int:family_id>/members', methods=['POST'])
def api_add_family_member(family_id):
    """대가족에 멤버 추가"""
    family = Family.query.get_or_404(family_id)
    data = request.get_json() or {}

    member_id = data.get('member_id')
    if not member_id:
        return jsonify({"error": "member_id가 필요합니다"}), 400

    member = Member.query.get(member_id)
    if not member:
        return jsonify({"error": "교인을 찾을 수 없습니다"}), 404

    member.family_id = family_id
    db.session.commit()

    return jsonify({
        "success": True,
        "message": f"{member.name}님이 {family.family_name}에 추가되었습니다"
    })


@app.route('/api/sync/auto-link-families', methods=['POST'])
def api_auto_link_families():
    """god4u 데이터 기반 자동 가족 연결"""
    data = request.get_json() or {}

    results = {
        "spouse_linked": 0,
        "families_created": 0,
        "errors": []
    }

    try:
        # 1. partner_id 기반 배우자 연결
        members_with_partner = Member.query.filter(Member.partner_id.isnot(None)).all()

        for member in members_with_partner:
            partner_external_id = member.partner_id
            partner = Member.query.filter_by(external_id=str(partner_external_id)).first()

            if partner:
                # 이미 배우자 관계가 있는지 확인
                existing = FamilyRelationship.query.filter_by(
                    member_id=member.id,
                    related_member_id=partner.id,
                    relationship_type='spouse'
                ).first()

                if not existing:
                    # 성별에 따라 상세 관계 설정
                    if member.gender == 'M':
                        detail, reverse_detail = '남편', '아내'
                    elif member.gender == 'F':
                        detail, reverse_detail = '아내', '남편'
                    else:
                        detail, reverse_detail = '배우자', '배우자'

                    relationships = FamilyRelationship.create_bidirectional(
                        member_id=member.id,
                        related_member_id=partner.id,
                        relationship_type='spouse',
                        detail=detail,
                        reverse_detail=reverse_detail
                    )

                    for rel in relationships:
                        db.session.add(rel)

                    # 같은 가족 그룹으로 연결
                    if not member.family_id and not partner.family_id:
                        family = Family(family_name=f"{member.name} 가정")
                        db.session.add(family)
                        db.session.flush()
                        member.family_id = family.id
                        partner.family_id = family.id
                        results["families_created"] += 1
                    elif member.family_id and not partner.family_id:
                        partner.family_id = member.family_id
                    elif not member.family_id and partner.family_id:
                        member.family_id = partner.family_id

                    results["spouse_linked"] += 1

        db.session.commit()

        return jsonify({
            "success": True,
            "results": results
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({"error": f"자동 연결 실패: {str(e)}"}), 500


def _member_to_dict(member):
    """Member 객체를 딕셔너리로 변환"""
    return {
        "id": member.id,
        "name": member.name,
        "phone": member.phone,
        "email": member.email,
        "address": member.address,
        "birth_date": member.birth_date.isoformat() if member.birth_date else None,
        "gender": member.gender,
        "age": member.age,
        "baptism_date": member.baptism_date.isoformat() if member.baptism_date else None,
        "registration_date": member.registration_date.isoformat() if member.registration_date else None,
        "member_type": member.member_type,
        "status": member.status,
        "notes": member.notes,
        "external_id": member.external_id,
        "photo_url": member.photo_url,
        "created_at": member.created_at.isoformat() if member.created_at else None,
        "updated_at": member.updated_at.isoformat() if member.updated_at else None,
    }


def _parse_date(date_str):
    """날짜 문자열을 date 객체로 변환"""
    if not date_str:
        return None
    if isinstance(date_str, date):
        return date_str
    try:
        # 여러 형식 시도
        for fmt in ['%Y-%m-%d', '%Y/%m/%d', '%Y.%m.%d']:
            try:
                return datetime.strptime(date_str, fmt).date()
            except ValueError:
                continue
        return None
    except:
        return None


# =============================================================================
# god4u 연동 (양방향 동기화)
# =============================================================================

GOD4U_BASE_URL = "http://god4u.dimode.co.kr"
GOD4U_API_URL = f"{GOD4U_BASE_URL}/Handler/GetPersonListMobileJSon.asmx/GetPersonSearchListDefault"
GOD4U_UPDATE_URL = f"{GOD4U_BASE_URL}/WebMobile/WebChurch/PersonModifyDetailExecute.cshtml"
GOD4U_PHOTO_URL = f"{GOD4U_BASE_URL}/Handler/DisplayImage.ashx"


@app.route('/sync')
def sync_page():
    """동기화 페이지"""
    return render_template('sync.html')


@app.route('/api/sync/backup-god4u', methods=['POST'])
def api_backup_god4u():
    """god4u 데이터 백업 API - JSON 파일로 다운로드"""
    import time
    from flask import Response

    data = request.get_json() or {}
    cookies = data.get('cookies', {})

    if not cookies.get('ASP.NET_SessionId') or not cookies.get('pastorinfo'):
        return jsonify({"error": "god4u 쿠키가 필요합니다"}), 400

    try:
        session = requests.Session()
        session.cookies.update(cookies)

        headers = {
            "Content-Type": "application/json; charset=UTF-8",
            "Accept": "application/json, text/javascript, */*; q=0.01",
            "X-Requested-With": "XMLHttpRequest",
            "Origin": GOD4U_BASE_URL,
            "Referer": f"{GOD4U_BASE_URL}/WebMobile/WebChurch/RangeList.cshtml",
        }

        # 첫 페이지 조회
        payload = _create_god4u_payload(page=1, page_size=100)
        response = session.post(GOD4U_API_URL, json=payload, headers=headers, timeout=60)

        if response.status_code != 200:
            return jsonify({"error": f"god4u 연결 실패: {response.status_code}"}), 500

        result = response.json()
        if "d" in result:
            result = json.loads(result["d"])

        total_pages = int(result.get("totalpage", 1))
        all_persons = result.get("personInfo", [])

        # 모든 페이지 조회
        for page in range(2, total_pages + 1):
            time.sleep(0.3)
            payload = _create_god4u_payload(page=page, page_size=100)
            response = session.post(GOD4U_API_URL, json=payload, headers=headers, timeout=60)
            if response.status_code == 200:
                page_data = response.json()
                if "d" in page_data:
                    page_data = json.loads(page_data["d"])
                all_persons.extend(page_data.get("personInfo", []))

        # JSON 백업 파일 생성
        backup_data = {
            "backup_date": datetime.now().isoformat(),
            "total_count": len(all_persons),
            "source": "god4u.dimode.co.kr",
            "members": all_persons
        }

        json_str = json.dumps(backup_data, ensure_ascii=False, indent=2)

        return Response(
            json_str,
            mimetype='application/json',
            headers={'Content-Disposition': f'attachment;filename=god4u_backup_{datetime.now().strftime("%Y%m%d_%H%M%S")}.json'}
        )

    except Exception as e:
        return jsonify({"error": f"백업 실패: {str(e)}"}), 500


@app.route('/api/sync/preview', methods=['POST'])
def api_sync_preview():
    """동기화 미리보기 API - 변경될 내용을 먼저 보여줌"""
    import time

    data = request.get_json() or {}
    cookies = data.get('cookies', {})

    if not cookies.get('ASP.NET_SessionId') or not cookies.get('pastorinfo'):
        return jsonify({"error": "god4u 쿠키가 필요합니다"}), 400

    try:
        session = requests.Session()
        session.cookies.update(cookies)

        headers = {
            "Content-Type": "application/json; charset=UTF-8",
            "Accept": "application/json, text/javascript, */*; q=0.01",
            "X-Requested-With": "XMLHttpRequest",
            "Origin": GOD4U_BASE_URL,
            "Referer": f"{GOD4U_BASE_URL}/WebMobile/WebChurch/RangeList.cshtml",
        }

        # 모든 페이지 조회
        payload = _create_god4u_payload(page=1, page_size=100)
        response = session.post(GOD4U_API_URL, json=payload, headers=headers, timeout=60)

        if response.status_code != 200:
            return jsonify({"error": f"god4u API 오류: {response.status_code}"}), 500

        # 세션 만료 감지
        content_type = response.headers.get('Content-Type', '')
        if 'text/html' in content_type or response.text.strip().startswith('<'):
            return jsonify({"error": "god4u 세션이 만료되었습니다. 다시 로그인 후 쿠키를 갱신해주세요."}), 401

        data = response.json()
        if "d" in data:
            data = json.loads(data["d"])

        total_pages = int(data.get("totalpage", 1))
        all_persons = data.get("personInfo", [])

        # 나머지 페이지 조회
        for page in range(2, total_pages + 1):
            time.sleep(0.5)
            payload = _create_god4u_payload(page=page, page_size=100)
            try:
                response = session.post(GOD4U_API_URL, json=payload, headers=headers, timeout=60)
                content_type = response.headers.get('Content-Type', '')
                if 'text/html' in content_type or response.text.strip().startswith('<'):
                    break
                if response.status_code == 200:
                    page_data = response.json()
                    if "d" in page_data:
                        page_data = json.loads(page_data["d"])
                    all_persons.extend(page_data.get("personInfo", []))
            except:
                break

        # 기존 회원 로드
        existing_members = {m.external_id: m for m in Member.query.filter(Member.external_id.isnot(None)).all()}

        # 변경사항 분석
        new_members = []
        changed_members = []

        for person in all_persons:
            external_id = person.get("id")
            existing = existing_members.get(external_id) if external_id else None

            god4u_data = {
                "name": person.get("name", ""),
                "phone": person.get("handphone", ""),
                "tel": person.get("tel", ""),
                "email": person.get("email", ""),
                "address": person.get("addr", ""),
                "district": person.get("range", ""),
                "section": person.get("range1", ""),
                "cell_group": person.get("range2", ""),
                "member_type": person.get("cvname", ""),
                "position_detail": person.get("cvname1", ""),
                "attendance_status": person.get("state3", ""),
            }

            if not existing:
                new_members.append({
                    "external_id": external_id,
                    "name": god4u_data["name"],
                    "phone": god4u_data["phone"],
                    "district": god4u_data["district"],
                    "member_type": god4u_data["member_type"],
                })
            else:
                # 변경된 필드 확인 (주요 필드만 비교)
                changes = []

                # 기본 정보
                if existing.name != god4u_data["name"] and god4u_data["name"]:
                    changes.append({"field": "이름", "old": existing.name, "new": god4u_data["name"]})
                if existing.phone != god4u_data["phone"] and god4u_data["phone"]:
                    changes.append({"field": "연락처", "old": existing.phone, "new": god4u_data["phone"]})
                if existing.email != god4u_data["email"] and god4u_data["email"]:
                    changes.append({"field": "이메일", "old": existing.email, "new": god4u_data["email"]})
                if existing.address != god4u_data["address"] and god4u_data["address"]:
                    changes.append({"field": "주소", "old": existing.address, "new": god4u_data["address"]})

                # 교회 조직 정보
                if (existing.district or "") != god4u_data["district"] and god4u_data["district"]:
                    changes.append({"field": "교구", "old": existing.district or "", "new": god4u_data["district"]})
                if (existing.section or "") != god4u_data["section"] and god4u_data["section"]:
                    changes.append({"field": "구역", "old": existing.section or "", "new": god4u_data["section"]})
                if (existing.cell_group or "") != god4u_data["cell_group"] and god4u_data["cell_group"]:
                    changes.append({"field": "속회", "old": existing.cell_group or "", "new": god4u_data["cell_group"]})

                # 직분 정보
                if (existing.member_type or "") != god4u_data["member_type"] and god4u_data["member_type"]:
                    changes.append({"field": "직분", "old": existing.member_type or "", "new": god4u_data["member_type"]})
                if (existing.position_detail or "") != god4u_data["position_detail"] and god4u_data["position_detail"]:
                    changes.append({"field": "상세직분", "old": existing.position_detail or "", "new": god4u_data["position_detail"]})

                # 출석 상태
                if (existing.attendance_status or "") != god4u_data["attendance_status"] and god4u_data["attendance_status"]:
                    changes.append({"field": "출석상태", "old": existing.attendance_status or "", "new": god4u_data["attendance_status"]})

                if changes:
                    changed_members.append({
                        "external_id": external_id,
                        "id": existing.id,
                        "name": existing.name,
                        "member_type": god4u_data["member_type"],
                        "changes": changes,
                    })

        return jsonify({
            "success": True,
            "summary": {
                "total_god4u": len(all_persons),
                "total_local": len(existing_members),
                "new_count": len(new_members),
                "changed_count": len(changed_members),
            },
            "new_members": new_members[:50],  # 최대 50명만 표시
            "changed_members": changed_members[:50],
        })

    except Exception as e:
        return jsonify({"error": f"미리보기 실패: {str(e)}"}), 500


@app.route('/api/sync/god4u-to-registry', methods=['POST'])
def api_sync_god4u_to_registry():
    """god4u → church-registry 동기화 API"""
    import time

    data = request.get_json() or {}
    cookies = data.get('cookies', {})
    test_only = data.get('test_only', False)

    if not cookies.get('ASP.NET_SessionId') or not cookies.get('pastorinfo'):
        return jsonify({"error": "god4u 쿠키가 필요합니다 (ASP.NET_SessionId, pastorinfo)"}), 400

    try:
        session = requests.Session()
        session.cookies.update(cookies)

        # 첫 페이지로 전체 개수 확인
        headers = {
            "Content-Type": "application/json; charset=UTF-8",
            "Accept": "application/json, text/javascript, */*; q=0.01",
            "X-Requested-With": "XMLHttpRequest",
            "Origin": GOD4U_BASE_URL,
            "Referer": f"{GOD4U_BASE_URL}/WebMobile/WebChurch/RangeList.cshtml",
        }

        # 연결 테스트 모드: 첫 페이지만 조회하여 연결 확인
        payload = _create_god4u_payload(page=1, page_size=10)
        response = session.post(GOD4U_API_URL, json=payload, headers=headers, timeout=30)

        if response.status_code != 200:
            return jsonify({"error": f"god4u API 오류: {response.status_code}"}), 500

        test_data = response.json()
        if "d" in test_data:
            test_data = json.loads(test_data["d"])

        total_count = int(test_data.get("totalcount", 0))

        # test_only 모드면 여기서 성공 반환
        if test_only:
            return jsonify({
                "success": True,
                "message": f"연결 성공! 총 {total_count}명의 교인 데이터가 있습니다.",
                "total_count": total_count
            })

        # 전체 동기화: 첫 페이지를 100개로 다시 조회
        results = {"created": 0, "updated": 0, "failed": 0, "total": total_count}

        payload = _create_god4u_payload(page=1, page_size=100)
        response = session.post(GOD4U_API_URL, json=payload, headers=headers, timeout=60)

        if response.status_code != 200:
            return jsonify({"error": f"god4u API 오류: {response.status_code}"}), 500

        data = response.json()
        if "d" in data:
            data = json.loads(data["d"])

        total_pages = int(data.get("totalpage", 1))

        # 모든 페이지 크롤링
        all_persons = data.get("personInfo", [])
        session_expired = False

        for page in range(2, total_pages + 1):
            time.sleep(0.5)  # 요청 간격 증가
            payload = _create_god4u_payload(page=page, page_size=100)
            try:
                response = session.post(GOD4U_API_URL, json=payload, headers=headers, timeout=60)

                # 세션 만료 감지: HTML 응답 체크
                content_type = response.headers.get('Content-Type', '')
                if 'text/html' in content_type or response.text.strip().startswith('<'):
                    session_expired = True
                    break

                if response.status_code == 200:
                    page_data = response.json()
                    if "d" in page_data:
                        page_data = json.loads(page_data["d"])
                    all_persons.extend(page_data.get("personInfo", []))
            except json.JSONDecodeError:
                session_expired = True
                break
            except Exception:
                break

        if session_expired and len(all_persons) == 0:
            return jsonify({"error": "god4u 세션이 만료되었습니다. 다시 로그인 후 쿠키를 갱신해주세요."}), 401

        # 기존 회원 미리 로드 (배치 처리)
        existing_members = {m.external_id: m for m in Member.query.filter(Member.external_id.isnot(None)).all()}

        # god4u 우선 필드 (동기화 시 덮어쓰기)
        GOD4U_FIELDS = [
            'name', 'phone', 'email', 'address', 'birth_date', 'gender', 'registration_date',
            'member_type', 'position_detail', 'district', 'section', 'cell_group', 'mission_group',
            'age_group', 'attendance_status', 'birth_lunar', 'last_visit_date', 'tel', 'zipcode',
            'occupation', 'previous_church', 'partner_id', 'photo_url',
            'car_number', 'family_members', 'notes'  # 추가: 차량번호, 가족정보, 메모
        ]
        # 로컬 우선 필드 (기존 값 유지)
        LOCAL_FIELDS = ['status', 'barnabas', 'referrer']  # notes는 god4u에서 동기화

        # 선택적 동기화: 특정 회원만 업데이트
        selected_ids = data.get('selected_ids', None)  # None이면 전체, 리스트면 해당 회원만

        # church-registry에 저장
        for person in all_persons:
            try:
                external_id = person.get("id")
                existing = existing_members.get(external_id) if external_id else None

                # god4u에서 가져온 데이터 (전체 필드 매핑)
                god4u_data = {
                    # 기본 정보
                    "name": person.get("name", ""),
                    "phone": person.get("handphone", ""),
                    "tel": person.get("tel", ""),  # 집 전화
                    "email": person.get("email", ""),
                    "address": person.get("addr", ""),
                    "zipcode": person.get("zipcode", ""),
                    "birth_date": _parse_date(person.get("birth", "")),
                    "birth_lunar": person.get("solar") == "음",  # 음력 여부
                    "gender": "M" if person.get("sex") == "남" else "F" if person.get("sex") == "여" else None,

                    # 등록/심방 정보
                    "registration_date": _parse_date(person.get("regday", "")),
                    "last_visit_date": _parse_date(person.get("lastvisitday", "")),

                    # 등록번호 (god4u id)
                    "registration_number": str(external_id) if external_id else None,

                    # 직분 정보
                    "member_type": person.get("cvname", ""),  # 권사, 집사 등
                    "position_detail": person.get("cvname1", ""),  # 시무권사, 은퇴권사 등

                    # 교회 조직
                    "district": person.get("range", ""),  # 교구 (3교구[2025])
                    "section": person.get("range1", ""),  # 구역 (18구역)
                    "cell_group": person.get("range2", ""),  # 속회 (서울1속)
                    "mission_group": person.get("range3", ""),  # 선교회

                    # 상태 정보
                    "age_group": person.get("state1", ""),  # 장년, 청년 등
                    "attendance_status": person.get("state3", ""),  # 예배출석, 장기결석 등
                    "member_status": person.get("state", ""),  # 재적, 별세, 타교회 등

                    # 기타 정보
                    "occupation": person.get("occu") or person.get("occu1", ""),
                    "previous_church": person.get("prechurch", ""),
                    "external_id": external_id,

                    # 사진 URL (god4u에서 직접 링크)
                    "photo_url": f"{GOD4U_PHOTO_URL}?id={external_id}" if external_id else None,
                }

                # 배우자 연결 (partnerid가 있으면 저장)
                partner_id_str = person.get("partnerid")
                if partner_id_str:
                    god4u_data["partner_id"] = partner_id_str  # 나중에 매핑 필요

                # 차량번호 (god4u carnum)
                carnum = person.get("carnum", "").strip()
                if carnum:
                    god4u_data["car_number"] = carnum

                # 가족 정보 (god4u ran1)
                ran1 = person.get("ran1", "").strip()
                if ran1:
                    god4u_data["family_members"] = ran1

                # 메모 (god4u etc) - 항상 덮어씀 (이전 "가족:" "차량:" 텍스트 제거)
                etc_notes = person.get("etc", "").strip()
                god4u_data["notes"] = etc_notes  # 빈 값이어도 저장하여 이전 데이터 정리

                # 상태 결정 로직
                # god4u state 필드 기반 (별세, 타교회 우선)
                member_state = person.get("state", "")
                attendance_state = person.get("state3", "")

                if member_state in ['별세', '소천']:
                    determined_status = "deceased"
                elif member_state in ['타교회', '타교인', '이명']:
                    determined_status = "transferred"
                elif attendance_state == "예배출석":
                    determined_status = "active"
                else:
                    determined_status = "inactive"

                local_data = {
                    "status": determined_status,
                }

                if existing:
                    # 선택적 동기화: selected_ids가 있으면 해당 회원만 업데이트
                    if selected_ids is not None and external_id not in selected_ids:
                        continue

                    # 기존 회원: god4u 필드 업데이트 (car_number, family_members, notes 포함)
                    for key, value in god4u_data.items():
                        if value is not None:
                            setattr(existing, key, value)

                    # 상태 업데이트: god4u state 기반으로 덮어씀
                    existing.status = determined_status

                    results["updated"] += 1
                else:
                    # 신규 회원: 전체 데이터로 생성
                    member_data = {**god4u_data, **local_data}
                    member = Member(**{k: v for k, v in member_data.items() if v is not None})
                    db.session.add(member)
                    results["created"] += 1

            except Exception as e:
                results["failed"] += 1

        db.session.commit()

        # ===== 가족 관계 자동 생성 =====
        family_results = {"created": 0, "skipped": 0}

        # 1. partner_id로 배우자 관계 생성
        members_with_partner = Member.query.filter(Member.partner_id.isnot(None)).all()
        for member in members_with_partner:
            try:
                # partner_id로 배우자 찾기 (external_id 기준)
                spouse = Member.query.filter_by(external_id=str(member.partner_id)).first()
                if spouse and spouse.id != member.id:
                    # 이미 관계가 있는지 확인
                    existing = FamilyRelationship.query.filter_by(
                        member_id=member.id,
                        related_member_id=spouse.id,
                        relationship_type='spouse'
                    ).first()
                    if not existing:
                        # 배우자 관계 생성 (양방향)
                        rel1 = FamilyRelationship(
                            member_id=member.id,
                            related_member_id=spouse.id,
                            relationship_type='spouse'
                        )
                        db.session.add(rel1)

                        # 역방향 관계도 없으면 생성
                        existing_reverse = FamilyRelationship.query.filter_by(
                            member_id=spouse.id,
                            related_member_id=member.id,
                            relationship_type='spouse'
                        ).first()
                        if not existing_reverse:
                            rel2 = FamilyRelationship(
                                member_id=spouse.id,
                                related_member_id=member.id,
                                relationship_type='spouse'
                            )
                            db.session.add(rel2)

                        family_results["created"] += 1
            except:
                pass

        # 중간 커밋 (연결 유지)
        db.session.commit()
        app.logger.info(f"[가족관계] partner_id 기반 배우자 {family_results['created']}건 생성 완료")

        # 2. family_members (ran1)로 가족 관계 생성
        members_with_family = Member.query.filter(
            Member.family_members.isnot(None),
            Member.family_members != ''
        ).all()

        app.logger.info(f"[가족관계] family_members가 있는 교인 수: {len(members_with_family)}")
        import re
        batch_count = 0

        for member in members_with_family:
            try:
                # 공백 또는 쉼표로 구분된 이름들 파싱 (예: "남궁일곤,김혜숙2" 또는 "남궁일곤 김혜숙2")
                family_names = [n.strip() for n in re.split(r'[,\s]+', member.family_members) if n.strip()]

                for name in family_names:
                    name = name.strip()
                    if not name or name == member.name:
                        continue

                    # 같은 이름의 교인 찾기 (동명이인 고려: 같은 교구/구역 우선)
                    related_candidates = Member.query.filter_by(name=name).all()

                    if not related_candidates:
                        app.logger.warning(f"[가족관계] {member.name}의 가족 '{name}'을 찾을 수 없음")
                        continue

                    # 후보가 여러 명이면 같은 교구/구역 우선
                    related = None
                    if len(related_candidates) == 1:
                        related = related_candidates[0]
                    else:
                        # 같은 교구 + 구역인 사람 우선
                        for candidate in related_candidates:
                            if candidate.district == member.district and candidate.section == member.section:
                                related = candidate
                                break
                        # 못 찾으면 같은 교구만이라도
                        if not related:
                            for candidate in related_candidates:
                                if candidate.district == member.district:
                                    related = candidate
                                    break
                        # 그래도 못 찾으면 첫 번째
                        if not related:
                            related = related_candidates[0]
                        app.logger.info(f"[가족관계] 동명이인 {len(related_candidates)}명 중 '{related.name}'({related.district}/{related.section}) 선택")

                    if related and related.id != member.id:
                        # 이미 관계가 있는지 확인 (모든 유형)
                        existing = FamilyRelationship.query.filter_by(
                            member_id=member.id,
                            related_member_id=related.id
                        ).first()
                        if not existing:
                            # 관계 유형 추론
                            rel_type = 'extended'  # 기본값
                            rel_detail = None

                            # 1. partner_id로 배우자 관계인지 확인 (가장 신뢰)
                            is_spouse = False
                            if member.partner_id and related.external_id:
                                if str(member.partner_id) == str(related.external_id):
                                    is_spouse = True
                            if related.partner_id and member.external_id:
                                if str(related.partner_id) == str(member.external_id):
                                    is_spouse = True

                            # 나이 차이 및 성별 확인
                            age_diff = None
                            is_diff_gender = False
                            member_age = None
                            related_age = None
                            from datetime import date
                            today = date.today()

                            if member.birth_date and related.birth_date:
                                age_diff = (member.birth_date - related.birth_date).days / 365
                                abs_age_diff = abs(age_diff)

                                # 각자의 나이 계산
                                member_age = (today - member.birth_date).days / 365
                                related_age = (today - related.birth_date).days / 365

                                # 성별 확인
                                m_gender = (member.gender or '').upper()
                                r_gender = (related.gender or '').upper()
                                if m_gender and r_gender:
                                    m_is_male = m_gender in ['M', '남', '남성', '남자']
                                    r_is_male = r_gender in ['M', '남', '남성', '남자']
                                    is_diff_gender = (m_is_male != r_is_male)

                            # 2. 관계 유형 결정
                            if is_spouse:
                                rel_type = 'spouse'
                            elif age_diff is not None:
                                abs_age_diff = abs(age_diff)

                                # 둘 중 한 명이라도 20세 미만이면 미성년자로 판단
                                is_minor = (member_age is not None and member_age < 20) or \
                                          (related_age is not None and related_age < 20)

                                if abs_age_diff >= 18:
                                    # 나이 차이 18년 이상 → 부모/자녀
                                    if age_diff >= 18:  # 내가 18살 이상 어리면 → 상대는 부모
                                        rel_type = 'child'
                                        rel_detail = '부모'
                                    else:  # 내가 18살 이상 많으면 → 상대는 자녀
                                        rel_type = 'parent'
                                        rel_detail = '자녀'
                                elif is_diff_gender and abs_age_diff < 5 and not is_minor:
                                    # 성별 다름 + 나이 차이 5년 미만 + 둘 다 성인 → 배우자 가능성 높음
                                    rel_type = 'spouse'
                                else:
                                    # 나이 차이 5~18년 또는 동성 또는 미성년자 → 형제자매
                                    rel_type = 'sibling'

                            rel = FamilyRelationship(
                                member_id=member.id,
                                related_member_id=related.id,
                                relationship_type=rel_type,
                                relationship_detail=rel_detail
                            )
                            db.session.add(rel)
                            family_results["created"] += 1
                        else:
                            family_results["skipped"] += 1
            except Exception as e:
                app.logger.error(f"[가족관계] {member.name} 처리 중 오류: {str(e)}")

            # 50명마다 중간 커밋 (연결 타임아웃 방지)
            batch_count += 1
            if batch_count % 50 == 0:
                db.session.commit()
                app.logger.info(f"[가족관계] ran1 처리 중간 커밋... ({batch_count}명 처리)")

        # ran1 처리 완료 커밋
        db.session.commit()
        app.logger.info(f"[가족관계] ran1 처리 완료: {family_results['created']}건 생성")

        # 3. 같은 자녀를 둔 부모들을 배우자로 연결
        # 자녀(child role)인 모든 관계를 찾아서 부모들을 그룹화
        app.logger.info("[가족관계] 부모 배우자 추론 시작...")
        child_relationships = FamilyRelationship.query.filter_by(relationship_type='child').all()

        # 자녀별 부모 목록 생성
        child_parents = {}  # {child_id: [parent_ids]}
        for rel in child_relationships:
            child_id = rel.member_id
            parent_id = rel.related_member_id
            if child_id not in child_parents:
                child_parents[child_id] = []
            child_parents[child_id].append(parent_id)

        # 같은 자녀를 둔 부모들끼리 배우자 관계 생성
        spouse_created = 0
        for child_id, parent_ids in child_parents.items():
            if len(parent_ids) >= 2:
                # 부모가 2명 이상이면 서로 배우자로 연결
                for i, p1_id in enumerate(parent_ids):
                    for p2_id in parent_ids[i+1:]:
                        # 이미 관계가 있는지 확인
                        existing = FamilyRelationship.query.filter_by(
                            member_id=p1_id,
                            related_member_id=p2_id,
                            relationship_type='spouse'
                        ).first()
                        if not existing:
                            existing_reverse = FamilyRelationship.query.filter_by(
                                member_id=p2_id,
                                related_member_id=p1_id,
                                relationship_type='spouse'
                            ).first()
                            if not existing_reverse:
                                p1 = Member.query.get(p1_id)
                                p2 = Member.query.get(p2_id)
                                if p1 and p2:
                                    # 성별에 따른 호칭 결정
                                    if p1.gender in ['M', '남', '남성', '남자']:
                                        detail1, detail2 = '남편', '아내'
                                    else:
                                        detail1, detail2 = '아내', '남편'

                                    rel1 = FamilyRelationship(
                                        member_id=p1_id,
                                        related_member_id=p2_id,
                                        relationship_type='spouse',
                                        relationship_detail=detail2
                                    )
                                    rel2 = FamilyRelationship(
                                        member_id=p2_id,
                                        related_member_id=p1_id,
                                        relationship_type='spouse',
                                        relationship_detail=detail1
                                    )
                                    db.session.add(rel1)
                                    db.session.add(rel2)
                                    spouse_created += 1
                                    app.logger.info(f"[가족관계] 부모 배우자 추론: {p1.name} ↔ {p2.name}")

        family_results["created"] += spouse_created * 2
        app.logger.info(f"[가족관계] 부모 배우자 추론 완료: {spouse_created}쌍 생성")

        db.session.commit()

        # 4. 같은 부모를 둔 자녀들을 형제로 연결
        app.logger.info("[가족관계] 형제 관계 추론 시작 (같은 부모 공유)...")

        # 부모(parent role)인 모든 관계를 찾아서 자녀들을 그룹화
        parent_relationships = FamilyRelationship.query.filter_by(relationship_type='parent').all()

        # 부모별 자녀 목록 생성
        parent_children = {}  # {parent_id: [child_ids]}
        for rel in parent_relationships:
            parent_id = rel.member_id
            child_id = rel.related_member_id
            if parent_id not in parent_children:
                parent_children[parent_id] = []
            if child_id not in parent_children[parent_id]:
                parent_children[parent_id].append(child_id)

        # 같은 부모를 둔 자녀들끼리 형제 관계 생성
        sibling_created = 0
        for parent_id, child_ids in parent_children.items():
            if len(child_ids) >= 2:
                # 자녀가 2명 이상이면 서로 형제로 연결
                for i, c1_id in enumerate(child_ids):
                    for c2_id in child_ids[i+1:]:
                        # 이미 관계가 있는지 확인
                        existing = FamilyRelationship.query.filter_by(
                            member_id=c1_id,
                            related_member_id=c2_id,
                            relationship_type='sibling'
                        ).first()
                        if not existing:
                            existing_reverse = FamilyRelationship.query.filter_by(
                                member_id=c2_id,
                                related_member_id=c1_id,
                                relationship_type='sibling'
                            ).first()
                            if not existing_reverse:
                                c1 = Member.query.get(c1_id)
                                c2 = Member.query.get(c2_id)
                                if c1 and c2:
                                    # 성별과 나이에 따른 호칭 결정
                                    detail1, detail2 = '형제', '형제'  # 기본값
                                    c1_male = c1.gender in ['M', '남', '남성', '남자']
                                    c2_male = c2.gender in ['M', '남', '남성', '남자']
                                    c1_older = True
                                    if c1.birth_date and c2.birth_date:
                                        c1_older = c1.birth_date < c2.birth_date

                                    if c1_male and c2_male:
                                        detail1 = '형' if not c1_older else '동생'
                                        detail2 = '형' if c1_older else '동생'
                                    elif not c1_male and not c2_male:
                                        detail1 = '언니' if not c1_older else '동생'
                                        detail2 = '언니' if c1_older else '동생'
                                    elif c1_male:  # c1 남, c2 여
                                        detail1 = '오빠' if not c1_older else '남동생'
                                        detail2 = '누나' if c1_older else '여동생'
                                    else:  # c1 여, c2 남
                                        detail1 = '누나' if not c1_older else '여동생'
                                        detail2 = '오빠' if c1_older else '남동생'

                                    rel1 = FamilyRelationship(
                                        member_id=c1_id,
                                        related_member_id=c2_id,
                                        relationship_type='sibling',
                                        relationship_detail=detail2
                                    )
                                    rel2 = FamilyRelationship(
                                        member_id=c2_id,
                                        related_member_id=c1_id,
                                        relationship_type='sibling',
                                        relationship_detail=detail1
                                    )
                                    db.session.add(rel1)
                                    db.session.add(rel2)
                                    sibling_created += 1
                                    app.logger.info(f"[가족관계] 형제 추론: {c1.name} ↔ {c2.name}")

        family_results["created"] += sibling_created * 2
        app.logger.info(f"[가족관계] 형제 추론 완료: {sibling_created}쌍 생성")

        db.session.commit()
        results["family_created"] = family_results["created"]

        # 세션 만료로 부분 동기화된 경우
        if session_expired:
            return jsonify({
                "success": True,
                "message": f"부분 동기화 완료 (세션 만료): {results['created']}명 생성, {results['updated']}명 업데이트",
                "results": results,
                "warning": "god4u 세션이 만료되어 일부 데이터만 동기화되었습니다."
            })

        return jsonify({
            "success": True,
            "message": f"동기화 완료: {results['created']}명 생성, {results['updated']}명 업데이트, 가족관계 {results.get('family_created', 0)}건 생성",
            "results": results
        })

    except Exception as e:
        return jsonify({"error": f"동기화 실패: {str(e)}"}), 500


@app.route('/api/sync/registry-to-god4u', methods=['POST'])
def api_sync_registry_to_god4u():
    """church-registry → god4u 동기화 API"""
    import time

    data = request.get_json() or {}
    cookies = data.get('cookies', {})

    if not cookies.get('ASP.NET_SessionId') or not cookies.get('pastorinfo'):
        return jsonify({"error": "god4u 쿠키가 필요합니다"}), 400

    results = {"success": 0, "failed": 0, "skipped": 0}

    # external_id가 있는 회원만 동기화
    members = Member.query.filter(Member.external_id.isnot(None)).all()

    session = requests.Session()
    session.cookies.update(cookies)

    for member in members:
        try:
            success = _sync_member_to_god4u(session, member)
            if success:
                results["success"] += 1
            else:
                results["failed"] += 1
        except:
            results["failed"] += 1

        time.sleep(0.3)

    return jsonify({
        "success": True,
        "message": f"god4u 동기화 완료: {results['success']}명 성공",
        "results": results
    })


def _create_god4u_payload(page=1, page_size=100):
    """god4u API 페이로드 생성"""
    return {
        "paramName": "", "paramEName": "", "paramIds": "",
        "paramFree1": "", "paramFree2": "", "paramFree3": "", "paramFree4": "",
        "paramFree5": "", "paramFree6": "", "paramFree7": "", "paramFree8": "",
        "paramFree9": "", "paramFree10": "", "paramFree11": "", "paramFree12": "",
        "paramRange": "", "paramRange1": "", "paramRange2": "", "paramRange3": "",
        "paramRvname": "", "paramSection1": "", "paramSection2": "",
        "paramSection3": "", "paramSection4": "", "paramRvname2": "",
        "paramCoreChk": "", "paramCarNum": "", "paramGJeon": "",
        "paramLastSchool": "", "paramOffName": "", "paramGJeon1": "",
        "paramCvname": "", "paramCvname1": "", "paramState": "",
        "paramState1": "", "paramState3": "", "encryptOpt": "ALL",
        "rangeLimitUse": "false", "paramPage": str(page),
        "paramPageSize": str(page_size), "paramOrder": "NAME",
        "paramOrder2": "", "paramOrderAsc": "ASC", "paramOrder2Asc": "ASC",
        "paramPType": "P", "paramAddr": "", "paramRegDateS": "", "paramRegDateE": "",
    }


def _sync_member_to_god4u(session, member):
    """개별 회원을 god4u로 동기화"""
    if not member.external_id:
        return False

    # 주소 파싱
    address = member.address or ""
    parts = address.split() if address else []
    sido = parts[0] if len(parts) >= 1 else ""
    gugun = parts[1] if len(parts) >= 2 else ""
    dong = parts[2] if len(parts) >= 3 else ""
    bunji = " ".join(parts[3:]) if len(parts) >= 4 else ""

    gender = "남" if member.gender == "M" else "여" if member.gender == "F" else ""

    payload = {
        "mode": "mod",
        "hidIdM": member.external_id,
        "txtHidM": member.external_id,
        "txtNameM": member.name or "",
        "txtHandphoneM": member.phone or "",
        "txtTelM": "",
        "txtEmailM": member.email or "",
        "txtBirthDayM": member.birth_date.isoformat() if member.birth_date else "",
        "ddlGenderM": gender,
        "txtSidoM": sido,
        "txtGugunM": gugun,
        "txtDongM": dong,
        "txtBunjiM": bunji,
        "ddlState3": "예배출석" if member.status == "active" else "결석",
        "txtRegDayM": member.registration_date.isoformat() if member.registration_date else "",
        "hidcvname": member.member_type or "",
        "hidcvname1": "",
        "hidstate": "교인",
        "hidstate1": "장년",
        "hidRange": "", "hidRange1": "", "hidRange2": "", "hidRange3": "",
        "txtENameF": "", "txtENameM": "", "txtENameL": "",
        "ddlSolarM": "양", "txtAgeM": "",
        "txtCoreM": member.name or "",
        "ddlRelative": "본인",
        "txtZipcodeM": "", "txtZipcodeMOrg": "",
        "txtSidoMOrg": "", "txtGugunMOrg": "", "txtDongMOrg": "", "txtBunjiMOrg": "",
        "txtCityM": "", "txtStM": "", "txtCityMOrg": "", "txtStMOrg": "",
        "ddlGYear": "2026",
        "txtRangeOrg": "", "txtRange1Org": "", "txtRange2Org": "", "txtRange3Org": "",
        "ddlCvAct": "", "txtCvDay": "", "txtAppointChurch": "",
        "txtCvnameOrg": "", "txtCvname1Org": "", "txtCvActorg": "",
        "txtCvDayOrg": "", "txtAppointChurchOrg": "",
        "txtStateDay": "", "txtStateOrg": "교인", "txtState1Org": "장년",
        "txtState3Org": "예배출석", "txtStateDayOrg": "",
        "hidvaccinename": "", "hidvaccinenumber": "",
        "txtVaccineDate": "", "txtVaccineNameOrg": "", "txtVaccineNumberOrg": "",
        "txtVaccineDateOrg": "", "txtVaccineIndex": "0",
        "txtLeaderidM": "0", "txtLeaderM": "",
        "txtOffnameM": "", "txtOfftelM": "",
        "ddlGrade": "", "txtBaptday": "", "txtBaptchurch": "", "txtBaptist": "",
        "txtGradeOrg": "", "txtBaptdayOrg": "", "txtBaptchurchOrg": "", "txtBaptistOrg": "",
        "txtCarKind": "", "txtCarNum": "", "txtCarKind1": "", "txtCarNum1": "",
        "txPrechurchM": "", "txtEtcM": member.notes or "",
        "txtFree1": "", "txtFree2": "", "txtFree3": "", "txtFree4": "",
        "txtFree5": "", "txtFree6": "", "txtFree7": "", "txtFree8": "",
        "txtFree9": "", "txtFree10": "", "txtFree11": "", "txtFree12": "",
    }

    headers = {
        "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
        "Accept": "text/html, */*; q=0.01",
        "X-Requested-With": "XMLHttpRequest",
        "Origin": GOD4U_BASE_URL,
        "Referer": f"{GOD4U_BASE_URL}/WebMobile/WebChurch/PersonModifyDetail.cshtml?id={member.external_id}",
    }

    try:
        response = session.post(GOD4U_UPDATE_URL, data=payload, headers=headers, timeout=30)
        return response.status_code == 200 and "정보수정 완료" in response.text
    except:
        return False


# =============================================================================
# 데이터베이스 초기화 및 마이그레이션
# =============================================================================

def run_migrations():
    """데이터베이스 스키마 마이그레이션 실행"""
    from sqlalchemy import text, inspect

    inspector = inspect(db.engine)

    # members 테이블 마이그레이션
    if 'members' in inspector.get_table_names():
        columns = [col['name'] for col in inspector.get_columns('members')]

        # external_id 컬럼 추가 (god4u 등 외부 시스템 연동용)
        if 'external_id' not in columns:
            db.session.execute(text(
                'ALTER TABLE members ADD COLUMN external_id VARCHAR(50)'
            ))
            print('[Migration] Added external_id column to members table')
            db.session.commit()

        # car_number 컬럼 추가 (차량번호)
        if 'car_number' not in columns:
            db.session.execute(text(
                'ALTER TABLE members ADD COLUMN car_number VARCHAR(20)'
            ))
            print('[Migration] Added car_number column to members table')
            db.session.commit()

        # family_members 컬럼 추가 (가족 정보 텍스트)
        if 'family_members' not in columns:
            db.session.execute(text(
                'ALTER TABLE members ADD COLUMN family_members VARCHAR(200)'
            ))
            print('[Migration] Added family_members column to members table')
            db.session.commit()

    # groups 테이블 마이그레이션
    if 'groups' in inspector.get_table_names():
        columns = [col['name'] for col in inspector.get_columns('groups')]

        # parent_id 컬럼 추가
        if 'parent_id' not in columns:
            db.session.execute(text(
                'ALTER TABLE groups ADD COLUMN parent_id INTEGER REFERENCES groups(id)'
            ))
            print('[Migration] Added parent_id column to groups table')

        # level 컬럼 추가
        if 'level' not in columns:
            db.session.execute(text(
                'ALTER TABLE groups ADD COLUMN level INTEGER DEFAULT 0'
            ))
            print('[Migration] Added level column to groups table')

        # description 컬럼 추가
        if 'description' not in columns:
            db.session.execute(text(
                'ALTER TABLE groups ADD COLUMN description VARCHAR(200)'
            ))
            print('[Migration] Added description column to groups table')

        db.session.commit()

with app.app_context():
    db.create_all()
    run_migrations()


if __name__ == '__main__':
    app.run(debug=True, port=5000)
