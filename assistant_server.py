"""
Personal AI Assistant - 메인 대시보드 서버
Mac(iCloud) 일정/미리알림과 웹 서버(DB) 데이터를 통합 관리하는 허브
"""
import os
import io
import csv
import json
import re
from datetime import datetime, date, timedelta
from flask import Blueprint, request, jsonify, render_template


def parse_korean_datetime(datetime_str):
    """
    한국어 날짜/시간 형식을 ISO timestamp로 변환
    예: "2025. 12. 3. 오전 9:00" -> "2025-12-03T09:00:00"
        "2025. 12. 3. 오후 2:30" -> "2025-12-03T14:30:00"
    """
    if not datetime_str:
        print(f"[DATE-PARSER] 빈 값 수신")
        return None

    # 문자열 정리 (앞뒤 공백 및 대괄호 제거)
    datetime_str = str(datetime_str).strip()
    datetime_str = datetime_str.strip('[]')  # 대괄호 제거
    datetime_str = datetime_str.strip()  # 대괄호 제거 후 남은 공백 제거

    if not datetime_str:
        print(f"[DATE-PARSER] 정리 후 빈 값")
        return None

    print(f"[DATE-PARSER] 입력값 (정리 후): '{datetime_str}'")

    # 이미 ISO 형식이면 그대로 반환 (timezone 포함 가능)
    if re.match(r'^\d{4}-\d{2}-\d{2}T', datetime_str):
        # timezone 제거 후 반환
        result = datetime_str[:19]  # "2025-12-03T09:00:00" 까지만
        print(f"[DATE-PARSER] ISO 형식 감지: {result}")
        return result

    # 한국어 날짜/시간 형식 파싱: "2025. 12. 3. 오전 9:00" 또는 "2025.12.3. 오전 9:00"
    match = re.match(r'(\d{4})\.\s*(\d{1,2})\.\s*(\d{1,2})\.\s*(오전|오후)\s*(\d{1,2}):(\d{2})', datetime_str)
    if match:
        year, month, day, ampm, hour, minute = match.groups()
        hour = int(hour)
        if ampm == '오후' and hour != 12:
            hour += 12
        elif ampm == '오전' and hour == 12:
            hour = 0
        result = f"{year}-{int(month):02d}-{int(day):02d}T{hour:02d}:{minute}:00"
        print(f"[DATE-PARSER] 한국어 형식 파싱 성공: {result}")
        return result

    # "2025년 12월 3일 오전 9:00" 형식
    match = re.match(r'(\d{4})년\s*(\d{1,2})월\s*(\d{1,2})일\s*(오전|오후)\s*(\d{1,2}):(\d{2})', datetime_str)
    if match:
        year, month, day, ampm, hour, minute = match.groups()
        hour = int(hour)
        if ampm == '오후' and hour != 12:
            hour += 12
        elif ampm == '오전' and hour == 12:
            hour = 0
        result = f"{year}-{int(month):02d}-{int(day):02d}T{hour:02d}:{minute}:00"
        print(f"[DATE-PARSER] 한국어 '년월일' 형식 파싱 성공: {result}")
        return result

    # 날짜만 있는 경우: "2025. 12. 3." 또는 "2025.12.3"
    match = re.match(r'(\d{4})\.\s*(\d{1,2})\.\s*(\d{1,2})\.?', datetime_str)
    if match:
        year, month, day = match.groups()
        result = f"{year}-{int(month):02d}-{int(day):02d}T00:00:00"
        print(f"[DATE-PARSER] 날짜만 파싱 성공: {result}")
        return result

    # "2025년 12월 3일" 형식 (시간 없음)
    match = re.match(r'(\d{4})년\s*(\d{1,2})월\s*(\d{1,2})일', datetime_str)
    if match:
        year, month, day = match.groups()
        result = f"{year}-{int(month):02d}-{int(day):02d}T00:00:00"
        print(f"[DATE-PARSER] 한국어 '년월일' 날짜만 파싱 성공: {result}")
        return result

    # "2025-12-03 09:00:00" 형식 (ISO-like with space)
    match = re.match(r'(\d{4})-(\d{2})-(\d{2})\s+(\d{2}):(\d{2}):?(\d{2})?', datetime_str)
    if match:
        year, month, day, hour, minute, second = match.groups()
        second = second or '00'
        result = f"{year}-{month}-{day}T{hour}:{minute}:{second}"
        print(f"[DATE-PARSER] ISO-like 형식 파싱 성공: {result}")
        return result

    print(f"[DATE-PARSER] 파싱 실패: '{datetime_str}'")
    return None


def parse_korean_date(date_str):
    """
    한국어 날짜 형식을 ISO 형식으로 변환
    예: "2025. 12. 4. 오전 12:00" -> "2025-12-04"
        "2025. 12. 4." -> "2025-12-04"
        "2025-12-04" -> "2025-12-04" (이미 ISO 형식)
    """
    if not date_str:
        return None

    # 이미 ISO 형식이면 그대로 반환
    if re.match(r'^\d{4}-\d{2}-\d{2}', date_str):
        return date_str[:10]

    # 한국어 날짜 형식 파싱: "2025. 12. 4. 오전 12:00" 또는 "2025. 12. 4."
    match = re.match(r'(\d{4})\.\s*(\d{1,2})\.\s*(\d{1,2})\.?', date_str)
    if match:
        year, month, day = match.groups()
        return f"{year}-{int(month):02d}-{int(day):02d}"

    # 다른 형식 시도: "12/4/2025" 등
    try:
        for fmt in ['%Y-%m-%d', '%Y/%m/%d', '%m/%d/%Y', '%d/%m/%Y']:
            try:
                parsed = datetime.strptime(date_str.split()[0], fmt)
                return parsed.strftime('%Y-%m-%d')
            except ValueError:
                continue
    except Exception:
        pass

    return None

# OpenAI (GPT-4o-mini for parsing)
from openai import OpenAI

def get_openai_client():
    key = (os.getenv("OPENAI_API_KEY") or "").strip()
    if not key:
        return None
    return OpenAI(api_key=key)

assistant_bp = Blueprint('assistant', __name__)

# ===== 데이터베이스 설정 =====
DATABASE_URL = os.getenv('DATABASE_URL')
USE_POSTGRES = DATABASE_URL is not None

if USE_POSTGRES:
    import psycopg2
    from psycopg2.extras import RealDictCursor

    # Render의 postgres:// URL을 postgresql://로 변경
    if DATABASE_URL.startswith('postgres://'):
        DATABASE_URL = DATABASE_URL.replace('postgres://', 'postgresql://', 1)

    def get_db_connection():
        conn = psycopg2.connect(DATABASE_URL, cursor_factory=RealDictCursor)
        return conn
else:
    import sqlite3
    DB_PATH = os.path.join(os.path.dirname(__file__), 'assistant_data.db')

    def get_db_connection():
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        return conn


# ===== 데이터베이스 초기화 =====
def init_assistant_db():
    """Assistant DB 테이블 초기화"""
    conn = get_db_connection()
    cursor = conn.cursor()

    if USE_POSTGRES:
        # PostgreSQL 스키마
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS events (
                id SERIAL PRIMARY KEY,
                title VARCHAR(500) NOT NULL,
                start_time TIMESTAMP,
                end_time TIMESTAMP,
                category VARCHAR(100),
                location VARCHAR(200),
                notes TEXT,
                source VARCHAR(50) DEFAULT 'web',
                sync_status VARCHAR(50) DEFAULT 'pending_to_mac',
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')

        # 기존 테이블에 location, notes, gcal_id 컬럼 추가 (이미 존재하면 무시)
        try:
            cursor.execute('ALTER TABLE events ADD COLUMN IF NOT EXISTS location VARCHAR(200)')
            cursor.execute('ALTER TABLE events ADD COLUMN IF NOT EXISTS notes TEXT')
            cursor.execute('ALTER TABLE events ADD COLUMN IF NOT EXISTS gcal_id VARCHAR(200)')
        except Exception:
            pass  # 컬럼이 이미 존재하면 무시

        cursor.execute('''
            CREATE TABLE IF NOT EXISTS tasks (
                id SERIAL PRIMARY KEY,
                title VARCHAR(500) NOT NULL,
                due_date DATE,
                priority VARCHAR(20) DEFAULT 'medium',
                category VARCHAR(100),
                is_completed BOOLEAN DEFAULT FALSE,
                source VARCHAR(50) DEFAULT 'web',
                sync_status VARCHAR(50) DEFAULT 'pending_to_mac',
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')

        cursor.execute('''
            CREATE TABLE IF NOT EXISTS attendance (
                id SERIAL PRIMARY KEY,
                name VARCHAR(200) NOT NULL,
                date DATE NOT NULL,
                status VARCHAR(20) DEFAULT 'present',
                group_name VARCHAR(100),
                source VARCHAR(50) DEFAULT 'web',
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')

        # 기존 테이블에 group_name 컬럼 추가
        try:
            cursor.execute('ALTER TABLE attendance ADD COLUMN IF NOT EXISTS group_name VARCHAR(100)')
        except Exception:
            pass

        # People 테이블 (인물 관리)
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS people (
                id SERIAL PRIMARY KEY,
                name VARCHAR(200) NOT NULL,
                category VARCHAR(100),
                phone VARCHAR(50),
                email VARCHAR(200),
                address TEXT,
                birthday DATE,
                notes TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')

        # birthday 컬럼 추가 (기존 테이블 마이그레이션)
        try:
            cursor.execute('ALTER TABLE people ADD COLUMN IF NOT EXISTS birthday DATE')
        except:
            pass  # 이미 존재하면 무시

        # People Notes 테이블 (인물별 누적 기록)
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS people_notes (
                id SERIAL PRIMARY KEY,
                person_id INTEGER REFERENCES people(id) ON DELETE CASCADE,
                note_date DATE NOT NULL,
                content TEXT NOT NULL,
                category VARCHAR(100),
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')

        # Projects 테이블 (프로젝트 관리)
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS projects (
                id SERIAL PRIMARY KEY,
                name VARCHAR(300) NOT NULL,
                description TEXT,
                status VARCHAR(50) DEFAULT 'active',
                start_date DATE,
                end_date DATE,
                priority VARCHAR(20) DEFAULT 'medium',
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')

        # Project Notes 테이블 (프로젝트별 누적 기록)
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS project_notes (
                id SERIAL PRIMARY KEY,
                project_id INTEGER REFERENCES projects(id) ON DELETE CASCADE,
                note_date DATE NOT NULL,
                content TEXT NOT NULL,
                category VARCHAR(100),
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')

    else:
        # SQLite 스키마
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS events (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                title TEXT NOT NULL,
                start_time DATETIME,
                end_time DATETIME,
                category TEXT,
                location TEXT,
                notes TEXT,
                source TEXT DEFAULT 'web',
                sync_status TEXT DEFAULT 'pending_to_mac',
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                updated_at DATETIME DEFAULT CURRENT_TIMESTAMP
            )
        ''')

        # 기존 테이블에 location, notes, gcal_id 컬럼 추가 (SQLite는 IF NOT EXISTS 미지원)
        try:
            cursor.execute('ALTER TABLE events ADD COLUMN location TEXT')
        except Exception:
            pass
        try:
            cursor.execute('ALTER TABLE events ADD COLUMN notes TEXT')
        except Exception:
            pass
        try:
            cursor.execute('ALTER TABLE events ADD COLUMN gcal_id TEXT')
        except Exception:
            pass

        cursor.execute('''
            CREATE TABLE IF NOT EXISTS tasks (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                title TEXT NOT NULL,
                due_date DATE,
                priority TEXT DEFAULT 'medium',
                category TEXT,
                is_completed INTEGER DEFAULT 0,
                source TEXT DEFAULT 'web',
                sync_status TEXT DEFAULT 'pending_to_mac',
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                updated_at DATETIME DEFAULT CURRENT_TIMESTAMP
            )
        ''')

        cursor.execute('''
            CREATE TABLE IF NOT EXISTS attendance (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                date DATE NOT NULL,
                status TEXT DEFAULT 'present',
                group_name TEXT,
                source TEXT DEFAULT 'web',
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP
            )
        ''')

        # 기존 테이블에 group_name 컬럼 추가 (SQLite는 IF NOT EXISTS 미지원)
        try:
            cursor.execute('ALTER TABLE attendance ADD COLUMN group_name TEXT')
        except Exception:
            pass

        # People 테이블 (인물 관리)
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS people (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                category TEXT,
                phone TEXT,
                email TEXT,
                address TEXT,
                birthday DATE,
                notes TEXT,
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                updated_at DATETIME DEFAULT CURRENT_TIMESTAMP
            )
        ''')

        # birthday 컬럼 추가 (기존 테이블 마이그레이션)
        try:
            cursor.execute('ALTER TABLE people ADD COLUMN birthday DATE')
        except Exception:
            pass  # 이미 존재하면 무시

        # People Notes 테이블 (인물별 누적 기록)
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS people_notes (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                person_id INTEGER NOT NULL,
                note_date DATE NOT NULL,
                content TEXT NOT NULL,
                category TEXT,
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (person_id) REFERENCES people(id) ON DELETE CASCADE
            )
        ''')

        # Projects 테이블 (프로젝트 관리)
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS projects (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                description TEXT,
                status TEXT DEFAULT 'active',
                start_date DATE,
                end_date DATE,
                priority TEXT DEFAULT 'medium',
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                updated_at DATETIME DEFAULT CURRENT_TIMESTAMP
            )
        ''')

        # Project Notes 테이블 (프로젝트별 누적 기록)
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS project_notes (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                project_id INTEGER NOT NULL,
                note_date DATE NOT NULL,
                content TEXT NOT NULL,
                category TEXT,
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (project_id) REFERENCES projects(id) ON DELETE CASCADE
            )
        ''')

    conn.commit()
    conn.close()
    print("[ASSISTANT-DB] 테이블 초기화 완료")


# ===== 메인 대시보드 라우트 =====
@assistant_bp.route('/assistant')
def assistant_dashboard():
    """메인 대시보드 페이지 렌더링"""
    return render_template('assistant.html')


# ===== 이벤트 API =====
@assistant_bp.route('/assistant/api/events', methods=['GET'])
def get_events():
    """이벤트 목록 조회 (start, end 파라미터로 기간 지정 가능)"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        # 쿼리 파라미터에서 날짜 범위 가져오기
        start_date_str = request.args.get('start')
        end_date_str = request.args.get('end')

        today = date.today()

        if start_date_str and end_date_str:
            # 지정된 기간으로 조회
            start_date = date.fromisoformat(start_date_str)
            end_date = date.fromisoformat(end_date_str)
        else:
            # 기본: 현재 달 전체 + 앞뒤 1주씩
            start_date = today.replace(day=1) - timedelta(days=7)
            # 다음 달 1일 - 1일 = 이번 달 마지막 날
            if today.month == 12:
                end_date = date(today.year + 1, 1, 1) + timedelta(days=7)
            else:
                end_date = date(today.year, today.month + 1, 1) + timedelta(days=7)

        if USE_POSTGRES:
            cursor.execute('''
                SELECT * FROM events
                WHERE DATE(start_time) >= %s AND DATE(start_time) <= %s
                ORDER BY start_time ASC
            ''', (start_date, end_date))
        else:
            cursor.execute('''
                SELECT * FROM events
                WHERE DATE(start_time) >= ? AND DATE(start_time) <= ?
                ORDER BY start_time ASC
            ''', (start_date.isoformat(), end_date.isoformat()))

        events = cursor.fetchall()
        conn.close()

        # Row 객체를 dict로 변환
        events_list = [dict(row) for row in events]

        # datetime 객체를 문자열로 변환
        for event in events_list:
            for key in ['start_time', 'end_time', 'created_at', 'updated_at']:
                if event.get(key) and not isinstance(event[key], str):
                    event[key] = event[key].isoformat()

        return jsonify({
            'success': True,
            'events': events_list,
            'start': start_date.isoformat(),
            'end': end_date.isoformat()
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@assistant_bp.route('/assistant/api/events', methods=['POST'])
def create_event():
    """새 이벤트 생성"""
    try:
        data = request.get_json()

        conn = get_db_connection()
        cursor = conn.cursor()

        if USE_POSTGRES:
            cursor.execute('''
                INSERT INTO events (title, start_time, end_time, category, source, sync_status)
                VALUES (%s, %s, %s, %s, %s, %s)
                RETURNING id
            ''', (
                data.get('title'),
                data.get('start_time'),
                data.get('end_time'),
                data.get('category', ''),
                data.get('source', 'web'),
                'pending_to_mac'
            ))
            event_id = cursor.fetchone()['id']
        else:
            cursor.execute('''
                INSERT INTO events (title, start_time, end_time, category, source, sync_status)
                VALUES (?, ?, ?, ?, ?, ?)
            ''', (
                data.get('title'),
                data.get('start_time'),
                data.get('end_time'),
                data.get('category', ''),
                data.get('source', 'web'),
                'pending_to_mac'
            ))
            event_id = cursor.lastrowid

        conn.commit()
        conn.close()

        return jsonify({
            'success': True,
            'id': event_id,
            'message': '이벤트가 생성되었습니다.'
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@assistant_bp.route('/assistant/api/events/<int:event_id>', methods=['GET'])
def get_event(event_id):
    """특정 이벤트 조회"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        if USE_POSTGRES:
            cursor.execute('SELECT * FROM events WHERE id = %s', (event_id,))
        else:
            cursor.execute('SELECT * FROM events WHERE id = ?', (event_id,))

        event = cursor.fetchone()
        conn.close()

        if event:
            return jsonify({'success': True, 'event': dict(event)})
        else:
            return jsonify({'success': False, 'error': '이벤트를 찾을 수 없습니다.'}), 404
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@assistant_bp.route('/assistant/api/events/<int:event_id>', methods=['PUT'])
def update_event(event_id):
    """이벤트 수정"""
    try:
        data = request.get_json()
        conn = get_db_connection()
        cursor = conn.cursor()

        if USE_POSTGRES:
            cursor.execute('''
                UPDATE events SET title = %s, start_time = %s, end_time = %s, category = %s
                WHERE id = %s
            ''', (data.get('title'), data.get('start_time'), data.get('end_time'), data.get('category', ''), event_id))
        else:
            cursor.execute('''
                UPDATE events SET title = ?, start_time = ?, end_time = ?, category = ?
                WHERE id = ?
            ''', (data.get('title'), data.get('start_time'), data.get('end_time'), data.get('category', ''), event_id))

        conn.commit()
        conn.close()
        return jsonify({'success': True, 'message': '이벤트가 수정되었습니다.'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@assistant_bp.route('/assistant/api/events/<int:event_id>', methods=['DELETE'])
def delete_event(event_id):
    """이벤트 삭제"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        if USE_POSTGRES:
            cursor.execute('DELETE FROM events WHERE id = %s', (event_id,))
        else:
            cursor.execute('DELETE FROM events WHERE id = ?', (event_id,))

        conn.commit()
        conn.close()
        return jsonify({'success': True, 'message': '이벤트가 삭제되었습니다.'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ===== 태스크 API =====
@assistant_bp.route('/assistant/api/tasks', methods=['GET'])
def get_tasks():
    """태스크 목록 조회 (all=true 파라미터로 전체 태스크 조회)"""
    try:
        include_all = request.args.get('all', 'true').lower() == 'true'

        conn = get_db_connection()
        cursor = conn.cursor()

        if include_all:
            # 전체 태스크 조회 (완료/미완료 모두)
            if USE_POSTGRES:
                cursor.execute('''
                    SELECT * FROM tasks
                    ORDER BY is_completed ASC, due_date ASC NULLS LAST, priority DESC
                ''')
            else:
                cursor.execute('''
                    SELECT * FROM tasks
                    ORDER BY is_completed ASC, due_date ASC, priority DESC
                ''')
        else:
            # 미완료 태스크만 조회
            if USE_POSTGRES:
                cursor.execute('''
                    SELECT * FROM tasks
                    WHERE is_completed = FALSE
                    ORDER BY due_date ASC NULLS LAST, priority DESC
                ''')
            else:
                cursor.execute('''
                    SELECT * FROM tasks
                    WHERE is_completed = 0
                    ORDER BY due_date ASC, priority DESC
                ''')

        tasks = cursor.fetchall()
        conn.close()

        tasks_list = [dict(row) for row in tasks]

        # date 객체를 문자열로 변환 및 status 필드 추가
        for task in tasks_list:
            for key in ['due_date', 'created_at', 'updated_at']:
                if task.get(key) and not isinstance(task[key], str):
                    task[key] = task[key].isoformat() if hasattr(task[key], 'isoformat') else str(task[key])
            # status 필드 추가 (프론트엔드 호환)
            task['status'] = 'completed' if task.get('is_completed') else 'pending'

        return jsonify({
            'success': True,
            'tasks': tasks_list
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@assistant_bp.route('/assistant/api/tasks', methods=['POST'])
def create_task():
    """새 태스크 생성"""
    try:
        data = request.get_json()

        conn = get_db_connection()
        cursor = conn.cursor()

        if USE_POSTGRES:
            cursor.execute('''
                INSERT INTO tasks (title, due_date, priority, category, source, sync_status)
                VALUES (%s, %s, %s, %s, %s, %s)
                RETURNING id
            ''', (
                data.get('title'),
                data.get('due_date'),
                data.get('priority', 'medium'),
                data.get('category', ''),
                data.get('source', 'web'),
                'pending_to_mac'
            ))
            task_id = cursor.fetchone()['id']
        else:
            cursor.execute('''
                INSERT INTO tasks (title, due_date, priority, category, source, sync_status)
                VALUES (?, ?, ?, ?, ?, ?)
            ''', (
                data.get('title'),
                data.get('due_date'),
                data.get('priority', 'medium'),
                data.get('category', ''),
                data.get('source', 'web'),
                'pending_to_mac'
            ))
            task_id = cursor.lastrowid

        conn.commit()
        conn.close()

        return jsonify({
            'success': True,
            'id': task_id,
            'message': '태스크가 생성되었습니다.'
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@assistant_bp.route('/assistant/api/tasks/<int:task_id>/complete', methods=['POST'])
def complete_task(task_id):
    """태스크 완료 처리"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        if USE_POSTGRES:
            cursor.execute('''
                UPDATE tasks SET is_completed = TRUE, updated_at = CURRENT_TIMESTAMP
                WHERE id = %s
            ''', (task_id,))
        else:
            cursor.execute('''
                UPDATE tasks SET is_completed = 1, updated_at = CURRENT_TIMESTAMP
                WHERE id = ?
            ''', (task_id,))

        conn.commit()
        conn.close()

        return jsonify({
            'success': True,
            'message': '태스크가 완료되었습니다.'
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@assistant_bp.route('/assistant/api/tasks/<int:task_id>/uncomplete', methods=['POST'])
def uncomplete_task(task_id):
    """태스크 완료 취소"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        if USE_POSTGRES:
            cursor.execute('''
                UPDATE tasks SET is_completed = FALSE, updated_at = CURRENT_TIMESTAMP
                WHERE id = %s
            ''', (task_id,))
        else:
            cursor.execute('''
                UPDATE tasks SET is_completed = 0, updated_at = CURRENT_TIMESTAMP
                WHERE id = ?
            ''', (task_id,))

        conn.commit()
        conn.close()

        return jsonify({
            'success': True,
            'message': '태스크 완료가 취소되었습니다.'
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@assistant_bp.route('/assistant/api/tasks/<int:task_id>', methods=['DELETE'])
def delete_task(task_id):
    """태스크 삭제"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        if USE_POSTGRES:
            cursor.execute('DELETE FROM tasks WHERE id = %s', (task_id,))
        else:
            cursor.execute('DELETE FROM tasks WHERE id = ?', (task_id,))

        conn.commit()
        conn.close()

        return jsonify({
            'success': True,
            'message': '태스크가 삭제되었습니다.'
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ===== Mac 연동 API (Shortcuts 호출용) =====
@assistant_bp.route('/assistant/api/sync/health', methods=['GET', 'POST'])
def sync_health():
    """Mac 단축어 연결 테스트용 health check"""
    return jsonify({
        'success': True,
        'message': 'Server is running',
        'timestamp': datetime.now().isoformat()
    })


@assistant_bp.route('/assistant/api/sync/from-mac', methods=['POST'])
def sync_from_mac():
    """
    Mac 단축어에서 Calendar/Reminders 데이터를 받아 저장
    Body: { "events": [...], "tasks": [...] }
    """
    try:
        # 요청 정보 로깅
        print(f"[SYNC-FROM-MAC] Content-Type: {request.content_type}")
        print(f"[SYNC-FROM-MAC] Content-Length: {request.content_length}")

        # JSON 파싱 시도 (여러 방법으로)
        data = None
        raw_data = None

        try:
            # 방법 1: get_json() 사용
            data = request.get_json(force=True, silent=True)
        except Exception as json_error:
            print(f"[SYNC-FROM-MAC] get_json 실패: {json_error}")

        if not data:
            # 방법 2: 원본 데이터에서 직접 파싱
            try:
                raw_data = request.get_data(as_text=True)
                print(f"[SYNC-FROM-MAC] 원본 데이터 (처음 500자): {raw_data[:500] if raw_data else 'EMPTY'}")
                if raw_data:
                    data = json.loads(raw_data)
            except Exception as raw_error:
                print(f"[SYNC-FROM-MAC] 원본 데이터 파싱 실패: {raw_error}")

        if not data:
            return jsonify({
                'success': False,
                'error': 'JSON 파싱 실패. Content-Type을 application/json으로 설정하세요.',
                'received_content_type': request.content_type,
                'raw_preview': raw_data[:200] if raw_data else None
            }), 400

        # 디버그 로그: 수신된 원본 데이터 출력
        print(f"[SYNC-FROM-MAC] 수신된 원본 데이터: {json.dumps(data, ensure_ascii=False, indent=2)}")

        conn = get_db_connection()
        cursor = conn.cursor()

        events_added = 0
        tasks_added = 0

        # 기존 mac_calendar 소스 이벤트 삭제 (전체 동기화)
        if data.get('events'):
            if USE_POSTGRES:
                cursor.execute("DELETE FROM events WHERE source = %s", ('mac_calendar',))
            else:
                cursor.execute("DELETE FROM events WHERE source = ?", ('mac_calendar',))
            print(f"[SYNC-FROM-MAC] 기존 mac_calendar 이벤트 삭제 완료")

        # 기존 mac_reminders 소스 태스크 삭제 (전체 동기화)
        if data.get('tasks'):
            if USE_POSTGRES:
                cursor.execute("DELETE FROM tasks WHERE source = %s", ('mac_reminders',))
            else:
                cursor.execute("DELETE FROM tasks WHERE source = ?", ('mac_reminders',))
            print(f"[SYNC-FROM-MAC] 기존 mac_reminders 태스크 삭제 완료")

        # 이벤트 저장
        for event in data.get('events', []):
            # 한국어 날짜/시간 형식을 ISO 형식으로 변환
            raw_start = event.get('start_time')
            raw_end = event.get('end_time')
            start_time = parse_korean_datetime(raw_start)
            end_time = parse_korean_datetime(raw_end)

            # 디버그 로그: 각 이벤트의 파싱 결과
            print(f"[SYNC-FROM-MAC] 이벤트: {event.get('title')}")
            print(f"  - 원본 start_time: {raw_start} -> 파싱 결과: {start_time}")
            print(f"  - 원본 end_time: {raw_end} -> 파싱 결과: {end_time}")

            if USE_POSTGRES:
                cursor.execute('''
                    INSERT INTO events (title, start_time, end_time, category, source, sync_status)
                    VALUES (%s, %s, %s, %s, %s, %s)
                ''', (
                    event.get('title'),
                    start_time,
                    end_time,
                    event.get('category', ''),
                    'mac_calendar',
                    'synced'
                ))
            else:
                cursor.execute('''
                    INSERT INTO events (title, start_time, end_time, category, source, sync_status)
                    VALUES (?, ?, ?, ?, ?, ?)
                ''', (
                    event.get('title'),
                    start_time,
                    end_time,
                    event.get('category', ''),
                    'mac_calendar',
                    'synced'
                ))
            events_added += 1

        # 태스크 저장
        for task in data.get('tasks', []):
            # 한국어 날짜 형식을 ISO 형식으로 변환
            due_date = parse_korean_date(task.get('due_date'))

            if USE_POSTGRES:
                cursor.execute('''
                    INSERT INTO tasks (title, due_date, priority, category, source, sync_status)
                    VALUES (%s, %s, %s, %s, %s, %s)
                ''', (
                    task.get('title'),
                    due_date,
                    task.get('priority', 'medium'),
                    task.get('category', ''),
                    'mac_reminders',
                    'synced'
                ))
            else:
                cursor.execute('''
                    INSERT INTO tasks (title, due_date, priority, category, source, sync_status)
                    VALUES (?, ?, ?, ?, ?, ?)
                ''', (
                    task.get('title'),
                    due_date,
                    task.get('priority', 'medium'),
                    task.get('category', ''),
                    'mac_reminders',
                    'synced'
                ))
            tasks_added += 1

        conn.commit()
        conn.close()

        result = {
            'success': True,
            'message': f'Mac에서 동기화 완료: 이벤트 {events_added}개, 태스크 {tasks_added}개',
            'events_added': events_added,
            'tasks_added': tasks_added
        }
        print(f"[SYNC-FROM-MAC] 완료: {result}")
        return jsonify(result)
    except Exception as e:
        import traceback
        error_detail = traceback.format_exc()
        print(f"[SYNC-FROM-MAC] 오류 발생: {error_detail}")
        return jsonify({
            'success': False,
            'error': str(e),
            'detail': error_detail
        }), 500


@assistant_bp.route('/assistant/api/sync/to-mac', methods=['GET'])
def sync_to_mac():
    """
    Mac 단축어가 가져갈 pending_to_mac 상태의 데이터 반환
    단축어가 이 데이터를 Calendar/Reminders에 추가

    쿼리 파라미터:
    - type: 'events' | 'tasks' | None (둘 다)
    - category: 특정 카테고리만 필터링 (선택)
    """
    try:
        sync_type = request.args.get('type')  # 'events', 'tasks', or None
        category_filter = request.args.get('category')  # 선택적 카테고리 필터

        conn = get_db_connection()
        cursor = conn.cursor()

        events = []
        tasks = []

        # 이벤트 조회 (type이 없거나 'events'일 때)
        if not sync_type or sync_type == 'events':
            if category_filter:
                if USE_POSTGRES:
                    cursor.execute('''
                        SELECT * FROM events WHERE sync_status = %s AND category = %s
                    ''', ('pending_to_mac', category_filter))
                else:
                    cursor.execute('''
                        SELECT * FROM events WHERE sync_status = ? AND category = ?
                    ''', ('pending_to_mac', category_filter))
            else:
                if USE_POSTGRES:
                    cursor.execute('''
                        SELECT * FROM events WHERE sync_status = %s
                    ''', ('pending_to_mac',))
                else:
                    cursor.execute('''
                        SELECT * FROM events WHERE sync_status = ?
                    ''', ('pending_to_mac',))
            events = [dict(row) for row in cursor.fetchall()]

        # 태스크 조회 (type이 없거나 'tasks'일 때)
        if not sync_type or sync_type == 'tasks':
            if category_filter:
                if USE_POSTGRES:
                    cursor.execute('''
                        SELECT * FROM tasks WHERE sync_status = %s AND category = %s
                    ''', ('pending_to_mac', category_filter))
                else:
                    cursor.execute('''
                        SELECT * FROM tasks WHERE sync_status = ? AND category = ?
                    ''', ('pending_to_mac', category_filter))
            else:
                if USE_POSTGRES:
                    cursor.execute('''
                        SELECT * FROM tasks WHERE sync_status = %s
                    ''', ('pending_to_mac',))
                else:
                    cursor.execute('''
                        SELECT * FROM tasks WHERE sync_status = ?
                    ''', ('pending_to_mac',))
            tasks = [dict(row) for row in cursor.fetchall()]

        conn.close()

        # datetime 객체를 문자열로 변환 + end_time 자동 계산
        for event in events:
            for key in ['start_time', 'end_time', 'created_at', 'updated_at']:
                if event.get(key) and not isinstance(event[key], str):
                    event[key] = event[key].isoformat()

            # end_time이 없으면 start_time + 1시간으로 계산
            if not event.get('end_time') and event.get('start_time'):
                try:
                    start = datetime.fromisoformat(event['start_time'].replace('Z', '+00:00'))
                    end = start + timedelta(hours=1)
                    event['end_time'] = end.isoformat()
                except:
                    pass

        for task in tasks:
            for key in ['due_date', 'created_at', 'updated_at']:
                if task.get(key) and not isinstance(task[key], str):
                    task[key] = task[key].isoformat() if hasattr(task[key], 'isoformat') else str(task[key])

        # 자동으로 synced 처리 (auto_mark 파라미터가 false가 아닌 경우)
        auto_mark = request.args.get('auto_mark', 'true').lower() != 'false'
        if auto_mark and (events or tasks):
            conn2 = get_db_connection()
            cursor2 = conn2.cursor()

            event_ids = [e['id'] for e in events]
            task_ids = [t['id'] for t in tasks]

            if event_ids:
                if USE_POSTGRES:
                    cursor2.execute('''
                        UPDATE events SET sync_status = %s, updated_at = CURRENT_TIMESTAMP
                        WHERE id = ANY(%s)
                    ''', ('synced', event_ids))
                else:
                    placeholders = ','.join('?' * len(event_ids))
                    cursor2.execute(f'''
                        UPDATE events SET sync_status = ?, updated_at = CURRENT_TIMESTAMP
                        WHERE id IN ({placeholders})
                    ''', ['synced'] + event_ids)

            if task_ids:
                if USE_POSTGRES:
                    cursor2.execute('''
                        UPDATE tasks SET sync_status = %s, updated_at = CURRENT_TIMESTAMP
                        WHERE id = ANY(%s)
                    ''', ('synced', task_ids))
                else:
                    placeholders = ','.join('?' * len(task_ids))
                    cursor2.execute(f'''
                        UPDATE tasks SET sync_status = ?, updated_at = CURRENT_TIMESTAMP
                        WHERE id IN ({placeholders})
                    ''', ['synced'] + task_ids)

            conn2.commit()
            conn2.close()
            print(f"[SYNC-TO-MAC] 자동 synced 처리: events={len(event_ids)}, tasks={len(task_ids)}")

        return jsonify({
            'success': True,
            'events': events,
            'tasks': tasks,
            'total_pending': len(events) + len(tasks),
            'auto_marked_synced': auto_mark
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@assistant_bp.route('/assistant/api/sync/mark-synced', methods=['POST'])
def mark_synced():
    """
    Mac 단축어가 동기화 완료 후 호출하여 sync_status를 'synced'로 변경
    Body: { "event_ids": [...], "task_ids": [...] }
    """
    try:
        data = request.get_json()

        conn = get_db_connection()
        cursor = conn.cursor()

        event_ids = data.get('event_ids', [])
        task_ids = data.get('task_ids', [])

        if event_ids:
            if USE_POSTGRES:
                cursor.execute('''
                    UPDATE events SET sync_status = %s, updated_at = CURRENT_TIMESTAMP
                    WHERE id = ANY(%s)
                ''', ('synced', event_ids))
            else:
                placeholders = ','.join('?' * len(event_ids))
                cursor.execute(f'''
                    UPDATE events SET sync_status = ?, updated_at = CURRENT_TIMESTAMP
                    WHERE id IN ({placeholders})
                ''', ['synced'] + event_ids)

        if task_ids:
            if USE_POSTGRES:
                cursor.execute('''
                    UPDATE tasks SET sync_status = %s, updated_at = CURRENT_TIMESTAMP
                    WHERE id = ANY(%s)
                ''', ('synced', task_ids))
            else:
                placeholders = ','.join('?' * len(task_ids))
                cursor.execute(f'''
                    UPDATE tasks SET sync_status = ?, updated_at = CURRENT_TIMESTAMP
                    WHERE id IN ({placeholders})
                ''', ['synced'] + task_ids)

        conn.commit()
        conn.close()

        return jsonify({
            'success': True,
            'message': f'동기화 완료 처리: 이벤트 {len(event_ids)}개, 태스크 {len(task_ids)}개'
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# 단축어 B 호환용 별칭 (mark-synced와 동일)
@assistant_bp.route('/assistant/api/sync/confirm', methods=['POST'])
def sync_confirm():
    """Mac 단축어 B에서 사용하는 동기화 완료 확인 API (mark-synced 별칭)"""
    return mark_synced()


# ===== AI 통합 파싱 API (GPT-5.1) =====
@assistant_bp.route('/assistant/api/parse', methods=['POST'])
def parse_input():
    """
    자유 형식 텍스트를 AI가 분석하여 이벤트/태스크/인물/프로젝트로 구조화
    입력: { "text": "이번주 금요일 청년부 총회 오후 2시" }
    출력: { "events": [...], "tasks": [...], "people": [...], "projects": [...] }
    """
    try:
        data = request.get_json()
        text = data.get('text', '').strip()

        if not text:
            return jsonify({'success': False, 'error': '텍스트가 비어있습니다.'}), 400

        client = get_openai_client()
        if not client:
            return jsonify({'success': False, 'error': 'OpenAI API 키가 설정되지 않았습니다.'}), 500

        today = date.today()
        default_category = data.get('default_category', None)

        # GPT-5.1 통합 파싱 프롬프트
        system_prompt = f"""[역할]
너는 '개인 비서용 통합 파서'이다.
사용자가 붙여넣은 한국어/영어 텍스트를 읽고,
그 안에 있는 일정(events), 할 일(tasks), 인물(people), 프로젝트(projects)를 구조화된 JSON으로 추출한다.

[입력 컨텍스트]
- 오늘 날짜: {today.isoformat()} ({today.strftime('%A')})
- 기본 카테고리: {default_category or '없음'}

[출력]
반드시 다음 JSON 형식으로만 출력하라:
{{
  "events": [
    {{
      "title": "string",
      "date": "yyyy-MM-dd",
      "time": "HH:mm or null",
      "end_time": "HH:mm or null",
      "category": "교회 | 사업 | 유튜브 | 가정 | 공부 | 기타",
      "location": "string or null",
      "notes": "string or null"
    }}
  ],
  "tasks": [
    {{
      "title": "string",
      "due_date": "yyyy-MM-dd or null",
      "category": "교회 | 사업 | 유튜브 | 가정 | 공부 | 기타",
      "priority": "high | normal | low",
      "notes": "string or null"
    }}
  ],
  "people": [
    {{
      "name": "string (이름만, 직분 제외)",
      "role": "string or null (권사, 집사, 장로, 목사 등 직분/직책)",
      "category": "교회 | 사업 | 유튜브 | 가정 | 공부 | 기타",
      "notes": "string or null (건강상태, 특이사항 등)"
    }}
  ],
  "projects": [
    {{
      "name": "string (프로젝트명)",
      "description": "string or null",
      "status": "active | planning | completed",
      "priority": "high | medium | low",
      "start_date": "yyyy-MM-dd or null",
      "end_date": "yyyy-MM-dd or null"
    }}
  ]
}}

[분류 규칙]

1. events (일정)
   - 특정 시간/날짜에 실제로 '열리는 모임/행사/예배/회의'
   - 예: "12월 15일 청년부 총회 오후 2시"

2. tasks (할 일)
   - 행사를 준비하기 위한 "해야 할 일" (자료 준비, 문자 발송, 설교 작성 등)
   - 예: "주일 설교 준비해야 함"

3. people (인물)
   - 사람 이름 + 직분/직책이 함께 언급된 경우
   - 이름 뒤에 권사, 집사, 장로, 목사, 사장, 부장 등이 붙은 경우
   - 건강/수술/입원/사망 등 개인 상황이 언급된 경우
   - 예: "홍길동 권사", "김영희 집사 12월 10일 수술 예정", "박철수 장로 심방 필요"

4. projects (프로젝트)
   - "프로젝트", "사업", "계획" 등의 단어가 포함되거나
   - 장기적인 목표/기간이 명시된 작업
   - 예: "장기부진자 편지 발송 프로젝트", "바나바 교육 프로그램"

[날짜/시간 처리]
- "이번 주 금요일", "다음 주일", "고난주간 전" 등 상대적 표현은 오늘 날짜를 기준으로 계산
- 시간이 없으면 null

[category 분류]
- 교회: 예배, 기도회, 총회, 구역모임, 출석, 심방, 교회 행사, 권사/집사/장로/목사 등
- 사업: 무역, 스마트스토어, 재고, 배송, 세금, 광고 등
- 유튜브: 촬영, 편집, 썸네일, 스크립트, 업로드 등
- 가정: 가족 모임, 아이 일정, 개인 건강/가사 관련 등
- 공부: 코딩 공부, 강의 수강, 책 읽기 등
- 기타: 위 분류에 명확히 속하지 않는 경우

[중요]
- 하나의 텍스트에서 여러 유형의 항목을 동시에 추출할 수 있다
- 인물 정보와 함께 일정이 언급되면 둘 다 추출 (예: "홍길동 권사 12월 10일 수술" → people + events)
- 이해가 불가능한 정보는 무시한다
- JSON 외의 다른 텍스트는 절대로 출력하지 마라"""

        # GPT-5.1 Responses API 사용
        response = client.responses.create(
            model="gpt-5.1",
            input=[
                {
                    "role": "system",
                    "content": [{"type": "input_text", "text": system_prompt}]
                },
                {
                    "role": "user",
                    "content": [{"type": "input_text", "text": text}]
                }
            ],
            temperature=0.3
        )

        # 결과 추출
        if getattr(response, "output_text", None):
            result_text = response.output_text.strip()
        else:
            text_chunks = []
            for item in getattr(response, "output", []) or []:
                for content in getattr(item, "content", []) or []:
                    if getattr(content, "type", "") == "text":
                        text_chunks.append(getattr(content, "text", ""))
            result_text = "\n".join(text_chunks).strip()

        # JSON 파싱 (마크다운 코드블록 제거)
        if result_text.startswith("```"):
            result_text = result_text.split("```")[1]
            if result_text.startswith("json"):
                result_text = result_text[4:]
        result_text = result_text.strip()
        result = json.loads(result_text)

        # save_to_db 파라미터가 true면 DB에 저장
        save_to_db = data.get('save_to_db', False)
        saved_events = []
        saved_tasks = []
        saved_people = []
        saved_projects = []

        if save_to_db:
            conn = get_db_connection()
            cursor = conn.cursor()

            # 이벤트 저장
            for event in result.get('events', []):
                start_time = None
                if event.get('date'):
                    time_str = event.get('time') or '00:00'
                    start_time = f"{event['date']}T{time_str}:00"

                end_time = None
                if event.get('date') and event.get('end_time'):
                    end_time = f"{event['date']}T{event['end_time']}:00"

                if USE_POSTGRES:
                    cursor.execute('''
                        INSERT INTO events (title, start_time, end_time, category, source, sync_status)
                        VALUES (%s, %s, %s, %s, %s, %s)
                        RETURNING id
                    ''', (
                        event.get('title'),
                        start_time,
                        end_time,
                        event.get('category', '기타'),
                        'gpt_parse',
                        'pending_to_mac'
                    ))
                    event_id = cursor.fetchone()['id']
                else:
                    cursor.execute('''
                        INSERT INTO events (title, start_time, end_time, category, source, sync_status)
                        VALUES (?, ?, ?, ?, ?, ?)
                    ''', (
                        event.get('title'),
                        start_time,
                        end_time,
                        event.get('category', '기타'),
                        'gpt_parse',
                        'pending_to_mac'
                    ))
                    event_id = cursor.lastrowid

                saved_events.append({'id': event_id, 'title': event.get('title')})

            # 태스크 저장
            for task in result.get('tasks', []):
                if USE_POSTGRES:
                    cursor.execute('''
                        INSERT INTO tasks (title, due_date, priority, category, source, sync_status)
                        VALUES (%s, %s, %s, %s, %s, %s)
                        RETURNING id
                    ''', (
                        task.get('title'),
                        task.get('due_date'),
                        task.get('priority', 'normal'),
                        task.get('category', '기타'),
                        'gpt_parse',
                        'pending_to_mac'
                    ))
                    task_id = cursor.fetchone()['id']
                else:
                    cursor.execute('''
                        INSERT INTO tasks (title, due_date, priority, category, source, sync_status)
                        VALUES (?, ?, ?, ?, ?, ?)
                    ''', (
                        task.get('title'),
                        task.get('due_date'),
                        task.get('priority', 'normal'),
                        task.get('category', '기타'),
                        'gpt_parse',
                        'pending_to_mac'
                    ))
                    task_id = cursor.lastrowid

                saved_tasks.append({'id': task_id, 'title': task.get('title')})

            # 인물 저장
            for person in result.get('people', []):
                if USE_POSTGRES:
                    cursor.execute('''
                        INSERT INTO people (name, role, category, notes)
                        VALUES (%s, %s, %s, %s)
                        RETURNING id
                    ''', (
                        person.get('name'),
                        person.get('role'),
                        person.get('category', '기타'),
                        person.get('notes')
                    ))
                    person_id = cursor.fetchone()['id']
                else:
                    cursor.execute('''
                        INSERT INTO people (name, role, category, notes)
                        VALUES (?, ?, ?, ?)
                    ''', (
                        person.get('name'),
                        person.get('role'),
                        person.get('category', '기타'),
                        person.get('notes')
                    ))
                    person_id = cursor.lastrowid

                saved_people.append({'id': person_id, 'name': person.get('name')})

            # 프로젝트 저장
            for project in result.get('projects', []):
                if USE_POSTGRES:
                    cursor.execute('''
                        INSERT INTO projects (name, description, status, priority, start_date, end_date)
                        VALUES (%s, %s, %s, %s, %s, %s)
                        RETURNING id
                    ''', (
                        project.get('name'),
                        project.get('description'),
                        project.get('status', 'active'),
                        project.get('priority', 'medium'),
                        project.get('start_date'),
                        project.get('end_date')
                    ))
                    project_id = cursor.fetchone()['id']
                else:
                    cursor.execute('''
                        INSERT INTO projects (name, description, status, priority, start_date, end_date)
                        VALUES (?, ?, ?, ?, ?, ?)
                    ''', (
                        project.get('name'),
                        project.get('description'),
                        project.get('status', 'active'),
                        project.get('priority', 'medium'),
                        project.get('start_date'),
                        project.get('end_date')
                    ))
                    project_id = cursor.lastrowid

                saved_projects.append({'id': project_id, 'name': project.get('name')})

            conn.commit()
            conn.close()

        return jsonify({
            'success': True,
            'parsed': result,
            'original_text': text,
            'saved_to_db': save_to_db,
            'saved_events': saved_events,
            'saved_tasks': saved_tasks,
            'saved_people': saved_people,
            'saved_projects': saved_projects
        })
    except json.JSONDecodeError as e:
        return jsonify({'success': False, 'error': f'JSON 파싱 오류: {str(e)}'}), 500
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ===== 출석 API =====
@assistant_bp.route('/assistant/api/attendance', methods=['POST'])
def upload_attendance():
    """
    출석 데이터 업로드 (파일 또는 JSON)
    Body: { "date": "YYYY-MM-DD", "records": [{"name": "홍길동", "status": "present"}, ...] }
    """
    try:
        data = request.get_json()

        attendance_date = data.get('date', date.today().isoformat())
        records = data.get('records', [])

        if not records:
            return jsonify({'success': False, 'error': '출석 데이터가 비어있습니다.'}), 400

        conn = get_db_connection()
        cursor = conn.cursor()

        added = 0
        for record in records:
            if USE_POSTGRES:
                cursor.execute('''
                    INSERT INTO attendance (name, date, status, source)
                    VALUES (%s, %s, %s, %s)
                ''', (
                    record.get('name'),
                    attendance_date,
                    record.get('status', 'present'),
                    'web'
                ))
            else:
                cursor.execute('''
                    INSERT INTO attendance (name, date, status, source)
                    VALUES (?, ?, ?, ?)
                ''', (
                    record.get('name'),
                    attendance_date,
                    record.get('status', 'present'),
                    'web'
                ))
            added += 1

        conn.commit()
        conn.close()

        return jsonify({
            'success': True,
            'message': f'{attendance_date} 출석 데이터 {added}건 저장 완료'
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@assistant_bp.route('/assistant/api/attendance', methods=['GET'])
def get_attendance():
    """특정 날짜의 출석 데이터 조회"""
    try:
        target_date = request.args.get('date', date.today().isoformat())

        conn = get_db_connection()
        cursor = conn.cursor()

        if USE_POSTGRES:
            cursor.execute('''
                SELECT * FROM attendance WHERE date = %s ORDER BY name
            ''', (target_date,))
        else:
            cursor.execute('''
                SELECT * FROM attendance WHERE date = ? ORDER BY name
            ''', (target_date,))

        records = [dict(row) for row in cursor.fetchall()]
        conn.close()

        # date 객체를 문자열로 변환
        for record in records:
            if record.get('date') and not isinstance(record['date'], str):
                record['date'] = record['date'].isoformat()
            if record.get('created_at') and not isinstance(record['created_at'], str):
                record['created_at'] = record['created_at'].isoformat()

        return jsonify({
            'success': True,
            'date': target_date,
            'records': records,
            'total': len(records)
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@assistant_bp.route('/assistant/api/attendance/upload', methods=['POST'])
def upload_attendance_file():
    """
    CSV/XLSX 파일 업로드로 출석 데이터 일괄 저장

    CSV 형식 예시:
    name,date,status,group_name
    홍길동,2024-12-01,present,청년부
    김철수,2024-12-01,absent,청년부

    또는 단순 형식 (날짜/그룹은 파라미터로):
    name,status
    홍길동,present
    김철수,absent
    """
    try:
        # 파일 체크
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': '파일이 없습니다.'}), 400

        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'error': '파일이 선택되지 않았습니다.'}), 400

        # 추가 파라미터 (CSV에 날짜/그룹이 없을 경우 사용)
        default_date = request.form.get('date', date.today().isoformat())
        default_group = request.form.get('group_name', '')

        filename = file.filename.lower()
        records = []

        if filename.endswith('.csv'):
            # CSV 파싱
            content = file.read().decode('utf-8-sig')  # BOM 처리
            reader = csv.DictReader(io.StringIO(content))

            for row in reader:
                name = row.get('name', row.get('이름', '')).strip()
                if not name:
                    continue

                records.append({
                    'name': name,
                    'date': row.get('date', row.get('날짜', default_date)),
                    'status': row.get('status', row.get('상태', 'present')),
                    'group_name': row.get('group_name', row.get('그룹', default_group))
                })

        elif filename.endswith('.xlsx') or filename.endswith('.xls'):
            # XLSX 파싱 (openpyxl 필요)
            try:
                from openpyxl import load_workbook

                wb = load_workbook(filename=io.BytesIO(file.read()))
                ws = wb.active

                headers = [cell.value for cell in ws[1]]
                header_map = {}
                for idx, h in enumerate(headers):
                    if h:
                        h_lower = str(h).lower().strip()
                        if h_lower in ['name', '이름']:
                            header_map['name'] = idx
                        elif h_lower in ['date', '날짜']:
                            header_map['date'] = idx
                        elif h_lower in ['status', '상태']:
                            header_map['status'] = idx
                        elif h_lower in ['group_name', '그룹', 'group']:
                            header_map['group_name'] = idx

                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row or not any(row):
                        continue

                    name = ''
                    if 'name' in header_map and row[header_map['name']]:
                        name = str(row[header_map['name']]).strip()

                    if not name:
                        continue

                    record_date = default_date
                    if 'date' in header_map and row[header_map['date']]:
                        d = row[header_map['date']]
                        if hasattr(d, 'isoformat'):
                            record_date = d.isoformat()[:10]
                        else:
                            record_date = str(d)[:10]

                    status = 'present'
                    if 'status' in header_map and row[header_map['status']]:
                        status = str(row[header_map['status']]).strip().lower()
                        # 한글 상태 변환
                        if status in ['출석', '참석', 'o', '○']:
                            status = 'present'
                        elif status in ['결석', '불참', 'x', '×']:
                            status = 'absent'

                    group_name = default_group
                    if 'group_name' in header_map and row[header_map['group_name']]:
                        group_name = str(row[header_map['group_name']]).strip()

                    records.append({
                        'name': name,
                        'date': record_date,
                        'status': status,
                        'group_name': group_name
                    })

            except ImportError:
                return jsonify({
                    'success': False,
                    'error': 'XLSX 파일 처리를 위한 openpyxl 라이브러리가 설치되지 않았습니다.'
                }), 500
        else:
            return jsonify({
                'success': False,
                'error': '지원하지 않는 파일 형식입니다. CSV 또는 XLSX 파일을 업로드해주세요.'
            }), 400

        if not records:
            return jsonify({'success': False, 'error': '유효한 출석 데이터가 없습니다.'}), 400

        # DB에 저장
        conn = get_db_connection()
        cursor = conn.cursor()

        added = 0
        for record in records:
            if USE_POSTGRES:
                cursor.execute('''
                    INSERT INTO attendance (name, date, status, group_name, source)
                    VALUES (%s, %s, %s, %s, %s)
                ''', (
                    record['name'],
                    record['date'],
                    record['status'],
                    record['group_name'],
                    'file_upload'
                ))
            else:
                cursor.execute('''
                    INSERT INTO attendance (name, date, status, group_name, source)
                    VALUES (?, ?, ?, ?, ?)
                ''', (
                    record['name'],
                    record['date'],
                    record['status'],
                    record['group_name'],
                    'file_upload'
                ))
            added += 1

        conn.commit()
        conn.close()

        return jsonify({
            'success': True,
            'message': f'출석 데이터 {added}건 업로드 완료',
            'records_added': added,
            'sample': records[:3] if len(records) > 3 else records
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@assistant_bp.route('/assistant/api/attendance/under-attending', methods=['GET'])
def get_under_attending():
    """
    부진자(지속 결석자) 조회 API

    쿼리 파라미터:
    - weeks: 연속 결석 기준 주 수 (기본: 2)
    - group: 특정 그룹만 필터링 (선택)

    응답:
    [{ name, absent_weeks, last_attended_date, group_name }]
    """
    try:
        weeks = int(request.args.get('weeks', 2))
        group_filter = request.args.get('group', None)

        conn = get_db_connection()
        cursor = conn.cursor()

        # 최근 N주간의 예배 날짜 조회 (일요일 기준)
        today = date.today()

        # 최근 예배 날짜들 조회 (출석 기록이 있는 날짜들)
        if group_filter:
            if USE_POSTGRES:
                cursor.execute('''
                    SELECT DISTINCT date FROM attendance
                    WHERE group_name = %s
                    ORDER BY date DESC
                    LIMIT %s
                ''', (group_filter, weeks + 2))  # 여유분 포함
            else:
                cursor.execute('''
                    SELECT DISTINCT date FROM attendance
                    WHERE group_name = ?
                    ORDER BY date DESC
                    LIMIT ?
                ''', (group_filter, weeks + 2))
        else:
            if USE_POSTGRES:
                cursor.execute('''
                    SELECT DISTINCT date FROM attendance
                    ORDER BY date DESC
                    LIMIT %s
                ''', (weeks + 2,))
            else:
                cursor.execute('''
                    SELECT DISTINCT date FROM attendance
                    ORDER BY date DESC
                    LIMIT ?
                ''', (weeks + 2,))

        recent_dates = [row['date'] if isinstance(row['date'], str) else row['date'].isoformat()
                       for row in cursor.fetchall()]

        if len(recent_dates) < weeks:
            conn.close()
            return jsonify({
                'success': True,
                'message': f'충분한 출석 데이터가 없습니다. (최소 {weeks}주 필요, 현재 {len(recent_dates)}주)',
                'under_attending': [],
                'recent_dates': recent_dates
            })

        # 최근 N주 날짜
        check_dates = recent_dates[:weeks]

        # 모든 멤버 조회
        if group_filter:
            if USE_POSTGRES:
                cursor.execute('''
                    SELECT DISTINCT name, group_name FROM attendance
                    WHERE group_name = %s
                ''', (group_filter,))
            else:
                cursor.execute('''
                    SELECT DISTINCT name, group_name FROM attendance
                    WHERE group_name = ?
                ''', (group_filter,))
        else:
            cursor.execute('''
                SELECT DISTINCT name, group_name FROM attendance
            ''')

        members = [dict(row) for row in cursor.fetchall()]

        under_attending = []

        for member in members:
            name = member['name']
            group_name = member.get('group_name', '')

            # 이 멤버의 최근 N주 출석 현황 확인
            if USE_POSTGRES:
                cursor.execute('''
                    SELECT date, status FROM attendance
                    WHERE name = %s AND date = ANY(%s)
                    ORDER BY date DESC
                ''', (name, check_dates))
            else:
                placeholders = ','.join('?' * len(check_dates))
                cursor.execute(f'''
                    SELECT date, status FROM attendance
                    WHERE name = ? AND date IN ({placeholders})
                    ORDER BY date DESC
                ''', [name] + check_dates)

            attendance_records = {
                (row['date'] if isinstance(row['date'], str) else row['date'].isoformat()): row['status']
                for row in cursor.fetchall()
            }

            # 연속 결석 횟수 계산
            absent_count = 0
            for check_date in check_dates:
                status = attendance_records.get(check_date, 'absent')  # 기록 없으면 결석으로 간주
                if status == 'absent':
                    absent_count += 1
                else:
                    break  # 출석 기록이 있으면 연속 결석 종료

            # 기준 이상 연속 결석인 경우 부진자로 분류
            if absent_count >= weeks:
                # 마지막 출석 날짜 조회
                if USE_POSTGRES:
                    cursor.execute('''
                        SELECT date FROM attendance
                        WHERE name = %s AND status = 'present'
                        ORDER BY date DESC
                        LIMIT 1
                    ''', (name,))
                else:
                    cursor.execute('''
                        SELECT date FROM attendance
                        WHERE name = ? AND status = 'present'
                        ORDER BY date DESC
                        LIMIT 1
                    ''', (name,))

                last_row = cursor.fetchone()
                last_attended = None
                if last_row:
                    last_attended = last_row['date'] if isinstance(last_row['date'], str) else last_row['date'].isoformat()

                under_attending.append({
                    'name': name,
                    'absent_weeks': absent_count,
                    'last_attended_date': last_attended,
                    'group_name': group_name
                })

        conn.close()

        # 결석 주수로 정렬 (많은 순)
        under_attending.sort(key=lambda x: x['absent_weeks'], reverse=True)

        return jsonify({
            'success': True,
            'under_attending': under_attending,
            'total': len(under_attending),
            'weeks_checked': weeks,
            'check_dates': check_dates,
            'group_filter': group_filter
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


# ===== 심방 문자 프롬프트 (프로필별) =====
def _get_youth_pastor_prompt(today, style):
    """청년부 전도사 톤 프롬프트"""
    style_guide = {
        '따뜻한': '위로와 안부를 중심으로 따뜻하게 작성한다. 요즘 어떻게 지내는지 진심으로 궁금해하는 느낌.',
        '격려': '용기와 소망을 주는 방향으로 작성한다. "함께 걸어가고 싶다", "응원한다"는 뉘앙스를 담는다.',
        '공식적인': '짧고 담백하게 안내/확인 말투로 작성한다. 너무 감성적이지 않게.'
    }

    return f"""[역할]
너는 청년부 전도사를 돕는 심방/안부 문자 작성 도우미이다.
청년부에서 연락이 뜸한 청년들에게 보낼 따뜻하고 친근한 안부 문자를 작성한다.

[오늘 날짜]
{today.isoformat()}

[대상]
10대 후반 ~ 20대 청년

[문체 스타일: {style}]
{style_guide.get(style, style_guide['따뜻한'])}

[규칙]
1. 반드시 한국어 존댓말을 사용하되, 친근하고 가볍게 작성한다. 무겁지 않게.
2. 각 사람마다 2~3문장으로 간결하게 작성한다.
3. 이름은 "OO님" 형태로 자연스럽게 포함한다.
4. 정죄, 비난, 눈치 주는 표현은 절대 사용하지 않는다.
5. "왜 안 나왔냐"보다는 "요즘 잘 지내는지 궁금해서 연락 드렸다"는 방향으로.
6. "함께 예배 드리고 싶다", "함께 있어주고 싶다"는 뉘앙스를 담는다.
7. 결석 주수를 직접적으로 언급하지 않는다 (예: "3주 동안 안 나오셨네요" X)
8. 이모티콘은 1개까지만 허용한다 (예: 😊, 🙌). 없어도 된다.
9. 혼자 감당하기 버거운 일이 있으면 알려달라는 제안을 자연스럽게 넣어도 좋다.

[출력 형식]
반드시 다음 JSON 형식으로만 출력한다:
{{"messages": [
  {{"name": "이름", "message": "문자 내용"}}
]}}

JSON 외의 다른 텍스트(설명, 주석 등)는 절대 출력하지 마라.

[예시 - 참고용]
"OO님, 요즘 잘 지내고 계신가요? 최근 예배에서 얼굴을 잘 뵙지 못해서 문득 생각이 났어요. 혹시 혼자 감당하기 버거운 일들이 있다면 언제든 편하게 알려 주세요. 청년부에서 함께 예배드리고, 같이 이야기 나눌 수 있으면 좋겠습니다. 😊"
"""


def _get_adult_pastor_prompt(today, style):
    """장년 목사 톤 프롬프트"""
    style_guide = {
        '따뜻한': '안부와 위로를 중심으로 작성한다. 요즘 어떻게 지내시는지, 건강은 어떠신지 묻는 따뜻한 마음.',
        '격려': '신앙, 기도, 소망에 대한 부드러운 격려를 한 문장 추가한다.',
        '공식적인': '안내성, 사실 전달 위주로 짧게 작성한다. 담백하고 공손하게.'
    }

    return f"""[역할]
너는 담임목사(또는 담당 목사)를 돕는 장년 성도 심방 문자 작성 도우미이다.
예배에 오랫동안 참석하지 못한 장년 성도님들에게 보낼 따뜻하고 공손한 안부 문자를 작성한다.

[오늘 날짜]
{today.isoformat()}

[대상]
장년 성도 (40~70대, 집사님/권사님/장로님 포함)

[문체 스타일: {style}]
{style_guide.get(style, style_guide['따뜻한'])}

[규칙]
1. 반드시 한국어 존댓말을 사용한다. 공손하고 목회자의 따뜻한 배려가 느껴지는 톤.
2. 각 사람마다 2~3문장으로 간결하게 작성한다.
3. 이름 앞에 적절한 호칭을 붙인다:
   - group_name이 "청년부"면 "OO님"
   - 그 외(장년부 등)면 "OO 집사님", "OO 성도님" 등 자연스럽게 선택
4. **절대로 "어르신"이라는 단어를 사용하지 마라.** 이 단어는 금지어이다.
5. 정죄, 비난, 압박 표현은 절대 사용하지 않는다.
6. "오랫동안 안 나오셨다"는 표현 대신 "최근에 예배에서 자주 뵙지 못했다" 정도로 부드럽게.
7. 결석 주수를 직접적으로 언급하지 않는다.
8. 건강과 안부를 묻고, 필요하면 "기도 제목 나누어 달라"는 제안을 포함한다.
9. 예배 자리를 부담 없이 다시 초대하는 뉘앙스를 담는다.
10. 이모티콘은 0~1개, 너무 가볍지 않게. 없어도 된다.

[출력 형식]
반드시 다음 JSON 형식으로만 출력한다:
{{"messages": [
  {{"name": "이름", "message": "문자 내용"}}
]}}

JSON 외의 다른 텍스트(설명, 주석 등)는 절대 출력하지 마라.

[예시 - 참고용]
"OO 집사님, 잘 지내고 계신지요? 최근 예배 자리에서 자주 뵙지 못해 안부가 궁금하여 연락을 드렸습니다. 혹시 기도 부탁하실 일이나 어려운 상황이 있으시면 언제든 말씀해 주시면 함께 기도하겠습니다."
"""


@assistant_bp.route('/assistant/api/attendance/messages', methods=['POST'])
def generate_care_messages():
    """
    부진자별 심방/안부 문자 GPT 생성 API

    Body:
    {
        "people": [{ "name": "홍길동", "absent_weeks": 3, "last_attended_date": "2024-11-10", "group_name": "청년부" }],
        "style": "따뜻한" | "격려" | "공식적인" (선택, 기본: 따뜻한),
        "profile": "youth" | "adult" (선택, 기본: adult)
    }

    응답:
    { "messages": [{ "name": "홍길동", "message": "..." }] }
    """
    try:
        data = request.get_json()
        people = data.get('people', [])
        style = data.get('style', '따뜻한')
        profile = data.get('profile', 'adult')  # youth 또는 adult

        if not people:
            return jsonify({'success': False, 'error': '대상자 목록이 비어있습니다.'}), 400

        client = get_openai_client()
        if not client:
            return jsonify({'success': False, 'error': 'OpenAI API 키가 설정되지 않았습니다.'}), 500

        today = date.today()

        # 프로필에 따른 system prompt 선택
        if profile == 'youth':
            system_prompt = _get_youth_pastor_prompt(today, style)
        else:
            system_prompt = _get_adult_pastor_prompt(today, style)

        # 대상자 정보 정리
        people_info = "\n".join([
            f"- {p['name']} ({p.get('group_name', '그룹 미지정')}): {p['absent_weeks']}주 연속 결석, 마지막 출석: {p.get('last_attended_date', '기록 없음')}"
            for p in people
        ])

        user_prompt = f"""다음 분들에게 보낼 안부 문자를 작성해주세요:

{people_info}"""

        response = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_prompt}
            ],
            temperature=0.7,
            response_format={"type": "json_object"}
        )

        result = json.loads(response.choices[0].message.content)

        # 결과 형식 정규화
        messages = result if isinstance(result, list) else result.get('messages', result.get('data', []))

        return jsonify({
            'success': True,
            'messages': messages,
            'total': len(messages),
            'style': style,
            'profile': profile
        })
    except json.JSONDecodeError as e:
        return jsonify({'success': False, 'error': f'JSON 파싱 오류: {str(e)}'}), 500
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@assistant_bp.route('/assistant/api/attendance/groups', methods=['GET'])
def get_attendance_groups():
    """출석부에 등록된 그룹 목록 조회"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        cursor.execute('''
            SELECT DISTINCT group_name FROM attendance
            WHERE group_name IS NOT NULL AND group_name != ''
            ORDER BY group_name
        ''')

        groups = [row['group_name'] for row in cursor.fetchall()]
        conn.close()

        return jsonify({
            'success': True,
            'groups': groups
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ===== 대시보드 요약 API =====
@assistant_bp.route('/assistant/api/dashboard', methods=['GET'])
def get_dashboard():
    """대시보드용 요약 데이터"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        today = date.today()
        week_end = today + timedelta(days=7)

        # 오늘 이벤트
        if USE_POSTGRES:
            cursor.execute('''
                SELECT * FROM events WHERE DATE(start_time) = %s ORDER BY start_time
            ''', (today,))
        else:
            cursor.execute('''
                SELECT * FROM events WHERE DATE(start_time) = ? ORDER BY start_time
            ''', (today.isoformat(),))
        today_events = [dict(row) for row in cursor.fetchall()]

        # 이번 주 이벤트
        if USE_POSTGRES:
            cursor.execute('''
                SELECT * FROM events
                WHERE DATE(start_time) > %s AND DATE(start_time) <= %s
                ORDER BY start_time
            ''', (today, week_end))
        else:
            cursor.execute('''
                SELECT * FROM events
                WHERE DATE(start_time) > ? AND DATE(start_time) <= ?
                ORDER BY start_time
            ''', (today.isoformat(), week_end.isoformat()))
        week_events = [dict(row) for row in cursor.fetchall()]

        # 미완료 태스크
        if USE_POSTGRES:
            cursor.execute('''
                SELECT * FROM tasks WHERE is_completed = FALSE
                ORDER BY due_date ASC NULLS LAST
            ''')
        else:
            cursor.execute('''
                SELECT * FROM tasks WHERE is_completed = 0
                ORDER BY due_date ASC
            ''')
        pending_tasks = [dict(row) for row in cursor.fetchall()]

        # Pending sync 카운트
        if USE_POSTGRES:
            cursor.execute('''
                SELECT COUNT(*) as count FROM events WHERE sync_status = %s
            ''', ('pending_to_mac',))
        else:
            cursor.execute('''
                SELECT COUNT(*) as count FROM events WHERE sync_status = ?
            ''', ('pending_to_mac',))
        pending_events_count = cursor.fetchone()['count']

        if USE_POSTGRES:
            cursor.execute('''
                SELECT COUNT(*) as count FROM tasks WHERE sync_status = %s
            ''', ('pending_to_mac',))
        else:
            cursor.execute('''
                SELECT COUNT(*) as count FROM tasks WHERE sync_status = ?
            ''', ('pending_to_mac',))
        pending_tasks_count = cursor.fetchone()['count']

        conn.close()

        # datetime 변환
        def convert_datetime(items):
            for item in items:
                for key in ['start_time', 'end_time', 'due_date', 'created_at', 'updated_at']:
                    if item.get(key) and not isinstance(item[key], str):
                        item[key] = item[key].isoformat() if hasattr(item[key], 'isoformat') else str(item[key])
            return items

        return jsonify({
            'success': True,
            'today': today.isoformat(),
            'today_events': convert_datetime(today_events),
            'week_events': convert_datetime(week_events),
            'pending_tasks': convert_datetime(pending_tasks),
            'pending_sync': {
                'events': pending_events_count,
                'tasks': pending_tasks_count,
                'total': pending_events_count + pending_tasks_count
            }
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@assistant_bp.route('/assistant/api/debug/all-events', methods=['GET'])
def debug_all_events():
    """디버그용: DB에 저장된 모든 이벤트 조회"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        cursor.execute('SELECT * FROM events ORDER BY id DESC LIMIT 20')
        all_events = [dict(row) for row in cursor.fetchall()]

        conn.close()

        # datetime 객체를 문자열로 변환
        for event in all_events:
            for key in ['start_time', 'end_time', 'created_at', 'updated_at']:
                if event.get(key) and not isinstance(event[key], str):
                    event[key] = event[key].isoformat() if hasattr(event[key], 'isoformat') else str(event[key])

        return jsonify({
            'success': True,
            'count': len(all_events),
            'events': all_events
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ===== 뉴스/이슈 분석 API =====
import requests
from xml.etree import ElementTree

# 뉴스 캐시 (간단한 메모리 캐시)
_news_cache = {
    'data': None,
    'updated_at': None
}

def fetch_rss_news():
    """Google News RSS 피드에서 뉴스 가져오기"""
    news_items = []

    # 한국 뉴스 RSS 피드들
    rss_feeds = [
        # Google News Korea - 주요 뉴스
        ('https://news.google.com/rss?hl=ko&gl=KR&ceid=KR:ko', '국내'),
        # Google News Korea - 세계 뉴스
        ('https://news.google.com/rss/topics/CAAqJggKIiBDQkFTRWdvSUwyMHZNRGx1YlY4U0FtdHZHZ0pMVWlnQVAB?hl=ko&gl=KR&ceid=KR:ko', '해외'),
    ]

    for feed_url, category in rss_feeds:
        try:
            response = requests.get(feed_url, timeout=10, headers={
                'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36'
            })
            if response.status_code == 200:
                root = ElementTree.fromstring(response.content)
                items = root.findall('.//item')[:10]  # 각 피드에서 최대 10개

                for item in items:
                    title = item.find('title')
                    link = item.find('link')
                    pub_date = item.find('pubDate')

                    if title is not None:
                        news_items.append({
                            'title': title.text,
                            'link': link.text if link is not None else '',
                            'pub_date': pub_date.text if pub_date is not None else '',
                            'category': category
                        })
        except Exception as e:
            print(f"[NEWS] RSS 피드 오류 ({category}): {e}")
            continue

    return news_items


def parse_rss_date(date_str):
    """RSS 날짜 문자열을 한국 시간으로 변환"""
    if not date_str:
        return None
    try:
        from email.utils import parsedate_to_datetime
        import pytz
        dt = parsedate_to_datetime(date_str)
        kst = pytz.timezone('Asia/Seoul')
        dt_kst = dt.astimezone(kst)
        return dt_kst.strftime('%Y-%m-%d %H:%M')
    except Exception:
        return None


def analyze_news_with_gpt(news_items):
    """GPT로 뉴스 분석 및 유튜브 영상 소재 평가"""
    client = get_openai_client()
    if not client:
        return None

    today = date.today()

    # 뉴스 헤드라인 목록 생성 (인덱스 포함)
    headlines_text = "\n".join([
        f"[{i}][{item['category']}] {item['title']}"
        for i, item in enumerate(news_items[:20])
    ])

    system_prompt = f"""[역할]
너는 뉴스 분석가이자 유튜브 이슈 채널 PD의 어시스턴트이다.
오늘 날짜: {today.isoformat()}

[임무]
주어진 뉴스 헤드라인들을 분석하여:
1. 중요한 뉴스를 선별하고 (중복/비슷한 내용은 하나로 통합)
2. 각 뉴스를 간략하게 요약하고
3. 해외 뉴스는 한국인이 이해하기 쉽게 맥락을 설명하고
4. 유튜브 이슈 영상 소재로서의 가능성을 평가한다

[영상 소재 가능성 평가 기준]
- high (높음): 논쟁적, 대중적 관심 높음, 조회수 기대, 트렌드 연관
- medium (보통): 정보 가치 있음, 일부 관심층 존재
- low (낮음): 일상적 뉴스, 특별한 관심 유발 어려움

[출력 형식]
반드시 다음 JSON 형식으로만 출력:
{{
  "news": [
    {{
      "original_index": 0,
      "category": "국내" | "해외",
      "title": "뉴스 제목 (간결하게)",
      "summary": "2-3문장 요약",
      "interpretation": "해외 뉴스인 경우 한국인을 위한 맥락 설명 (국내 뉴스면 null)",
      "video_potential": "high" | "medium" | "low",
      "video_reason": "영상 소재로서의 이유 (1문장)"
    }}
  ]
}}

[규칙]
1. 최대 8개의 중요 뉴스만 선별한다
2. 국내/해외 균형있게 포함한다 (각각 3-4개)
3. 단순 연예/스포츠 결과보다는 사회적 이슈를 우선한다
4. 같은 사건의 후속 보도는 하나로 통합한다
5. original_index는 입력된 뉴스의 [인덱스] 번호를 그대로 사용한다
6. JSON 외의 텍스트는 출력하지 마라"""

    try:
        response = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": f"다음 뉴스 헤드라인들을 분석해주세요:\n\n{headlines_text}"}
            ],
            temperature=0.5,
            response_format={"type": "json_object"}
        )

        result = json.loads(response.choices[0].message.content)
        analyzed = result.get('news', [])

        # 원본 뉴스에서 링크와 발행 시간 추가
        for item in analyzed:
            idx = item.get('original_index', -1)
            if 0 <= idx < len(news_items):
                original = news_items[idx]
                item['link'] = original.get('link', '')
                item['pub_date'] = parse_rss_date(original.get('pub_date', ''))
            else:
                item['link'] = ''
                item['pub_date'] = None

        return analyzed
    except Exception as e:
        print(f"[NEWS] GPT 분석 오류: {e}")
        return None


@assistant_bp.route('/assistant/api/news', methods=['GET'])
def get_news():
    """
    오늘의 뉴스/이슈 조회 API

    쿼리 파라미터:
    - refresh: true면 캐시 무시하고 새로 가져오기

    응답:
    {
        "success": true,
        "news": [
            {
                "category": "국내" | "해외",
                "title": "제목",
                "summary": "요약",
                "interpretation": "해석 (해외 뉴스만)",
                "video_potential": "high" | "medium" | "low",
                "video_reason": "영상 소재 이유"
            }
        ],
        "updated_at": "2025-12-04T09:00:00"
    }
    """
    try:
        refresh = request.args.get('refresh', 'false').lower() == 'true'

        # 캐시 확인 (30분 이내면 캐시 사용)
        now = datetime.now()
        if not refresh and _news_cache['data'] and _news_cache['updated_at']:
            cache_age = (now - _news_cache['updated_at']).total_seconds()
            if cache_age < 1800:  # 30분 = 1800초
                return jsonify({
                    'success': True,
                    'news': _news_cache['data'],
                    'updated_at': _news_cache['updated_at'].isoformat(),
                    'cached': True
                })

        # 뉴스 가져오기
        raw_news = fetch_rss_news()

        if not raw_news:
            # RSS 실패 시 기본 응답
            return jsonify({
                'success': True,
                'news': [],
                'updated_at': now.isoformat(),
                'message': '뉴스를 가져오는 중 오류가 발생했습니다. 잠시 후 다시 시도해주세요.'
            })

        # GPT로 분석
        analyzed_news = analyze_news_with_gpt(raw_news)

        if analyzed_news:
            # 캐시 업데이트
            _news_cache['data'] = analyzed_news
            _news_cache['updated_at'] = now

            return jsonify({
                'success': True,
                'news': analyzed_news,
                'updated_at': now.isoformat(),
                'cached': False
            })
        else:
            # GPT 분석 실패 시 원본 뉴스 반환
            fallback_news = [
                {
                    'category': item['category'],
                    'title': item['title'],
                    'summary': '',
                    'interpretation': None,
                    'video_potential': 'medium',
                    'video_reason': '분석 대기 중'
                }
                for item in raw_news[:8]
            ]
            return jsonify({
                'success': True,
                'news': fallback_news,
                'updated_at': now.isoformat(),
                'fallback': True
            })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@assistant_bp.route('/assistant/api/news/script', methods=['POST'])
def generate_news_script():
    """
    뉴스 기반 유튜브 대본 생성 API

    Body:
    {
        "title": "뉴스 제목",
        "summary": "뉴스 요약",
        "interpretation": "해석 (선택)",
        "category": "국내" | "해외"
    }

    응답:
    {
        "success": true,
        "script": "생성된 대본 텍스트"
    }
    """
    try:
        data = request.get_json()
        title = data.get('title', '')
        summary = data.get('summary', '')
        interpretation = data.get('interpretation', '')
        category = data.get('category', '국내')

        if not title:
            return jsonify({'success': False, 'error': '뉴스 제목이 필요합니다.'}), 400

        client = get_openai_client()
        if not client:
            return jsonify({'success': False, 'error': 'OpenAI API 키가 설정되지 않았습니다.'}), 500

        today = date.today()

        system_prompt = f"""[역할]
너는 유튜브 이슈/뉴스 채널의 전문 대본 작가이다.
주어진 뉴스를 바탕으로 10분 분량의 유튜브 영상 대본을 작성한다.

[오늘 날짜]
{today.isoformat()}

[대본 구성]
1. 오프닝 (30초): 시청자의 관심을 끄는 도입부 - 핵심 이슈를 간결하게 소개
2. 배경 설명 (2분): 이 뉴스가 왜 중요한지, 관련 배경 정보
3. 본론 1 (2분 30초): 뉴스의 핵심 내용 상세 설명
4. 본론 2 (2분): 다양한 시각/반응/영향 분석
5. 전망 및 의견 (2분): 앞으로의 전망, 시청자가 주목해야 할 점
6. 클로징 (1분): 요약 + 구독/좋아요 유도 멘트

[작성 규칙]
1. 자연스러운 구어체로 작성한다 (유튜브 대본이므로)
2. 각 섹션은 [섹션명] 형식으로 구분한다
3. 시청자를 '여러분'으로 호칭한다
4. 너무 딱딱하지 않게, 친근하면서도 신뢰감 있는 톤을 유지한다
5. 각 섹션마다 대략적인 시간을 표기한다
6. 총 분량은 약 2000-2500자 (말하면 약 10분)
7. 해외 뉴스인 경우 한국 시청자가 이해하기 쉽게 맥락을 설명한다
8. 정치적으로 편향되지 않게 객관적으로 작성한다
9. 확인되지 않은 정보는 "~로 알려졌습니다", "~라는 보도가 있습니다" 형식으로 작성한다

[출력]
대본 텍스트만 출력한다. JSON이 아닌 순수 텍스트로 출력."""

        user_content = f"""다음 뉴스를 바탕으로 10분 분량의 유튜브 대본을 작성해주세요.

[카테고리] {category}
[제목] {title}
[요약] {summary}
{f'[해석/맥락] {interpretation}' if interpretation else ''}
"""

        response = client.chat.completions.create(
            model="gpt-5.1",  # GPT 5.1
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_content}
            ],
            max_tokens=4000
        )

        script = response.choices[0].message.content

        return jsonify({
            'success': True,
            'script': script,
            'title': title
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@assistant_bp.route('/assistant/api/news/refresh', methods=['POST'])
def refresh_news():
    """뉴스 새로고침 (캐시 무시)"""
    # GET API에 refresh=true 파라미터 전달하는 것과 동일
    try:
        raw_news = fetch_rss_news()

        if not raw_news:
            return jsonify({
                'success': False,
                'error': '뉴스를 가져오는 데 실패했습니다.'
            }), 500

        analyzed_news = analyze_news_with_gpt(raw_news)

        now = datetime.now()

        if analyzed_news:
            _news_cache['data'] = analyzed_news
            _news_cache['updated_at'] = now

            return jsonify({
                'success': True,
                'news': analyzed_news,
                'updated_at': now.isoformat()
            })
        else:
            return jsonify({
                'success': False,
                'error': 'GPT 분석에 실패했습니다.'
            }), 500

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ===== People (인물 관리) API =====
@assistant_bp.route('/assistant/api/people', methods=['GET'])
def get_people():
    """인물 목록 조회"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        search = request.args.get('search', '')
        category = request.args.get('category', '')

        if USE_POSTGRES:
            query = 'SELECT * FROM people WHERE 1=1'
            params = []
            if search:
                query += ' AND name ILIKE %s'
                params.append(f'%{search}%')
            if category:
                query += ' AND category = %s'
                params.append(category)
            query += ' ORDER BY updated_at DESC'
            cursor.execute(query, params)
        else:
            query = 'SELECT * FROM people WHERE 1=1'
            params = []
            if search:
                query += ' AND name LIKE ?'
                params.append(f'%{search}%')
            if category:
                query += ' AND category = ?'
                params.append(category)
            query += ' ORDER BY updated_at DESC'
            cursor.execute(query, params)

        people = [dict(row) for row in cursor.fetchall()]
        conn.close()

        return jsonify({'success': True, 'people': people})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@assistant_bp.route('/assistant/api/people/check-duplicate', methods=['POST'])
def check_duplicate_person():
    """동명이인 확인"""
    try:
        data = request.get_json()
        name = data.get('name', '').strip()
        if not name:
            return jsonify({'success': True, 'duplicates': []})

        conn = get_db_connection()
        cursor = conn.cursor()

        # 이름이 같거나 비슷한 인물 검색
        if USE_POSTGRES:
            cursor.execute('''
                SELECT id, name, category, phone, birthday FROM people
                WHERE name ILIKE %s OR name ILIKE %s
                ORDER BY updated_at DESC
            ''', (name, f'%{name}%'))
        else:
            cursor.execute('''
                SELECT id, name, category, phone, birthday FROM people
                WHERE name LIKE ? OR name LIKE ?
                ORDER BY updated_at DESC
            ''', (name, f'%{name}%'))

        duplicates = [dict(row) for row in cursor.fetchall()]
        conn.close()

        return jsonify({'success': True, 'duplicates': duplicates})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@assistant_bp.route('/assistant/api/people', methods=['POST'])
def create_person():
    """새 인물 생성 (중복 확인 및 생일 이벤트 자동 생성)"""
    try:
        data = request.get_json()
        name = data.get('name', '').strip()
        if not name:
            return jsonify({'success': False, 'error': '이름은 필수입니다.'}), 400

        # 중복 확인 옵션 (force=true면 중복 무시)
        force = data.get('force', False)

        conn = get_db_connection()
        cursor = conn.cursor()

        # 중복 확인 (force가 아닐 때만)
        if not force:
            if USE_POSTGRES:
                cursor.execute('SELECT id, name, category, phone, birthday FROM people WHERE name ILIKE %s', (name,))
            else:
                cursor.execute('SELECT id, name, category, phone, birthday FROM people WHERE name LIKE ?', (name,))

            duplicates = [dict(row) for row in cursor.fetchall()]
            if duplicates:
                conn.close()
                return jsonify({
                    'success': False,
                    'error': 'duplicate_found',
                    'message': f'"{name}"과(와) 동일하거나 비슷한 이름의 인물이 이미 있습니다.',
                    'duplicates': duplicates
                }), 409  # Conflict

        birthday = data.get('birthday')

        if USE_POSTGRES:
            cursor.execute('''
                INSERT INTO people (name, category, phone, email, address, birthday, notes)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
                RETURNING id
            ''', (name, data.get('category'), data.get('phone'), data.get('email'), data.get('address'), birthday, data.get('notes')))
            person_id = cursor.fetchone()['id']
        else:
            cursor.execute('''
                INSERT INTO people (name, category, phone, email, address, birthday, notes)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            ''', (name, data.get('category'), data.get('phone'), data.get('email'), data.get('address'), birthday, data.get('notes')))
            person_id = cursor.lastrowid

        # 생일이 있으면 캘린더 이벤트 자동 생성
        birthday_event_created = False
        if birthday:
            try:
                # 올해 생일 이벤트 생성
                year = datetime.now().year
                birthday_date = datetime.strptime(birthday, '%Y-%m-%d')
                this_year_birthday = birthday_date.replace(year=year)

                # 이미 지났으면 내년으로
                if this_year_birthday.date() < date.today():
                    this_year_birthday = this_year_birthday.replace(year=year + 1)

                birthday_str = this_year_birthday.strftime('%Y-%m-%d')

                if USE_POSTGRES:
                    cursor.execute('''
                        INSERT INTO events (title, start_time, end_time, category, source, sync_status)
                        VALUES (%s, %s, %s, %s, %s, %s)
                    ''', (f'🎂 {name} 생일', f'{birthday_str}T00:00:00', f'{birthday_str}T23:59:59', '생일', 'web', 'pending_to_mac'))
                else:
                    cursor.execute('''
                        INSERT INTO events (title, start_time, end_time, category, source, sync_status)
                        VALUES (?, ?, ?, ?, ?, ?)
                    ''', (f'🎂 {name} 생일', f'{birthday_str}T00:00:00', f'{birthday_str}T23:59:59', '생일', 'web', 'pending_to_mac'))
                birthday_event_created = True
            except Exception as e:
                print(f"[PERSON] 생일 이벤트 생성 실패: {e}")

        conn.commit()
        conn.close()

        message = f'{name} 인물이 추가되었습니다.'
        if birthday_event_created:
            message += ' 생일 이벤트가 캘린더에 추가되었습니다.'

        return jsonify({'success': True, 'id': person_id, 'message': message, 'birthday_event_created': birthday_event_created})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@assistant_bp.route('/assistant/api/people/<int:person_id>', methods=['GET'])
def get_person(person_id):
    """특정 인물 상세 조회 (노트 포함)"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        if USE_POSTGRES:
            cursor.execute('SELECT * FROM people WHERE id = %s', (person_id,))
        else:
            cursor.execute('SELECT * FROM people WHERE id = ?', (person_id,))

        person = cursor.fetchone()
        if not person:
            conn.close()
            return jsonify({'success': False, 'error': '인물을 찾을 수 없습니다.'}), 404

        person = dict(person)

        # 노트 목록도 함께 조회
        if USE_POSTGRES:
            cursor.execute('''
                SELECT * FROM people_notes WHERE person_id = %s ORDER BY note_date DESC, created_at DESC
            ''', (person_id,))
        else:
            cursor.execute('''
                SELECT * FROM people_notes WHERE person_id = ? ORDER BY note_date DESC, created_at DESC
            ''', (person_id,))

        notes = [dict(row) for row in cursor.fetchall()]
        person['notes_list'] = notes

        conn.close()
        return jsonify({'success': True, 'person': person})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@assistant_bp.route('/assistant/api/people/<int:person_id>', methods=['PUT'])
def update_person(person_id):
    """인물 정보 수정 (생일 변경 시 캘린더 이벤트 업데이트)"""
    try:
        data = request.get_json()
        conn = get_db_connection()
        cursor = conn.cursor()

        name = data.get('name')
        birthday = data.get('birthday')

        if USE_POSTGRES:
            cursor.execute('''
                UPDATE people SET name = %s, category = %s, phone = %s, email = %s, address = %s, birthday = %s, notes = %s, updated_at = CURRENT_TIMESTAMP
                WHERE id = %s
            ''', (name, data.get('category'), data.get('phone'), data.get('email'), data.get('address'), birthday, data.get('notes'), person_id))
        else:
            cursor.execute('''
                UPDATE people SET name = ?, category = ?, phone = ?, email = ?, address = ?, birthday = ?, notes = ?, updated_at = CURRENT_TIMESTAMP
                WHERE id = ?
            ''', (name, data.get('category'), data.get('phone'), data.get('email'), data.get('address'), birthday, data.get('notes'), person_id))

        conn.commit()
        conn.close()
        return jsonify({'success': True, 'message': '인물 정보가 수정되었습니다.'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@assistant_bp.route('/assistant/api/people/<int:person_id>', methods=['DELETE'])
def delete_person(person_id):
    """인물 삭제"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        if USE_POSTGRES:
            cursor.execute('DELETE FROM people WHERE id = %s', (person_id,))
        else:
            cursor.execute('DELETE FROM people WHERE id = ?', (person_id,))

        conn.commit()
        conn.close()
        return jsonify({'success': True, 'message': '인물이 삭제되었습니다.'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@assistant_bp.route('/assistant/api/people/<int:person_id>/notes', methods=['POST'])
def add_person_note(person_id):
    """인물에 노트 추가"""
    try:
        data = request.get_json()
        content = data.get('content', '').strip()
        if not content:
            return jsonify({'success': False, 'error': '내용은 필수입니다.'}), 400

        note_date = data.get('note_date', date.today().isoformat())
        category = data.get('category')

        conn = get_db_connection()
        cursor = conn.cursor()

        if USE_POSTGRES:
            cursor.execute('''
                INSERT INTO people_notes (person_id, note_date, content, category)
                VALUES (%s, %s, %s, %s)
                RETURNING id
            ''', (person_id, note_date, content, category))
            note_id = cursor.fetchone()['id']
            # 인물의 updated_at 갱신
            cursor.execute('UPDATE people SET updated_at = CURRENT_TIMESTAMP WHERE id = %s', (person_id,))
        else:
            cursor.execute('''
                INSERT INTO people_notes (person_id, note_date, content, category)
                VALUES (?, ?, ?, ?)
            ''', (person_id, note_date, content, category))
            note_id = cursor.lastrowid
            cursor.execute('UPDATE people SET updated_at = CURRENT_TIMESTAMP WHERE id = ?', (person_id,))

        conn.commit()
        conn.close()
        return jsonify({'success': True, 'id': note_id, 'message': '노트가 추가되었습니다.'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@assistant_bp.route('/assistant/api/people/notes/<int:note_id>', methods=['DELETE'])
def delete_person_note(note_id):
    """인물 노트 삭제"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        if USE_POSTGRES:
            cursor.execute('DELETE FROM people_notes WHERE id = %s', (note_id,))
        else:
            cursor.execute('DELETE FROM people_notes WHERE id = ?', (note_id,))

        conn.commit()
        conn.close()
        return jsonify({'success': True, 'message': '노트가 삭제되었습니다.'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ===== Projects (프로젝트 관리) API =====
@assistant_bp.route('/assistant/api/projects', methods=['GET'])
def get_projects():
    """프로젝트 목록 조회"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        status = request.args.get('status', '')
        search = request.args.get('search', '')

        if USE_POSTGRES:
            query = 'SELECT * FROM projects WHERE 1=1'
            params = []
            if status:
                query += ' AND status = %s'
                params.append(status)
            if search:
                query += ' AND name ILIKE %s'
                params.append(f'%{search}%')
            query += ' ORDER BY updated_at DESC'
            cursor.execute(query, params)
        else:
            query = 'SELECT * FROM projects WHERE 1=1'
            params = []
            if status:
                query += ' AND status = ?'
                params.append(status)
            if search:
                query += ' AND name LIKE ?'
                params.append(f'%{search}%')
            query += ' ORDER BY updated_at DESC'
            cursor.execute(query, params)

        projects = [dict(row) for row in cursor.fetchall()]
        conn.close()

        return jsonify({'success': True, 'projects': projects})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@assistant_bp.route('/assistant/api/projects/check-duplicate', methods=['POST'])
def check_duplicate_project():
    """유사 프로젝트 확인"""
    try:
        data = request.get_json()
        name = data.get('name', '').strip()
        if not name:
            return jsonify({'success': True, 'duplicates': []})

        conn = get_db_connection()
        cursor = conn.cursor()

        if USE_POSTGRES:
            cursor.execute('''
                SELECT id, name, status, start_date, end_date FROM projects
                WHERE name ILIKE %s OR name ILIKE %s
                ORDER BY updated_at DESC
            ''', (name, f'%{name}%'))
        else:
            cursor.execute('''
                SELECT id, name, status, start_date, end_date FROM projects
                WHERE name LIKE ? OR name LIKE ?
                ORDER BY updated_at DESC
            ''', (name, f'%{name}%'))

        duplicates = [dict(row) for row in cursor.fetchall()]
        conn.close()

        return jsonify({'success': True, 'duplicates': duplicates})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@assistant_bp.route('/assistant/api/projects', methods=['POST'])
def create_project():
    """새 프로젝트 생성 (중복 확인 및 일정 자동 생성)"""
    try:
        data = request.get_json()
        name = data.get('name', '').strip()
        if not name:
            return jsonify({'success': False, 'error': '프로젝트 이름은 필수입니다.'}), 400

        # 중복 확인 옵션 (force=true면 중복 무시)
        force = data.get('force', False)

        conn = get_db_connection()
        cursor = conn.cursor()

        # 중복 확인 (force가 아닐 때만)
        if not force:
            if USE_POSTGRES:
                cursor.execute('SELECT id, name, status, start_date, end_date FROM projects WHERE name ILIKE %s', (name,))
            else:
                cursor.execute('SELECT id, name, status, start_date, end_date FROM projects WHERE name LIKE ?', (name,))

            duplicates = [dict(row) for row in cursor.fetchall()]
            if duplicates:
                conn.close()
                return jsonify({
                    'success': False,
                    'error': 'duplicate_found',
                    'message': f'"{name}"과(와) 동일하거나 비슷한 이름의 프로젝트가 이미 있습니다.',
                    'duplicates': duplicates
                }), 409  # Conflict

        start_date = data.get('start_date')
        end_date = data.get('end_date')

        if USE_POSTGRES:
            cursor.execute('''
                INSERT INTO projects (name, description, status, start_date, end_date, priority)
                VALUES (%s, %s, %s, %s, %s, %s)
                RETURNING id
            ''', (name, data.get('description'), data.get('status', 'active'), start_date, end_date, data.get('priority', 'medium')))
            project_id = cursor.fetchone()['id']
        else:
            cursor.execute('''
                INSERT INTO projects (name, description, status, start_date, end_date, priority)
                VALUES (?, ?, ?, ?, ?, ?)
            ''', (name, data.get('description'), data.get('status', 'active'), start_date, end_date, data.get('priority', 'medium')))
            project_id = cursor.lastrowid

        # 날짜가 있으면 캘린더 이벤트 자동 생성
        events_created = []
        if start_date:
            try:
                if USE_POSTGRES:
                    cursor.execute('''
                        INSERT INTO events (title, start_time, end_time, category, source, sync_status)
                        VALUES (%s, %s, %s, %s, %s, %s)
                    ''', (f'📋 {name} 시작', f'{start_date}T09:00:00', f'{start_date}T10:00:00', '프로젝트', 'web', 'pending_to_mac'))
                else:
                    cursor.execute('''
                        INSERT INTO events (title, start_time, end_time, category, source, sync_status)
                        VALUES (?, ?, ?, ?, ?, ?)
                    ''', (f'📋 {name} 시작', f'{start_date}T09:00:00', f'{start_date}T10:00:00', '프로젝트', 'web', 'pending_to_mac'))
                events_created.append('시작일')
            except Exception as e:
                print(f"[PROJECT] 시작일 이벤트 생성 실패: {e}")

        if end_date:
            try:
                if USE_POSTGRES:
                    cursor.execute('''
                        INSERT INTO events (title, start_time, end_time, category, source, sync_status)
                        VALUES (%s, %s, %s, %s, %s, %s)
                    ''', (f'🏁 {name} 마감', f'{end_date}T09:00:00', f'{end_date}T10:00:00', '프로젝트', 'web', 'pending_to_mac'))
                else:
                    cursor.execute('''
                        INSERT INTO events (title, start_time, end_time, category, source, sync_status)
                        VALUES (?, ?, ?, ?, ?, ?)
                    ''', (f'🏁 {name} 마감', f'{end_date}T09:00:00', f'{end_date}T10:00:00', '프로젝트', 'web', 'pending_to_mac'))
                events_created.append('마감일')
            except Exception as e:
                print(f"[PROJECT] 마감일 이벤트 생성 실패: {e}")

        conn.commit()
        conn.close()

        message = f'{name} 프로젝트가 생성되었습니다.'
        if events_created:
            message += f' 캘린더에 {", ".join(events_created)} 이벤트가 추가되었습니다.'

        return jsonify({'success': True, 'id': project_id, 'message': message, 'events_created': events_created})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@assistant_bp.route('/assistant/api/projects/<int:project_id>', methods=['GET'])
def get_project(project_id):
    """특정 프로젝트 상세 조회 (노트 포함)"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        if USE_POSTGRES:
            cursor.execute('SELECT * FROM projects WHERE id = %s', (project_id,))
        else:
            cursor.execute('SELECT * FROM projects WHERE id = ?', (project_id,))

        project = cursor.fetchone()
        if not project:
            conn.close()
            return jsonify({'success': False, 'error': '프로젝트를 찾을 수 없습니다.'}), 404

        project = dict(project)

        # 노트 목록도 함께 조회
        if USE_POSTGRES:
            cursor.execute('''
                SELECT * FROM project_notes WHERE project_id = %s ORDER BY note_date DESC, created_at DESC
            ''', (project_id,))
        else:
            cursor.execute('''
                SELECT * FROM project_notes WHERE project_id = ? ORDER BY note_date DESC, created_at DESC
            ''', (project_id,))

        notes = [dict(row) for row in cursor.fetchall()]
        project['notes_list'] = notes

        conn.close()
        return jsonify({'success': True, 'project': project})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@assistant_bp.route('/assistant/api/projects/<int:project_id>', methods=['PUT'])
def update_project(project_id):
    """프로젝트 정보 수정"""
    try:
        data = request.get_json()
        conn = get_db_connection()
        cursor = conn.cursor()

        if USE_POSTGRES:
            cursor.execute('''
                UPDATE projects SET name = %s, description = %s, status = %s, start_date = %s, end_date = %s, priority = %s, updated_at = CURRENT_TIMESTAMP
                WHERE id = %s
            ''', (data.get('name'), data.get('description'), data.get('status'), data.get('start_date'), data.get('end_date'), data.get('priority'), project_id))
        else:
            cursor.execute('''
                UPDATE projects SET name = ?, description = ?, status = ?, start_date = ?, end_date = ?, priority = ?, updated_at = CURRENT_TIMESTAMP
                WHERE id = ?
            ''', (data.get('name'), data.get('description'), data.get('status'), data.get('start_date'), data.get('end_date'), data.get('priority'), project_id))

        conn.commit()
        conn.close()
        return jsonify({'success': True, 'message': '프로젝트 정보가 수정되었습니다.'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@assistant_bp.route('/assistant/api/projects/<int:project_id>', methods=['DELETE'])
def delete_project(project_id):
    """프로젝트 삭제"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        if USE_POSTGRES:
            cursor.execute('DELETE FROM projects WHERE id = %s', (project_id,))
        else:
            cursor.execute('DELETE FROM projects WHERE id = ?', (project_id,))

        conn.commit()
        conn.close()
        return jsonify({'success': True, 'message': '프로젝트가 삭제되었습니다.'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@assistant_bp.route('/assistant/api/projects/<int:project_id>/notes', methods=['POST'])
def add_project_note(project_id):
    """프로젝트에 노트 추가"""
    try:
        data = request.get_json()
        content = data.get('content', '').strip()
        if not content:
            return jsonify({'success': False, 'error': '내용은 필수입니다.'}), 400

        note_date = data.get('note_date', date.today().isoformat())
        category = data.get('category')

        conn = get_db_connection()
        cursor = conn.cursor()

        if USE_POSTGRES:
            cursor.execute('''
                INSERT INTO project_notes (project_id, note_date, content, category)
                VALUES (%s, %s, %s, %s)
                RETURNING id
            ''', (project_id, note_date, content, category))
            note_id = cursor.fetchone()['id']
            cursor.execute('UPDATE projects SET updated_at = CURRENT_TIMESTAMP WHERE id = %s', (project_id,))
        else:
            cursor.execute('''
                INSERT INTO project_notes (project_id, note_date, content, category)
                VALUES (?, ?, ?, ?)
            ''', (project_id, note_date, content, category))
            note_id = cursor.lastrowid
            cursor.execute('UPDATE projects SET updated_at = CURRENT_TIMESTAMP WHERE id = ?', (project_id,))

        conn.commit()
        conn.close()
        return jsonify({'success': True, 'id': note_id, 'message': '노트가 추가되었습니다.'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@assistant_bp.route('/assistant/api/projects/notes/<int:note_id>', methods=['DELETE'])
def delete_project_note(note_id):
    """프로젝트 노트 삭제"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        if USE_POSTGRES:
            cursor.execute('DELETE FROM project_notes WHERE id = %s', (note_id,))
        else:
            cursor.execute('DELETE FROM project_notes WHERE id = ?', (note_id,))

        conn.commit()
        conn.close()
        return jsonify({'success': True, 'message': '노트가 삭제되었습니다.'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ===== AI 분석 (Quick Input for People/Project) =====
@assistant_bp.route('/assistant/api/analyze-input-people', methods=['POST'])
def analyze_input_people():
    """
    Quick Input 텍스트를 AI로 분석하여 인물/프로젝트 정보 추출
    예: "홍길동 집사 12월 10일 담낭암 수술" -> 인물: 홍길동 집사, 노트: 12월 10일 담낭암 수술
    """
    try:
        data = request.get_json()
        input_text = data.get('text', '').strip()
        input_type = data.get('type', 'people')  # 'people' or 'project'

        if not input_text:
            return jsonify({'success': False, 'error': '입력 텍스트가 없습니다.'}), 400

        client = OpenAI()

        # 오늘 날짜 정보
        today = date.today()
        today_str = today.isoformat()
        year = today.year

        if input_type == 'people':
            system_prompt = f"""당신은 텍스트에서 인물 정보를 추출하는 AI입니다.
오늘 날짜: {today_str}

입력된 텍스트에서 다음을 추출하세요:
1. 인물 이름 (name): 이름과 직함/호칭을 함께 (예: "홍길동 집사", "김영희 권사")
2. 카테고리 (category): 교회 직분 또는 관계 (예: "집사", "권사", "장로", "청년부" 등)
3. 전화번호 (phone): 전화번호가 언급되면 추출 (예: "010-1234-5678")
4. 이메일 (email): 이메일이 언급되면 추출
5. 주소 (address): 주소가 언급되면 추출
6. 생일 (birthday): 생일이 언급되면 YYYY-MM-DD 형식으로 추출 (예: "12월 10일 생일" → "{year}-12-10")
7. 노트 날짜 (note_date): 특정 일정이나 사건 날짜 (YYYY-MM-DD 형식, 없으면 오늘 날짜)
8. 노트 내용 (note_content): 인물에 관한 상황/내용 요약

중요: 날짜를 정확히 변환하세요.
- "12월 10일" → "{year}-12-10"
- "다음주 월요일" → 실제 날짜로 변환
- "내년 1월" → "{year+1}-01-XX"

JSON 형식으로 반환 (해당 없는 필드는 null):
{{
  "name": "홍길동 집사",
  "category": "집사",
  "phone": "010-1234-5678",
  "email": null,
  "address": null,
  "birthday": "1960-05-15",
  "note_date": "{today_str}",
  "note_content": "담낭암 수술 예정"
}}
"""
        else:  # project
            system_prompt = f"""당신은 텍스트에서 프로젝트 정보를 추출하는 AI입니다.
오늘 날짜: {today_str}

입력된 텍스트에서 다음을 추출하세요:
1. 프로젝트 이름 (name): 프로젝트/업무 명칭
2. 설명 (description): 프로젝트 간략 설명
3. 시작일 (start_date): 프로젝트 시작 날짜 (YYYY-MM-DD 형식)
4. 마감일 (end_date): 프로젝트 마감/완료 예정 날짜 (YYYY-MM-DD 형식)
5. 노트 날짜 (note_date): 언급된 날짜 (YYYY-MM-DD 형식, 없으면 오늘 날짜)
6. 노트 내용 (note_content): 진행 상황/내용 요약
7. 우선순위 (priority): high, medium, low 중 하나

중요: 날짜를 정확히 변환하세요.
- "12월 10일" → "{year}-12-10"
- "다음주 금요일까지" → 실제 날짜로 변환
- "내년 1월" → "{year+1}-01-XX"

JSON 형식으로 반환 (해당 없는 필드는 null):
{{
  "name": "교회 리모델링",
  "description": "교회 본당 리모델링 프로젝트",
  "start_date": "{today_str}",
  "end_date": "{year}-12-31",
  "note_date": "{today_str}",
  "note_content": "설계 도면 검토 완료",
  "priority": "high"
}}
"""

        response = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": input_text}
            ],
            max_tokens=500,
            temperature=0.3
        )

        result_text = response.choices[0].message.content.strip()

        # JSON 파싱
        import json
        # ```json ... ``` 형식 제거
        if '```' in result_text:
            result_text = result_text.split('```')[1]
            if result_text.startswith('json'):
                result_text = result_text[4:]
            result_text = result_text.strip()

        parsed_result = json.loads(result_text)

        return jsonify({'success': True, 'parsed': parsed_result, 'type': input_type})
    except json.JSONDecodeError:
        return jsonify({'success': False, 'error': 'AI 응답을 파싱할 수 없습니다.'}), 500
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@assistant_bp.route('/assistant/api/quick-add-people', methods=['POST'])
def quick_add_people():
    """
    AI 분석 결과를 바탕으로 인물/노트를 빠르게 추가
    - 동명이인 발견 시 duplicate_found 에러 반환 (사용자 확인 필요)
    - force=true면 중복 무시하고 추가
    - 기존 인물이면 노트만 추가, 새 인물이면 인물 생성 후 노트 추가
    - 생일이 있으면 캘린더 이벤트 자동 생성
    """
    try:
        data = request.get_json()
        name = data.get('name', '').strip()
        category = data.get('category')
        phone = data.get('phone')
        email = data.get('email')
        address = data.get('address')
        birthday = data.get('birthday')
        note_date = data.get('note_date', date.today().isoformat())
        note_content = data.get('note_content', '').strip()
        force = data.get('force', False)
        use_existing_id = data.get('use_existing_id')  # 기존 인물에 추가할 경우

        if not name:
            return jsonify({'success': False, 'error': '이름은 필수입니다.'}), 400

        conn = get_db_connection()
        cursor = conn.cursor()

        person_id = None
        is_new = False
        birthday_event_created = False

        # 기존 인물 ID가 지정되면 그 인물 사용
        if use_existing_id:
            person_id = use_existing_id
            message = f'기존 인물에 노트가 추가되었습니다.'
        else:
            # 동명이인 검색
            if USE_POSTGRES:
                cursor.execute('SELECT id, name, category, phone, birthday FROM people WHERE name ILIKE %s', (f'%{name}%',))
            else:
                cursor.execute('SELECT id, name, category, phone, birthday FROM people WHERE name LIKE ?', (f'%{name}%',))

            similar_people = [dict(row) for row in cursor.fetchall()]

            # 정확히 일치하는 인물 찾기
            exact_match = None
            for p in similar_people:
                if p['name'].lower() == name.lower():
                    exact_match = p
                    break

            if exact_match:
                # 정확히 일치하는 인물이 있으면 그 인물에 노트 추가
                person_id = exact_match['id']
                message = f'기존 인물 "{name}"에 노트가 추가되었습니다.'
            elif similar_people and not force:
                # 비슷한 이름이 있고 force가 아니면 확인 요청
                conn.close()
                return jsonify({
                    'success': False,
                    'error': 'duplicate_found',
                    'message': f'"{name}"과(와) 비슷한 이름의 인물이 있습니다. 동일 인물인지 확인해주세요.',
                    'duplicates': similar_people,
                    'parsed_data': {
                        'name': name,
                        'category': category,
                        'phone': phone,
                        'email': email,
                        'address': address,
                        'birthday': birthday,
                        'note_date': note_date,
                        'note_content': note_content
                    }
                }), 409
            else:
                # 새 인물 생성
                is_new = True
                if USE_POSTGRES:
                    cursor.execute('''
                        INSERT INTO people (name, category, phone, email, address, birthday)
                        VALUES (%s, %s, %s, %s, %s, %s)
                        RETURNING id
                    ''', (name, category, phone, email, address, birthday))
                    person_id = cursor.fetchone()['id']
                else:
                    cursor.execute('''
                        INSERT INTO people (name, category, phone, email, address, birthday)
                        VALUES (?, ?, ?, ?, ?, ?)
                    ''', (name, category, phone, email, address, birthday))
                    person_id = cursor.lastrowid

                message = f'새 인물 "{name}"이(가) 생성되었습니다.'

                # 생일이 있으면 캘린더 이벤트 자동 생성
                if birthday:
                    try:
                        year = datetime.now().year
                        birthday_date = datetime.strptime(birthday, '%Y-%m-%d')
                        this_year_birthday = birthday_date.replace(year=year)

                        if this_year_birthday.date() < date.today():
                            this_year_birthday = this_year_birthday.replace(year=year + 1)

                        birthday_str = this_year_birthday.strftime('%Y-%m-%d')

                        if USE_POSTGRES:
                            cursor.execute('''
                                INSERT INTO events (title, start_time, end_time, category, source, sync_status)
                                VALUES (%s, %s, %s, %s, %s, %s)
                            ''', (f'🎂 {name} 생일', f'{birthday_str}T00:00:00', f'{birthday_str}T23:59:59', '생일', 'web', 'pending_to_mac'))
                        else:
                            cursor.execute('''
                                INSERT INTO events (title, start_time, end_time, category, source, sync_status)
                                VALUES (?, ?, ?, ?, ?, ?)
                            ''', (f'🎂 {name} 생일', f'{birthday_str}T00:00:00', f'{birthday_str}T23:59:59', '생일', 'web', 'pending_to_mac'))
                        birthday_event_created = True
                        message += ' 생일 이벤트가 캘린더에 추가되었습니다.'
                    except Exception as e:
                        print(f"[QUICK-ADD] 생일 이벤트 생성 실패: {e}")

        # 노트 추가
        note_event_created = False
        if note_content:
            if USE_POSTGRES:
                cursor.execute('''
                    INSERT INTO people_notes (person_id, note_date, content, category)
                    VALUES (%s, %s, %s, %s)
                ''', (person_id, note_date, note_content, category))
                cursor.execute('UPDATE people SET updated_at = CURRENT_TIMESTAMP WHERE id = %s', (person_id,))
            else:
                cursor.execute('''
                    INSERT INTO people_notes (person_id, note_date, content, category)
                    VALUES (?, ?, ?, ?)
                ''', (person_id, note_date, note_content, category))
                cursor.execute('UPDATE people SET updated_at = CURRENT_TIMESTAMP WHERE id = ?', (person_id,))

            if not is_new:
                message += ' 노트가 추가되었습니다.'

            # 노트에 날짜가 있으면 캘린더 이벤트 자동 생성 (오늘이 아닌 미래 날짜만)
            if note_date and note_date != date.today().isoformat():
                try:
                    note_date_obj = datetime.strptime(note_date, '%Y-%m-%d').date()
                    if note_date_obj >= date.today():  # 미래 또는 오늘 날짜만
                        # 이벤트 제목: 인물명 - 노트 내용 요약
                        event_title = f'👤 {name} - {note_content[:30]}{"..." if len(note_content) > 30 else ""}'
                        if USE_POSTGRES:
                            cursor.execute('''
                                INSERT INTO events (title, start_time, end_time, category, source, sync_status)
                                VALUES (%s, %s, %s, %s, %s, %s)
                            ''', (event_title, f'{note_date}T09:00:00', f'{note_date}T10:00:00', '심방', 'web', 'pending_to_mac'))
                        else:
                            cursor.execute('''
                                INSERT INTO events (title, start_time, end_time, category, source, sync_status)
                                VALUES (?, ?, ?, ?, ?, ?)
                            ''', (event_title, f'{note_date}T09:00:00', f'{note_date}T10:00:00', '심방', 'web', 'pending_to_mac'))
                        note_event_created = True
                        message += f' 캘린더에 {note_date} 일정이 추가되었습니다.'
                except Exception as e:
                    print(f"[QUICK-ADD] 노트 일정 이벤트 생성 실패: {e}")

        conn.commit()
        conn.close()

        return jsonify({
            'success': True,
            'person_id': person_id,
            'message': message,
            'is_new': is_new,
            'birthday_event_created': birthday_event_created,
            'note_event_created': note_event_created
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@assistant_bp.route('/assistant/api/quick-add-project', methods=['POST'])
def quick_add_project():
    """
    AI 분석 결과를 바탕으로 프로젝트/노트를 빠르게 추가
    - 유사 프로젝트 발견 시 duplicate_found 에러 반환
    - force=true면 중복 무시하고 추가
    - 기존 프로젝트면 노트만 추가, 새 프로젝트면 프로젝트 생성 후 노트 추가
    - start_date, end_date가 있으면 캘린더 이벤트 자동 생성
    """
    try:
        data = request.get_json()
        name = data.get('name', '').strip()
        description = data.get('description')
        priority = data.get('priority', 'medium')
        start_date = data.get('start_date')
        end_date = data.get('end_date')
        note_date = data.get('note_date', date.today().isoformat())
        note_content = data.get('note_content', '').strip()
        force = data.get('force', False)
        use_existing_id = data.get('use_existing_id')

        if not name:
            return jsonify({'success': False, 'error': '프로젝트 이름은 필수입니다.'}), 400

        conn = get_db_connection()
        cursor = conn.cursor()

        project_id = None
        is_new = False
        events_created = []

        # 기존 프로젝트 ID가 지정되면 그 프로젝트 사용
        if use_existing_id:
            project_id = use_existing_id
            message = f'기존 프로젝트에 노트가 추가되었습니다.'
        else:
            # 유사 프로젝트 검색
            if USE_POSTGRES:
                cursor.execute('SELECT id, name, status, start_date, end_date FROM projects WHERE name ILIKE %s', (f'%{name}%',))
            else:
                cursor.execute('SELECT id, name, status, start_date, end_date FROM projects WHERE name LIKE ?', (f'%{name}%',))

            similar_projects = [dict(row) for row in cursor.fetchall()]

            # 정확히 일치하는 프로젝트 찾기
            exact_match = None
            for p in similar_projects:
                if p['name'].lower() == name.lower():
                    exact_match = p
                    break

            if exact_match:
                project_id = exact_match['id']
                message = f'기존 프로젝트 "{name}"에 노트가 추가되었습니다.'
            elif similar_projects and not force:
                conn.close()
                return jsonify({
                    'success': False,
                    'error': 'duplicate_found',
                    'message': f'"{name}"과(와) 비슷한 이름의 프로젝트가 있습니다. 동일 프로젝트인지 확인해주세요.',
                    'duplicates': similar_projects,
                    'parsed_data': {
                        'name': name,
                        'description': description,
                        'priority': priority,
                        'start_date': start_date,
                        'end_date': end_date,
                        'note_date': note_date,
                        'note_content': note_content
                    }
                }), 409
            else:
                # 새 프로젝트 생성
                is_new = True
                if USE_POSTGRES:
                    cursor.execute('''
                        INSERT INTO projects (name, description, priority, start_date, end_date)
                        VALUES (%s, %s, %s, %s, %s)
                        RETURNING id
                    ''', (name, description, priority, start_date, end_date))
                    project_id = cursor.fetchone()['id']
                else:
                    cursor.execute('''
                        INSERT INTO projects (name, description, priority, start_date, end_date)
                        VALUES (?, ?, ?, ?, ?)
                    ''', (name, description, priority, start_date, end_date))
                    project_id = cursor.lastrowid

                message = f'새 프로젝트 "{name}"이(가) 생성되었습니다.'

                # 날짜가 있으면 캘린더 이벤트 자동 생성
                if start_date:
                    try:
                        if USE_POSTGRES:
                            cursor.execute('''
                                INSERT INTO events (title, start_time, end_time, category, source, sync_status)
                                VALUES (%s, %s, %s, %s, %s, %s)
                            ''', (f'📋 {name} 시작', f'{start_date}T09:00:00', f'{start_date}T10:00:00', '프로젝트', 'web', 'pending_to_mac'))
                        else:
                            cursor.execute('''
                                INSERT INTO events (title, start_time, end_time, category, source, sync_status)
                                VALUES (?, ?, ?, ?, ?, ?)
                            ''', (f'📋 {name} 시작', f'{start_date}T09:00:00', f'{start_date}T10:00:00', '프로젝트', 'web', 'pending_to_mac'))
                        events_created.append('시작일')
                    except Exception as e:
                        print(f"[QUICK-ADD-PROJECT] 시작일 이벤트 생성 실패: {e}")

                if end_date:
                    try:
                        if USE_POSTGRES:
                            cursor.execute('''
                                INSERT INTO events (title, start_time, end_time, category, source, sync_status)
                                VALUES (%s, %s, %s, %s, %s, %s)
                            ''', (f'🏁 {name} 마감', f'{end_date}T09:00:00', f'{end_date}T10:00:00', '프로젝트', 'web', 'pending_to_mac'))
                        else:
                            cursor.execute('''
                                INSERT INTO events (title, start_time, end_time, category, source, sync_status)
                                VALUES (?, ?, ?, ?, ?, ?)
                            ''', (f'🏁 {name} 마감', f'{end_date}T09:00:00', f'{end_date}T10:00:00', '프로젝트', 'web', 'pending_to_mac'))
                        events_created.append('마감일')
                    except Exception as e:
                        print(f"[QUICK-ADD-PROJECT] 마감일 이벤트 생성 실패: {e}")

                if events_created:
                    message += f' 캘린더에 {", ".join(events_created)} 이벤트가 추가되었습니다.'

        # 노트 추가
        if note_content:
            if USE_POSTGRES:
                cursor.execute('''
                    INSERT INTO project_notes (project_id, note_date, content)
                    VALUES (%s, %s, %s)
                ''', (project_id, note_date, note_content))
                cursor.execute('UPDATE projects SET updated_at = CURRENT_TIMESTAMP WHERE id = %s', (project_id,))
            else:
                cursor.execute('''
                    INSERT INTO project_notes (project_id, note_date, content)
                    VALUES (?, ?, ?)
                ''', (project_id, note_date, note_content))
                cursor.execute('UPDATE projects SET updated_at = CURRENT_TIMESTAMP WHERE id = ?', (project_id,))

            if not is_new:
                message += ' 노트가 추가되었습니다.'

        conn.commit()
        conn.close()

        return jsonify({
            'success': True,
            'project_id': project_id,
            'message': message,
            'is_new': is_new,
            'events_created': events_created
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ===== Google Calendar API 연동 =====

# Google Calendar 토큰 저장소 (메모리 - 실제 운영에서는 DB나 파일로 저장 권장)
_gcal_credentials = {}

def get_gcal_credentials():
    """저장된 Google Calendar 인증 정보 반환"""
    global _gcal_credentials
    return _gcal_credentials.get('credentials')

def save_gcal_credentials(credentials):
    """Google Calendar 인증 정보 저장"""
    global _gcal_credentials
    _gcal_credentials['credentials'] = credentials


@assistant_bp.route('/assistant/api/gcal/auth-status', methods=['GET'])
def gcal_auth_status():
    """Google Calendar 인증 상태 확인"""
    try:
        creds = get_gcal_credentials()
        if creds:
            return jsonify({
                'success': True,
                'authenticated': True,
                'message': 'Google Calendar 연동됨'
            })
        return jsonify({
            'success': True,
            'authenticated': False,
            'message': '인증 필요'
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@assistant_bp.route('/assistant/api/gcal/auth', methods=['GET'])
def gcal_auth():
    """Google Calendar OAuth 인증 시작"""
    try:
        from google_auth_oauthlib.flow import Flow

        # 환경변수에서 OAuth 클라이언트 정보 가져오기
        client_id = os.getenv('GOOGLE_CLIENT_ID')
        client_secret = os.getenv('GOOGLE_CLIENT_SECRET')

        if not client_id or not client_secret:
            return jsonify({
                'success': False,
                'error': 'GOOGLE_CLIENT_ID와 GOOGLE_CLIENT_SECRET 환경변수를 설정해주세요.'
            }), 400

        # OAuth 클라이언트 설정
        client_config = {
            "web": {
                "client_id": client_id,
                "client_secret": client_secret,
                "auth_uri": "https://accounts.google.com/o/oauth2/auth",
                "token_uri": "https://oauth2.googleapis.com/token",
                "redirect_uris": [request.host_url.rstrip('/') + '/assistant/api/gcal/callback']
            }
        }

        # Calendar API 스코프
        scopes = [
            'https://www.googleapis.com/auth/calendar',
            'https://www.googleapis.com/auth/calendar.events'
        ]

        flow = Flow.from_client_config(client_config, scopes=scopes)
        flow.redirect_uri = request.host_url.rstrip('/') + '/assistant/api/gcal/callback'

        authorization_url, state = flow.authorization_url(
            access_type='offline',
            include_granted_scopes='true',
            prompt='consent'
        )

        # state 저장 (간단히 전역 변수 사용, 실제로는 세션 등 사용)
        _gcal_credentials['flow_state'] = state
        _gcal_credentials['flow'] = flow

        return jsonify({
            'success': True,
            'auth_url': authorization_url
        })

    except ImportError:
        return jsonify({
            'success': False,
            'error': 'Google 인증 라이브러리가 필요합니다. pip install google-auth-oauthlib google-api-python-client'
        }), 500
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@assistant_bp.route('/assistant/api/gcal/callback', methods=['GET'])
def gcal_callback():
    """Google Calendar OAuth 콜백"""
    try:
        from google_auth_oauthlib.flow import Flow

        # 환경변수에서 OAuth 클라이언트 정보 가져오기
        client_id = os.getenv('GOOGLE_CLIENT_ID')
        client_secret = os.getenv('GOOGLE_CLIENT_SECRET')

        if not client_id or not client_secret:
            return '<script>alert("Google OAuth 환경변수가 설정되지 않았습니다."); window.close();</script>'

        # Flow 재생성 (서버 재시작 대응)
        client_config = {
            "web": {
                "client_id": client_id,
                "client_secret": client_secret,
                "auth_uri": "https://accounts.google.com/o/oauth2/auth",
                "token_uri": "https://oauth2.googleapis.com/token",
                "redirect_uris": [request.host_url.rstrip('/') + '/assistant/api/gcal/callback']
            }
        }

        scopes = [
            'https://www.googleapis.com/auth/calendar',
            'https://www.googleapis.com/auth/calendar.events'
        ]

        flow = Flow.from_client_config(client_config, scopes=scopes)
        flow.redirect_uri = request.host_url.rstrip('/') + '/assistant/api/gcal/callback'

        # authorization response에서 토큰 가져오기
        flow.fetch_token(authorization_response=request.url)
        credentials = flow.credentials

        # 인증 정보 저장
        save_gcal_credentials({
            'token': credentials.token,
            'refresh_token': credentials.refresh_token,
            'token_uri': credentials.token_uri,
            'client_id': credentials.client_id,
            'client_secret': credentials.client_secret,
            'scopes': list(credentials.scopes) if credentials.scopes else scopes
        })

        # 성공 메시지와 함께 창 닫기
        return '''
        <html>
        <head><title>Google Calendar 연동 완료</title></head>
        <body style="font-family: sans-serif; text-align: center; padding: 50px;">
            <h2>✅ Google Calendar 연동 완료!</h2>
            <p>이 창을 닫고 대시보드로 돌아가세요.</p>
            <script>
                setTimeout(function() {
                    if (window.opener) {
                        window.opener.location.reload();
                    }
                    window.close();
                }, 2000);
            </script>
        </body>
        </html>
        '''

    except Exception as e:
        import traceback
        traceback.print_exc()
        return f'<script>alert("인증 오류: {str(e)}"); window.close();</script>'


@assistant_bp.route('/assistant/api/gcal/disconnect', methods=['POST'])
def gcal_disconnect():
    """Google Calendar 연동 해제"""
    global _gcal_credentials
    _gcal_credentials = {}
    return jsonify({'success': True, 'message': 'Google Calendar 연동이 해제되었습니다.'})


def get_gcal_service():
    """Google Calendar API 서비스 객체 반환"""
    from google.oauth2.credentials import Credentials
    from googleapiclient.discovery import build

    creds_data = get_gcal_credentials()
    if not creds_data:
        return None

    credentials = Credentials(
        token=creds_data.get('token'),
        refresh_token=creds_data.get('refresh_token'),
        token_uri=creds_data.get('token_uri'),
        client_id=creds_data.get('client_id'),
        client_secret=creds_data.get('client_secret'),
        scopes=creds_data.get('scopes')
    )

    # 토큰 만료 시 갱신
    if credentials.expired and credentials.refresh_token:
        from google.auth.transport.requests import Request
        credentials.refresh(Request())
        # 갱신된 토큰 저장
        save_gcal_credentials({
            'token': credentials.token,
            'refresh_token': credentials.refresh_token,
            'token_uri': credentials.token_uri,
            'client_id': credentials.client_id,
            'client_secret': credentials.client_secret,
            'scopes': credentials.scopes
        })

    return build('calendar', 'v3', credentials=credentials)


@assistant_bp.route('/assistant/api/gcal/events', methods=['GET'])
def gcal_get_events():
    """Google Calendar에서 일정 가져오기"""
    try:
        service = get_gcal_service()
        if not service:
            return jsonify({'success': False, 'error': 'Google Calendar 인증이 필요합니다.'}), 401

        # 기간 설정 (기본: 오늘부터 30일)
        days = request.args.get('days', 30, type=int)
        time_min = datetime.utcnow().isoformat() + 'Z'
        time_max = (datetime.utcnow() + timedelta(days=days)).isoformat() + 'Z'

        # 캘린더 목록 가져오기
        calendar_id = request.args.get('calendar_id', 'primary')

        events_result = service.events().list(
            calendarId=calendar_id,
            timeMin=time_min,
            timeMax=time_max,
            maxResults=100,
            singleEvents=True,
            orderBy='startTime'
        ).execute()

        events = events_result.get('items', [])

        # 이벤트 정리
        formatted_events = []
        for event in events:
            start = event['start'].get('dateTime', event['start'].get('date'))
            end = event['end'].get('dateTime', event['end'].get('date'))

            formatted_events.append({
                'id': event['id'],
                'title': event.get('summary', '(제목 없음)'),
                'start_time': start,
                'end_time': end,
                'location': event.get('location'),
                'description': event.get('description'),
                'source': 'google_calendar'
            })

        return jsonify({
            'success': True,
            'events': formatted_events,
            'count': len(formatted_events)
        })

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@assistant_bp.route('/assistant/api/gcal/sync', methods=['POST'])
def gcal_sync():
    """로컬 일정을 Google Calendar와 동기화"""
    try:
        service = get_gcal_service()
        if not service:
            return jsonify({'success': False, 'error': 'Google Calendar 인증이 필요합니다.'}), 401

        data = request.get_json() or {}
        direction = data.get('direction', 'both')  # 'to_google', 'from_google', 'both'
        calendar_id = data.get('calendar_id', 'primary')

        synced_to_google = 0
        synced_from_google = 0

        conn = get_db_connection()
        cursor = conn.cursor()

        # === 1. 로컬 → Google Calendar (pending_to_mac 상태인 이벤트) ===
        if direction in ['to_google', 'both']:
            if USE_POSTGRES:
                cursor.execute("SELECT * FROM events WHERE sync_status = 'pending_to_mac'")
            else:
                cursor.execute("SELECT * FROM events WHERE sync_status = 'pending_to_mac'")

            local_events = cursor.fetchall()

            for event in local_events:
                try:
                    # Google Calendar 이벤트 생성
                    gcal_event = {
                        'summary': event['title'],
                        'start': {},
                        'end': {}
                    }

                    # 시작 시간 설정
                    if event['start_time']:
                        start_dt = event['start_time']
                        if isinstance(start_dt, str):
                            if 'T' in start_dt:
                                gcal_event['start']['dateTime'] = start_dt
                                gcal_event['start']['timeZone'] = 'Asia/Seoul'
                            else:
                                gcal_event['start']['date'] = start_dt
                        else:
                            gcal_event['start']['dateTime'] = start_dt.isoformat()
                            gcal_event['start']['timeZone'] = 'Asia/Seoul'

                    # 종료 시간 설정
                    if event['end_time']:
                        end_dt = event['end_time']
                        if isinstance(end_dt, str):
                            if 'T' in end_dt:
                                gcal_event['end']['dateTime'] = end_dt
                                gcal_event['end']['timeZone'] = 'Asia/Seoul'
                            else:
                                gcal_event['end']['date'] = end_dt
                        else:
                            gcal_event['end']['dateTime'] = end_dt.isoformat()
                            gcal_event['end']['timeZone'] = 'Asia/Seoul'
                    else:
                        # 종료 시간이 없으면 시작 시간 + 1시간
                        gcal_event['end'] = gcal_event['start'].copy()

                    if event.get('location'):
                        gcal_event['location'] = event['location']

                    # Google Calendar에 추가
                    created_event = service.events().insert(
                        calendarId=calendar_id,
                        body=gcal_event
                    ).execute()

                    # 동기화 상태 업데이트
                    if USE_POSTGRES:
                        cursor.execute(
                            "UPDATE events SET sync_status = 'synced', gcal_id = %s WHERE id = %s",
                            (created_event['id'], event['id'])
                        )
                    else:
                        cursor.execute(
                            "UPDATE events SET sync_status = 'synced', gcal_id = ? WHERE id = ?",
                            (created_event['id'], event['id'])
                        )

                    synced_to_google += 1

                except Exception as e:
                    print(f"[GCAL-SYNC] 이벤트 동기화 오류 (ID: {event['id']}): {e}")

        # === 2. Google Calendar → 로컬 ===
        if direction in ['from_google', 'both']:
            # 최근 30일 + 미래 60일 이벤트 가져오기
            time_min = (datetime.utcnow() - timedelta(days=30)).isoformat() + 'Z'
            time_max = (datetime.utcnow() + timedelta(days=60)).isoformat() + 'Z'

            events_result = service.events().list(
                calendarId=calendar_id,
                timeMin=time_min,
                timeMax=time_max,
                maxResults=200,
                singleEvents=True,
                orderBy='startTime'
            ).execute()

            gcal_events = events_result.get('items', [])

            for gcal_event in gcal_events:
                gcal_id = gcal_event['id']

                # 이미 동기화된 이벤트인지 확인
                if USE_POSTGRES:
                    cursor.execute("SELECT id FROM events WHERE gcal_id = %s", (gcal_id,))
                else:
                    cursor.execute("SELECT id FROM events WHERE gcal_id = ?", (gcal_id,))

                existing = cursor.fetchone()

                if not existing:
                    # 새 이벤트 추가
                    title = gcal_event.get('summary', '(제목 없음)')
                    start = gcal_event['start'].get('dateTime', gcal_event['start'].get('date'))
                    end = gcal_event['end'].get('dateTime', gcal_event['end'].get('date'))
                    location = gcal_event.get('location')

                    if USE_POSTGRES:
                        cursor.execute('''
                            INSERT INTO events (title, start_time, end_time, location, category, source, sync_status, gcal_id)
                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                        ''', (title, start, end, location, '기타', 'google_calendar', 'synced', gcal_id))
                    else:
                        cursor.execute('''
                            INSERT INTO events (title, start_time, end_time, location, category, source, sync_status, gcal_id)
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                        ''', (title, start, end, location, '기타', 'google_calendar', 'synced', gcal_id))

                    synced_from_google += 1

        conn.commit()
        conn.close()

        return jsonify({
            'success': True,
            'synced_to_google': synced_to_google,
            'synced_from_google': synced_from_google,
            'message': f'동기화 완료: Google로 {synced_to_google}개, 로컬로 {synced_from_google}개'
        })

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@assistant_bp.route('/assistant/api/gcal/calendars', methods=['GET'])
def gcal_list_calendars():
    """사용 가능한 캘린더 목록 조회"""
    try:
        service = get_gcal_service()
        if not service:
            return jsonify({'success': False, 'error': 'Google Calendar 인증이 필요합니다.'}), 401

        calendar_list = service.calendarList().list().execute()
        calendars = []

        for cal in calendar_list.get('items', []):
            calendars.append({
                'id': cal['id'],
                'summary': cal.get('summary', ''),
                'primary': cal.get('primary', False),
                'backgroundColor': cal.get('backgroundColor')
            })

        return jsonify({
            'success': True,
            'calendars': calendars
        })

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ===== Google Calendar Webhook (실시간 동기화) =====

_gcal_webhook_channel = {}  # 채널 정보 저장

def get_webhook_channel():
    """저장된 Webhook 채널 정보 반환"""
    global _gcal_webhook_channel
    # 메모리에 없으면 DB에서 로드 시도
    if not _gcal_webhook_channel:
        try:
            conn = get_db_connection()
            cursor = conn.cursor()
            if USE_POSTGRES:
                cursor.execute("SELECT key, value FROM settings WHERE key LIKE 'gcal_webhook_%'")
            else:
                cursor.execute("SELECT key, value FROM settings WHERE key LIKE 'gcal_webhook_%'")
            rows = cursor.fetchall()
            conn.close()
            for row in rows:
                key = row['key'].replace('gcal_webhook_', '')
                _gcal_webhook_channel[key] = row['value']
        except:
            pass
    return _gcal_webhook_channel

def save_webhook_channel(channel_data):
    """Webhook 채널 정보 저장"""
    global _gcal_webhook_channel
    _gcal_webhook_channel = channel_data
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        for key, value in channel_data.items():
            db_key = f'gcal_webhook_{key}'
            if USE_POSTGRES:
                cursor.execute('''
                    INSERT INTO settings (key, value) VALUES (%s, %s)
                    ON CONFLICT (key) DO UPDATE SET value = EXCLUDED.value
                ''', (db_key, str(value)))
            else:
                cursor.execute('INSERT OR REPLACE INTO settings (key, value) VALUES (?, ?)', (db_key, str(value)))
        conn.commit()
        conn.close()
    except Exception as e:
        print(f"[GCAL-WEBHOOK] 채널 저장 오류: {e}")


@assistant_bp.route('/assistant/api/gcal/webhook/register', methods=['POST'])
def gcal_webhook_register():
    """Google Calendar Webhook 채널 등록"""
    try:
        import uuid

        service = get_gcal_service()
        if not service:
            return jsonify({'success': False, 'error': 'Google Calendar 인증이 필요합니다.'}), 401

        data = request.get_json() or {}
        calendar_id = data.get('calendar_id', 'primary')

        # 기존 채널이 있으면 먼저 해제
        existing_channel = get_webhook_channel()
        if existing_channel.get('id') and existing_channel.get('resource_id'):
            try:
                service.channels().stop(body={
                    'id': existing_channel['id'],
                    'resourceId': existing_channel['resource_id']
                }).execute()
                print(f"[GCAL-WEBHOOK] 기존 채널 해제: {existing_channel['id']}")
            except Exception as e:
                print(f"[GCAL-WEBHOOK] 기존 채널 해제 실패 (무시): {e}")

        # 새 채널 ID 생성
        channel_id = str(uuid.uuid4())

        # Webhook URL 설정
        webhook_url = request.host_url.rstrip('/') + '/assistant/api/gcal/webhook'

        # 채널 만료 시간 (최대 7일, 권장 1일)
        expiration = int((datetime.utcnow() + timedelta(days=1)).timestamp() * 1000)

        # Watch 요청
        watch_request = {
            'id': channel_id,
            'type': 'web_hook',
            'address': webhook_url,
            'expiration': expiration
        }

        response = service.events().watch(
            calendarId=calendar_id,
            body=watch_request
        ).execute()

        # 채널 정보 저장
        channel_data = {
            'id': response['id'],
            'resource_id': response['resourceId'],
            'resource_uri': response.get('resourceUri', ''),
            'expiration': response['expiration'],
            'calendar_id': calendar_id,
            'created_at': datetime.utcnow().isoformat()
        }
        save_webhook_channel(channel_data)

        # 만료 시간 계산
        expiration_dt = datetime.fromtimestamp(int(response['expiration']) / 1000)

        return jsonify({
            'success': True,
            'channel_id': response['id'],
            'expiration': expiration_dt.isoformat(),
            'webhook_url': webhook_url,
            'message': '실시간 동기화가 활성화되었습니다.'
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@assistant_bp.route('/assistant/api/gcal/webhook', methods=['POST'])
def gcal_webhook_callback():
    """Google Calendar Webhook 콜백 (실시간 알림 수신)"""
    try:
        # Google에서 보내는 헤더 확인
        channel_id = request.headers.get('X-Goog-Channel-ID')
        resource_id = request.headers.get('X-Goog-Resource-ID')
        resource_state = request.headers.get('X-Goog-Resource-State')

        print(f"[GCAL-WEBHOOK] 알림 수신: state={resource_state}, channel={channel_id}")

        # 채널 검증
        saved_channel = get_webhook_channel()
        if saved_channel.get('id') != channel_id:
            print(f"[GCAL-WEBHOOK] 알 수 없는 채널: {channel_id}")
            return '', 200  # Google에게는 항상 200 응답

        # sync 또는 exists 상태일 때 동기화 실행
        if resource_state in ['sync', 'exists']:
            # 비동기로 동기화 실행 (응답 지연 방지)
            import threading
            def do_sync():
                try:
                    _sync_from_google_internal()
                except Exception as e:
                    print(f"[GCAL-WEBHOOK] 동기화 오류: {e}")

            thread = threading.Thread(target=do_sync)
            thread.start()

        return '', 200

    except Exception as e:
        print(f"[GCAL-WEBHOOK] 콜백 오류: {e}")
        return '', 200  # 오류 시에도 200 응답


def _sync_from_google_internal():
    """Google Calendar에서 변경된 이벤트 동기화 (내부 함수)"""
    service = get_gcal_service()
    if not service:
        return

    channel = get_webhook_channel()
    calendar_id = channel.get('calendar_id', 'primary')

    # 최근 30일 + 미래 60일
    time_min = (datetime.utcnow() - timedelta(days=30)).isoformat() + 'Z'
    time_max = (datetime.utcnow() + timedelta(days=60)).isoformat() + 'Z'

    events_result = service.events().list(
        calendarId=calendar_id,
        timeMin=time_min,
        timeMax=time_max,
        maxResults=200,
        singleEvents=True,
        orderBy='startTime'
    ).execute()

    gcal_events = events_result.get('items', [])

    conn = get_db_connection()
    cursor = conn.cursor()

    synced_count = 0

    for gcal_event in gcal_events:
        gcal_id = gcal_event['id']

        # 이미 동기화된 이벤트인지 확인
        if USE_POSTGRES:
            cursor.execute("SELECT id FROM events WHERE gcal_id = %s", (gcal_id,))
        else:
            cursor.execute("SELECT id FROM events WHERE gcal_id = ?", (gcal_id,))

        existing = cursor.fetchone()

        if not existing:
            title = gcal_event.get('summary', '(제목 없음)')
            start = gcal_event['start'].get('dateTime', gcal_event['start'].get('date'))
            end = gcal_event['end'].get('dateTime', gcal_event['end'].get('date'))
            location = gcal_event.get('location')

            if USE_POSTGRES:
                cursor.execute('''
                    INSERT INTO events (title, start_time, end_time, location, category, source, sync_status, gcal_id)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                ''', (title, start, end, location, '기타', 'google_calendar', 'synced', gcal_id))
            else:
                cursor.execute('''
                    INSERT INTO events (title, start_time, end_time, location, category, source, sync_status, gcal_id)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                ''', (title, start, end, location, '기타', 'google_calendar', 'synced', gcal_id))

            synced_count += 1

    conn.commit()
    conn.close()

    print(f"[GCAL-WEBHOOK] 자동 동기화 완료: {synced_count}개 새 이벤트")


@assistant_bp.route('/assistant/api/gcal/webhook/status', methods=['GET'])
def gcal_webhook_status():
    """Webhook 채널 상태 확인"""
    try:
        channel = get_webhook_channel()

        if not channel.get('id'):
            return jsonify({
                'success': True,
                'active': False,
                'message': '실시간 동기화가 비활성화 상태입니다.'
            })

        # 만료 시간 확인
        expiration_ms = int(channel.get('expiration', 0))
        expiration_dt = datetime.fromtimestamp(expiration_ms / 1000)
        is_expired = datetime.utcnow() > expiration_dt

        return jsonify({
            'success': True,
            'active': not is_expired,
            'channel_id': channel.get('id'),
            'expiration': expiration_dt.isoformat(),
            'is_expired': is_expired,
            'calendar_id': channel.get('calendar_id', 'primary')
        })

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@assistant_bp.route('/assistant/api/gcal/webhook/stop', methods=['POST'])
def gcal_webhook_stop():
    """Webhook 채널 해제"""
    try:
        service = get_gcal_service()
        if not service:
            return jsonify({'success': False, 'error': 'Google Calendar 인증이 필요합니다.'}), 401

        channel = get_webhook_channel()

        if not channel.get('id') or not channel.get('resource_id'):
            return jsonify({'success': True, 'message': '활성화된 채널이 없습니다.'})

        # 채널 해제
        service.channels().stop(body={
            'id': channel['id'],
            'resourceId': channel['resource_id']
        }).execute()

        # 저장된 채널 정보 삭제
        save_webhook_channel({})

        return jsonify({
            'success': True,
            'message': '실시간 동기화가 비활성화되었습니다.'
        })

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ===== Google Sheets API 연동 =====

_gsheets_credentials = {}

def get_gsheets_service():
    """Google Sheets API 서비스 객체 반환 (Calendar와 동일한 인증 사용)"""
    from google.oauth2.credentials import Credentials
    from googleapiclient.discovery import build

    # Google Calendar 인증 정보 재사용 (스코프 추가 필요)
    creds_data = get_gcal_credentials()
    if not creds_data:
        return None

    credentials = Credentials(
        token=creds_data.get('token'),
        refresh_token=creds_data.get('refresh_token'),
        token_uri=creds_data.get('token_uri'),
        client_id=creds_data.get('client_id'),
        client_secret=creds_data.get('client_secret'),
        scopes=creds_data.get('scopes')
    )

    if credentials.expired and credentials.refresh_token:
        from google.auth.transport.requests import Request
        credentials.refresh(Request())

    return build('sheets', 'v4', credentials=credentials)


@assistant_bp.route('/assistant/api/gsheets/auth-status', methods=['GET'])
def gsheets_auth_status():
    """Google Sheets 인증 상태 확인"""
    try:
        creds = load_gsheets_credentials()
        if creds:
            return jsonify({'authenticated': True, 'success': True})
        return jsonify({'authenticated': False, 'success': True})
    except Exception as e:
        return jsonify({'authenticated': False, 'success': False, 'error': str(e)})


@assistant_bp.route('/assistant/api/gsheets/auth', methods=['GET'])
def gsheets_auth():
    """Google Sheets OAuth 인증 시작 (Calendar + Sheets 통합)"""
    try:
        from google_auth_oauthlib.flow import Flow

        client_id = os.getenv('GOOGLE_CLIENT_ID')
        client_secret = os.getenv('GOOGLE_CLIENT_SECRET')

        if not client_id or not client_secret:
            return jsonify({
                'success': False,
                'error': 'GOOGLE_CLIENT_ID와 GOOGLE_CLIENT_SECRET 환경변수를 설정해주세요.'
            }), 400

        client_config = {
            "web": {
                "client_id": client_id,
                "client_secret": client_secret,
                "auth_uri": "https://accounts.google.com/o/oauth2/auth",
                "token_uri": "https://oauth2.googleapis.com/token",
                "redirect_uris": [request.host_url.rstrip('/') + '/assistant/api/gsheets/callback']
            }
        }

        # Calendar + Sheets 스코프 통합
        scopes = [
            'https://www.googleapis.com/auth/calendar',
            'https://www.googleapis.com/auth/calendar.events',
            'https://www.googleapis.com/auth/spreadsheets',
            'https://www.googleapis.com/auth/spreadsheets.readonly'
        ]

        flow = Flow.from_client_config(client_config, scopes=scopes)
        flow.redirect_uri = request.host_url.rstrip('/') + '/assistant/api/gsheets/callback'

        authorization_url, state = flow.authorization_url(
            access_type='offline',
            include_granted_scopes='true',
            prompt='consent'
        )

        _gsheets_credentials['flow'] = flow

        return jsonify({
            'success': True,
            'auth_url': authorization_url
        })

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@assistant_bp.route('/assistant/api/gsheets/callback', methods=['GET'])
def gsheets_callback():
    """Google Sheets OAuth 콜백"""
    try:
        from google_auth_oauthlib.flow import Flow

        # 환경변수에서 OAuth 클라이언트 정보 가져오기
        client_id = os.getenv('GOOGLE_CLIENT_ID')
        client_secret = os.getenv('GOOGLE_CLIENT_SECRET')

        if not client_id or not client_secret:
            return '<script>alert("Google OAuth 환경변수가 설정되지 않았습니다."); window.close();</script>'

        # Flow 재생성 (서버 재시작 대응)
        client_config = {
            "web": {
                "client_id": client_id,
                "client_secret": client_secret,
                "auth_uri": "https://accounts.google.com/o/oauth2/auth",
                "token_uri": "https://oauth2.googleapis.com/token",
                "redirect_uris": [request.host_url.rstrip('/') + '/assistant/api/gsheets/callback']
            }
        }

        # Calendar + Sheets 스코프 통합
        scopes = [
            'https://www.googleapis.com/auth/calendar',
            'https://www.googleapis.com/auth/calendar.events',
            'https://www.googleapis.com/auth/spreadsheets',
            'https://www.googleapis.com/auth/spreadsheets.readonly'
        ]

        flow = Flow.from_client_config(client_config, scopes=scopes)
        flow.redirect_uri = request.host_url.rstrip('/') + '/assistant/api/gsheets/callback'

        flow.fetch_token(authorization_response=request.url)
        credentials = flow.credentials

        # Calendar 인증과 통합하여 저장
        save_gcal_credentials({
            'token': credentials.token,
            'refresh_token': credentials.refresh_token,
            'token_uri': credentials.token_uri,
            'client_id': credentials.client_id,
            'client_secret': credentials.client_secret,
            'scopes': list(credentials.scopes) if credentials.scopes else scopes
        })

        return '''
        <html>
        <head><title>Google 연동 완료</title></head>
        <body style="font-family: sans-serif; text-align: center; padding: 50px;">
            <h2>✅ Google Calendar + Sheets 연동 완료!</h2>
            <p>이 창을 닫고 대시보드로 돌아가세요.</p>
            <script>
                setTimeout(function() {
                    if (window.opener) window.opener.location.reload();
                    window.close();
                }, 2000);
            </script>
        </body>
        </html>
        '''
    except Exception as e:
        import traceback
        traceback.print_exc()
        return f'<script>alert("인증 오류: {str(e)}"); window.close();</script>'


@assistant_bp.route('/assistant/api/gsheets/read', methods=['GET'])
def gsheets_read():
    """Google Sheets 데이터 읽기"""
    try:
        service = get_gsheets_service()
        if not service:
            return jsonify({'success': False, 'error': 'Google 인증이 필요합니다.'}), 401

        spreadsheet_id = request.args.get('spreadsheet_id')
        range_name = request.args.get('range', 'Sheet1!A1:Z100')

        if not spreadsheet_id:
            return jsonify({'success': False, 'error': 'spreadsheet_id가 필요합니다.'}), 400

        result = service.spreadsheets().values().get(
            spreadsheetId=spreadsheet_id,
            range=range_name
        ).execute()

        values = result.get('values', [])

        return jsonify({
            'success': True,
            'data': values,
            'rows': len(values)
        })

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@assistant_bp.route('/assistant/api/gsheets/write', methods=['POST'])
def gsheets_write():
    """Google Sheets에 데이터 쓰기"""
    try:
        service = get_gsheets_service()
        if not service:
            return jsonify({'success': False, 'error': 'Google 인증이 필요합니다.'}), 401

        data = request.get_json()
        spreadsheet_id = data.get('spreadsheet_id')
        range_name = data.get('range', 'Sheet1!A1')
        values = data.get('values', [])

        if not spreadsheet_id:
            return jsonify({'success': False, 'error': 'spreadsheet_id가 필요합니다.'}), 400

        body = {'values': values}

        result = service.spreadsheets().values().update(
            spreadsheetId=spreadsheet_id,
            range=range_name,
            valueInputOption='USER_ENTERED',
            body=body
        ).execute()

        return jsonify({
            'success': True,
            'updated_cells': result.get('updatedCells', 0),
            'updated_range': result.get('updatedRange', '')
        })

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@assistant_bp.route('/assistant/api/gsheets/append', methods=['POST'])
def gsheets_append():
    """Google Sheets에 행 추가"""
    try:
        service = get_gsheets_service()
        if not service:
            return jsonify({'success': False, 'error': 'Google 인증이 필요합니다.'}), 401

        data = request.get_json()
        spreadsheet_id = data.get('spreadsheet_id')
        range_name = data.get('range', 'Sheet1')
        values = data.get('values', [])

        if not spreadsheet_id:
            return jsonify({'success': False, 'error': 'spreadsheet_id가 필요합니다.'}), 400

        body = {'values': values}

        result = service.spreadsheets().values().append(
            spreadsheetId=spreadsheet_id,
            range=range_name,
            valueInputOption='USER_ENTERED',
            insertDataOption='INSERT_ROWS',
            body=body
        ).execute()

        return jsonify({
            'success': True,
            'updated_range': result.get('updates', {}).get('updatedRange', ''),
            'updated_rows': result.get('updates', {}).get('updatedRows', 0)
        })

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@assistant_bp.route('/assistant/api/gsheets/export-people', methods=['POST'])
def gsheets_export_people():
    """인물 데이터를 Google Sheets로 내보내기"""
    try:
        service = get_gsheets_service()
        if not service:
            return jsonify({'success': False, 'error': 'Google 인증이 필요합니다.'}), 401

        data = request.get_json() or {}
        spreadsheet_id = data.get('spreadsheet_id')

        if not spreadsheet_id:
            return jsonify({'success': False, 'error': 'spreadsheet_id가 필요합니다.'}), 400

        # DB에서 인물 데이터 가져오기
        conn = get_db_connection()
        cursor = conn.cursor()

        if USE_POSTGRES:
            cursor.execute('SELECT name, role, category, notes FROM people ORDER BY name')
        else:
            cursor.execute('SELECT name, role, category, notes FROM people ORDER BY name')

        people = cursor.fetchall()
        conn.close()

        # 헤더 + 데이터 준비
        values = [['이름', '직분/역할', '카테고리', '비고']]
        for person in people:
            values.append([
                person['name'] or '',
                person['role'] or '',
                person['category'] or '',
                person['notes'] or ''
            ])

        # Sheets에 쓰기
        body = {'values': values}
        result = service.spreadsheets().values().update(
            spreadsheetId=spreadsheet_id,
            range='인물!A1',
            valueInputOption='USER_ENTERED',
            body=body
        ).execute()

        return jsonify({
            'success': True,
            'exported_count': len(people),
            'message': f'{len(people)}명의 인물 데이터를 내보냈습니다.'
        })

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@assistant_bp.route('/assistant/api/gsheets/export-events', methods=['POST'])
def gsheets_export_events():
    """일정 데이터를 Google Sheets로 내보내기"""
    try:
        service = get_gsheets_service()
        if not service:
            return jsonify({'success': False, 'error': 'Google 인증이 필요합니다.'}), 401

        data = request.get_json() or {}
        spreadsheet_id = data.get('spreadsheet_id')

        if not spreadsheet_id:
            return jsonify({'success': False, 'error': 'spreadsheet_id가 필요합니다.'}), 400

        conn = get_db_connection()
        cursor = conn.cursor()

        if USE_POSTGRES:
            cursor.execute('SELECT title, start_time, end_time, category, location FROM events ORDER BY start_time')
        else:
            cursor.execute('SELECT title, start_time, end_time, category, location FROM events ORDER BY start_time')

        events = cursor.fetchall()
        conn.close()

        values = [['제목', '시작', '종료', '카테고리', '장소']]
        for event in events:
            values.append([
                event['title'] or '',
                str(event['start_time'] or ''),
                str(event['end_time'] or ''),
                event['category'] or '',
                event['location'] or ''
            ])

        body = {'values': values}
        result = service.spreadsheets().values().update(
            spreadsheetId=spreadsheet_id,
            range='일정!A1',
            valueInputOption='USER_ENTERED',
            body=body
        ).execute()

        return jsonify({
            'success': True,
            'exported_count': len(events),
            'message': f'{len(events)}개의 일정을 내보냈습니다.'
        })

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# 모듈 로드 시 DB 초기화
try:
    init_assistant_db()
except Exception as e:
    print(f"[ASSISTANT-DB] 초기화 오류: {e}")
